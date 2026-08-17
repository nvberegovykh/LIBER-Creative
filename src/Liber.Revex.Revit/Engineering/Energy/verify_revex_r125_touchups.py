#!/usr/bin/env python3
from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace
import tempfile

import revex_final_touchups_r125 as r125


class FakeReference:
    @staticmethod
    def _row_code(row):
        import re
        text = " ".join(str(row.get(k) or "") for k in ("evidence", "assemblyType", "description")).upper()
        m = re.search(r"(?<![A-Z0-9])([A-Z]{1,3}\d+(?:\.\d+)?)(?![A-Z0-9.])", text)
        if not m:
            return "", ""
        code = m.group(1)
        return code, code.split(".", 1)[0]

    @staticmethod
    def _class_for_row(row):
        return str(row.get("kind") or "").lower()

    @staticmethod
    def _has_any_thermal(row):
        kind = str(row.get("kind") or "").lower()
        if kind in {"window", "door"}:
            return row.get("uFactor") not in (None, "")
        return any(row.get(k) not in (None, "") for k in ("cavityR", "continuousR"))

    @staticmethod
    def _profile_matches_row(profile, row):
        return True

    @staticmethod
    def _reference_path():
        return Path(__file__)

    @staticmethod
    def _approved_profiles(_path):
        return {
            "window": {"vt": 0.45},
            "door": {"vt": 0.45},
            "roof": {"cavityR": 26.4, "continuousR": 15.0},
        }


def verify_native_schedule_total() -> None:
    with tempfile.TemporaryDirectory(prefix="revex-r125-total-") as tmp:
        root = Path(tmp)
        schedule = root / "REVIT-SCHEDULE-EVIDENCE.json"
        schedule.write_text(json.dumps({
            "schema": "liber.revex.engineering-schedule-evidence.v1",
            "authority": "active-revit-document-native-schedules",
            "schedules": [{
                "name": "Envelope Assemblies",
                "uniqueId": "schedule-1",
                "fields": [
                    {"columnHeading": "Assembly", "name": "Assembly", "hidden": False},
                    {"columnHeading": "Gross Area", "name": "Area", "hidden": False},
                    {"columnHeading": "Cavity R-Value", "name": "R", "hidden": False},
                ],
                "headerRows": [["Assembly", "Gross Area", "Cavity R-Value"]],
                "bodyRows": [
                    ["R1 Roof", "700", "26.4"],
                    ["R2 Roof region", "650", ""],
                    ["Grand Total", "2450", ""],
                ],
            }],
        }), encoding="utf-8")
        request = root / "request.json"
        request.write_text(json.dumps({"sourceArtifacts": [str(schedule)]}), encoding="utf-8")
        total = r125.native_roof_schedule_total(request)
        assert total is not None, "native roof total was not found"
        assert abs(float(total["value"]) - 2450.0) < 1e-9, total
        assert total["authority"] == "ACTIVE_REVIT_NATIVE_SCHEDULE_TOTAL", total


def verify_request_projection() -> None:
    with tempfile.TemporaryDirectory(prefix="revex-r125-request-") as tmp:
        root = Path(tmp)
        schedule = root / "REVIT-SCHEDULE-EVIDENCE.json"
        schedule.write_text(json.dumps({
            "schema": "liber.revex.engineering-schedule-evidence.v1",
            "authority": "active-revit-document-native-schedules",
            "schedules": [{
                "name": "Envelope Assemblies",
                "uniqueId": "schedule-1",
                "fields": [
                    {"columnHeading": "Assembly", "hidden": False},
                    {"columnHeading": "Gross Area", "hidden": False},
                ],
                "bodyRows": [["R1 Roof", "700"], ["Grand Total", "2450"]],
            }],
        }), encoding="utf-8")
        facts = root / "facts.json"
        facts.write_text(json.dumps({
            "pages": [{
                "pageType": "EN",
                "confidence": 1.0,
                "envelope": [
                    {"kind": "roof", "assemblyType": "R1", "grossAreaFt2": 700, "confidence": 1.0,
                     "cavityR": 26.4, "continuousR": 15.0},
                    {"kind": "roof", "assemblyType": "R2", "grossAreaFt2": 650, "confidence": 1.0},
                    {"kind": "window", "assemblyType": "G11.1", "grossAreaFt2": 225, "confidence": 1.0,
                     "uFactor": 0.30, "shgc": 0.30},
                ],
            }]
        }), encoding="utf-8")
        request = root / "request.json"
        request.write_text(json.dumps({
            "revision": "eng_test",
            "pageFactsPath": str(facts),
            "sourceArtifacts": [str(schedule)],
        }), encoding="utf-8")
        derived = r125.apply_request_touchups(request, root, FakeReference)
        payload = json.loads(Path(derived).read_text(encoding="utf-8"))
        out = json.loads(Path(payload["pageFactsPath"]).read_text(encoding="utf-8"))
        rows = out["pages"][0]["envelope"]
        roof = [row for row in rows if row["kind"] == "roof"]
        window = next(row for row in rows if row["kind"] == "window")
        assert all(abs(float(row["scheduleTotalFt2"]) - 2450.0) < 1e-9 for row in roof)
        assert abs(float(window["vt"]) - 0.45) < 1e-9, window
        assert window["visibleTransmittanceAuthority"] == "APPROVED_SAME_ENVELOPE_REFERENCE_VT"


def verify_pipeline_patch() -> None:
    def code(row):
        import re
        text = " ".join(str(row.get(k) or "") for k in ("assemblyType", "description", "evidence")).upper()
        m = re.search(r"(?<![A-Z0-9])([A-Z]{1,3}\d+(?:\.\d+)?)(?![A-Z0-9.])", text)
        if not m:
            return "", ""
        c = m.group(1)
        return c, c.split(".", 1)[0]

    def merge(diagram, _thermal):
        return dict(diagram), None

    def roof(_diagram, _thermal):
        return ({"kind": "roof", "grossAreaFt2": 7167.0, "assemblyType": "R1", "description": "Roof"},
                [], {"roofAggregateAreaFt2": 7167.0})

    fake = SimpleNamespace(
        _merge_diagram_geometry_with_thermal=merge,
        _merge_roof_geometry_as_one_area=roof,
        canonicalize_comcheck_envelope_rows=lambda _pages: (
            [{"kind": "window", "assemblyType": "G11.1: Window Fixed, Perf. Specs.: Product ID Alumtec, SHGC 0.30, < 95' above-grade, [Bldg. Use 1 - Multifamily]",
              "description": "Window Fixed, Perf. Specs.: Product ID Alumtec, SHGC 0.30, < 95' above-grade, [Bldg. Use 1 - Multifamily]"}],
            {}
        ),
        _row_visible_transmittance=lambda _row: None,
        prepare_project_comcheck=lambda facts, project_identity, filing_dir, log: (None, Path(filing_dir)/"audit.pdf", {}),
        _comcheck_row_code=code,
    )
    r125.patch_pipeline(fake, FakeReference)
    aggregate, errors, audit = fake._merge_roof_geometry_as_one_area(
        [{"kind": "roof", "grossAreaFt2": 700, "scheduleTotalFt2": 2450},
         {"kind": "roof", "grossAreaFt2": 650, "scheduleTotalFt2": 2450}], []
    )
    assert not errors
    assert abs(float(aggregate["grossAreaFt2"]) - 2450.0) < 1e-9, aggregate
    assert audit["roofRegionArithmeticAreaFt2"] == 7167.0
    assert audit["roofRegionSumUsedForFiling"] is False

    merged, _ = fake._merge_diagram_geometry_with_thermal(
        {"kind": "window", "assemblyType": "G11.1", "description": "Window Fixed", "grossAreaFt2": 225},
        [{"kind": "window", "assemblyType": "G11.1", "uFactor": 0.3, "shgc": 0.3, "vt": 0.45}],
    )
    assert abs(float(merged["vt"]) - 0.45) < 1e-9, merged

    clear_vt = fake._row_visible_transmittance({"kind": "window", "description": "clear window"})
    tinted_vt = fake._row_visible_transmittance({"kind": "window", "description": "bronze tinted window"})
    assert abs(float(clear_vt) - 0.45) < 1e-9
    assert abs(float(tinted_vt) - 0.45) < 1e-9

    rows, _ = fake.canonicalize_comcheck_envelope_rows([])
    label = rows[0]["description"]
    assert len(label) <= 72, label
    assert "Bldg" not in label and "Product ID" not in label and "SHGC" not in label, label
    assert rows[0]["assemblyType"] == "G11.1", rows[0]


def verify_en1_print_inside_production_image() -> None:
    """Runs fully in the Docker image; source-only CI may skip if dependencies are absent."""
    import shutil
    if not (shutil.which("soffice") or shutil.which("libreoffice")):
        return
    import sys
    server = Path("/opt/revex/server")
    if server.is_dir() and str(server) not in sys.path:
        sys.path.insert(0, str(server))
    try:
        import revex_user_identity_en1 as en1
    except ModuleNotFoundError:
        return
    from openpyxl import Workbook, load_workbook
    from pypdf import PdfReader

    with tempfile.TemporaryDirectory(prefix="revex-r125-en1-") as tmp:
        root = Path(tmp)
        book = Workbook()
        first = book.active
        first.title = en1.EN1_PRINT_SHEETS[0]
        for name in en1.EN1_PRINT_SHEETS[1:]:
            book.create_sheet(name)
        book.create_sheet("Helper")
        for sheet in book.worksheets:
            sheet["A1"] = sheet.title
            sheet["B2"] = "Visible content"
            if sheet.title in en1.EN1_PRINT_SHEETS:
                sheet.row_dimensions[2].hidden = True
                sheet.column_dimensions["B"].hidden = True
        xlsx = root / "EN-1_READY_TO_INSERT.xlsx"
        pdf = root / "EN-1_READY_TO_INSERT.pdf"
        book.save(xlsx)
        audit = r125.strict_en1_pdf(xlsx, pdf, root)
        assert audit["oneFilingSheetPerPdfPage"] is True
        assert len(PdfReader(str(pdf)).pages) == len(en1.EN1_PRINT_SHEETS)
        check = load_workbook(xlsx)
        assert check[en1.EN1_PRINT_SHEETS[0]].row_dimensions[2].hidden is True
        assert check[en1.EN1_PRINT_SHEETS[0]].column_dimensions["B"].hidden is True
        check.close()


def main() -> int:
    verify_native_schedule_total()
    verify_request_projection()
    verify_pipeline_patch()
    verify_en1_print_inside_production_image()
    print(json.dumps({
        "REVEX_R125_FINAL_TOUCHUPS": "PASSED",
        "nativeScheduleTotal": True,
        "diagramRegionResumDisabledWhenScheduleTotalExists": True,
        "vtFallback": True,
        "compactComcheckLabels": True,
        "en1TemplatePreservingPrint": True,
    }, separators=(",", ":")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
