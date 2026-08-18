#!/usr/bin/env python3
"""REVEX r125 final Energy touch-ups.

Narrow reusable fixes applied after the r124 resumable-publication release:
- native Revit schedule totals outrank arithmetic re-summing of drawing regions;
- VT/VLT never stays blank when current evidence is missing it;
- COMcheck labels are compact so the official report remains readable;
- EN-1 PDF uses the original workbook appearance with visibility/fit-only print changes;
- r124 keeps using byte-verified prior simulation outputs and publishes the normal release ZIP.

Immutable Engineering evidence is never edited.
"""
from __future__ import annotations

import copy
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile
import xml.etree.ElementTree as ET
from typing import Any

VERSION = "20260817r127-fixed-vt0451"
MISSING_VT = 0.45
VT_CLEAR_FALLBACK = MISSING_VT
VT_TINTED_FALLBACK = MISSING_VT
TOTAL_MARKER = re.compile(r"\b(?:grand\s+total|total(?:s)?|sum)\b", re.I)
ROOF_TOKEN = re.compile(r"(?<![A-Z0-9])R\d+(?:\.\d+)?(?![A-Z0-9.])", re.I)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _text(value).lower()).strip()


def _number(value: Any) -> float | None:
    text = _text(value).replace(",", "")
    match = re.search(r"(?<![A-Za-z0-9])(-?\d+(?:\.\d+)?)", text)
    if not match:
        return None
    try:
        number = float(match.group(1))
    except ValueError:
        return None
    return number if number > 0 else None


def _schedule_path(request: dict) -> Path | None:
    for raw in list(request.get("sourceArtifacts") or []):
        path = Path(_text(raw))
        if path.is_file() and path.name.lower().endswith("revit-schedule-evidence.json"):
            return path
    return None


def _headings(schedule: dict) -> list[str]:
    fields = [row for row in list(schedule.get("fields") or []) if not bool(row.get("hidden"))]
    field_headings = [_text(row.get("columnHeading") or row.get("name")) for row in fields]
    headers = [[_text(v) for v in list(row or [])] for row in list(schedule.get("headerRows") or [])]
    body = [list(row or []) for row in list(schedule.get("bodyRows") or [])]
    width = max([len(field_headings), *(len(row) for row in headers), *(len(row) for row in body), 0])
    out = []
    for col in range(width):
        tokens: list[str] = []
        for row in headers:
            value = row[col] if col < len(row) else ""
            if value and _norm(value) not in {_norm(t) for t in tokens}:
                tokens.append(value)
        if col < len(field_headings):
            value = field_headings[col]
            if value and _norm(value) not in {_norm(t) for t in tokens}:
                tokens.append(value)
        out.append(" / ".join(tokens))
    return out


def native_roof_schedule_total(request_path: Path) -> dict | None:
    """Read the native Revit schedule's own total instead of summing diagram regions.

    Existing r116 Engineering revisions already contain bodyRows, so this works without
    a new Revit sync when the total row has a visible Total/Grand Total label. Newer
    exporter metadata (ShowGrandTotal / field DisplayType) is also consumed when present.
    """
    request = json.loads(Path(request_path).read_text(encoding="utf-8"))
    schedule_path = _schedule_path(request)
    if schedule_path is None:
        return None
    try:
        evidence = json.loads(schedule_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return None

    candidates: list[dict] = []
    for schedule in list(evidence.get("schedules") or []):
        headings = _headings(schedule)
        area_cols = [
            i for i, h in enumerate(headings)
            if "area" in _norm(h)
            and not any(t in _norm(h) for t in ("u factor", "r value", "perimeter", "ratio", "deduction"))
        ]
        if not area_cols:
            continue
        rows = [[_text(v) for v in list(raw or [])] for raw in list(schedule.get("bodyRows") or [])]
        schedule_text = " ".join(
            [_text(schedule.get("name")), *headings, *(cell for row in rows for cell in row[:3])]
        )
        roof_context = "roof" in _norm(schedule_text) or any(ROOF_TOKEN.search(" ".join(row)) for row in rows)
        if not roof_context:
            continue

        definition = dict(schedule.get("definition") or {})
        grand_total_enabled = _norm(definition.get("showGrandTotal")) in {"true", "1", "yes"}
        grand_total_title = _text(definition.get("grandTotalTitle"))

        for row_index, row in enumerate(rows):
            row_text = " ".join(row)
            explicit = bool(TOTAL_MARKER.search(row_text))
            if grand_total_title and grand_total_title.casefold() in row_text.casefold():
                explicit = True
            # When Revit suppresses the grand-total title, the final body row is still
            # the native total if the schedule definition explicitly says ShowGrandTotal.
            implicit_native_total = grand_total_enabled and row_index == len(rows) - 1
            if not explicit and not implicit_native_total:
                continue
            # Do not accept a component detail row that merely contains a word like
            # "total" inside a descriptive note.
            if ROOF_TOKEN.search(row_text) and not explicit:
                continue
            for col in area_cols:
                if col >= len(row):
                    continue
                value = _number(row[col])
                if value is None:
                    continue
                candidates.append({
                    "value": round(value, 6),
                    "scheduleName": _text(schedule.get("name")),
                    "scheduleUniqueId": _text(schedule.get("uniqueId")),
                    "rowIndex": row_index,
                    "columnIndex": col,
                    "columnHeading": headings[col] if col < len(headings) else "",
                    "cellText": row[col],
                    "rowText": row_text,
                    "authority": "ACTIVE_REVIT_NATIVE_SCHEDULE_TOTAL",
                    "explicitTotalMarker": explicit,
                    "showGrandTotal": grand_total_enabled,
                    "scheduleArtifact": schedule_path.name,
                })

        # Backward-compatible current-revision fallback: a single roof schedule row
        # carrying both a roof code and an area is more authoritative than re-summing
        # repeated diagram occurrences, but only when the schedule contains no competing
        # roof-area values.
        if not any(c["scheduleUniqueId"] == _text(schedule.get("uniqueId")) for c in candidates):
            roof_area_values: list[tuple[int, int, float, str]] = []
            for row_index, row in enumerate(rows):
                row_text = " ".join(row)
                if not ROOF_TOKEN.search(row_text) and "roof" not in _norm(row_text):
                    continue
                for col in area_cols:
                    if col < len(row):
                        value = _number(row[col])
                        if value is not None:
                            roof_area_values.append((row_index, col, round(value, 6), row_text))
            distinct = sorted({value for _, _, value, _ in roof_area_values})
            if len(distinct) == 1 and roof_area_values:
                row_index, col, value, row_text = roof_area_values[0]
                candidates.append({
                    "value": value,
                    "scheduleName": _text(schedule.get("name")),
                    "scheduleUniqueId": _text(schedule.get("uniqueId")),
                    "rowIndex": row_index,
                    "columnIndex": col,
                    "columnHeading": headings[col] if col < len(headings) else "",
                    "cellText": rows[row_index][col],
                    "rowText": row_text,
                    "authority": "ACTIVE_REVIT_NATIVE_SCHEDULE_SINGLE_ROOF_AREA",
                    "explicitTotalMarker": False,
                    "showGrandTotal": grand_total_enabled,
                    "scheduleArtifact": schedule_path.name,
                })

    if not candidates:
        return None
    explicit = [c for c in candidates if c["explicitTotalMarker"]]
    pool = explicit or candidates
    distinct = sorted({round(float(c["value"]), 3) for c in pool})
    if len(distinct) != 1:
        return None
    chosen = dict(pool[0])
    chosen["candidateCount"] = len(pool)
    chosen["corroboratingCandidates"] = pool
    return chosen


def _row_has_vt(row: dict) -> bool:
    for key in ("vt", "vlt", "visibleTransmittance"):
        if row.get(key) not in (None, ""):
            try:
                float(row[key])
                return True
            except (TypeError, ValueError):
                pass
    text = " ".join(_text(row.get(k)) for k in ("evidence", "description", "assemblyType"))
    return bool(re.search(r"\bV(?:T|LT)\s*[:=]?\s*(?:0?(?:\.\d+)|1(?:\.0+)?)", text, re.I))


def _tinted(row: dict) -> bool:
    text = _norm(" ".join(_text(row.get(k)) for k in ("description", "assemblyType", "product", "evidence")))
    return any(token in text for token in ("tint", "tinted", "bronze", "gray glass", "grey glass", "reflective", "frit"))


def apply_request_touchups(request_path: Path, output_root: Path, reference_envelope) -> Path:
    """Attach native roof total and fill missing VT without mutating source evidence."""
    request_path = Path(request_path).resolve()
    request = json.loads(request_path.read_text(encoding="utf-8"))
    facts_path = Path(_text(request.get("pageFactsPath")))
    if not facts_path.is_file():
        return request_path

    facts = json.loads(facts_path.read_text(encoding="utf-8"))
    derived = copy.deepcopy(facts)
    pages = [p for p in list(derived.get("pages") or []) if _text(p.get("pageType")).upper() == "EN"]
    if not pages:
        return request_path

    changed = False
    roof_total = native_roof_schedule_total(request_path)
    if roof_total is not None:
        for page in pages:
            for row in list(page.get("envelope") or []):
                if _text(row.get("kind")).lower() != "roof" or _number(row.get("grossAreaFt2")) is None:
                    continue
                row["scheduleTotalFt2"] = roof_total["value"]
                row["scheduleTotalAuthority"] = roof_total["authority"]
                row["scheduleTotalEvidence"] = {
                    k: roof_total[k] for k in (
                        "scheduleName", "scheduleUniqueId", "rowIndex", "columnIndex",
                        "columnHeading", "cellText", "scheduleArtifact"
                    ) if k in roof_total
                }
                changed = True

    profiles = {}
    reference_proven = False
    try:
        profiles = reference_envelope._approved_profiles(reference_envelope._reference_path())
        all_rows = [row for page in pages for row in list(page.get("envelope") or [])
                    if float(row.get("confidence") or 0) >= 0.90]
        for row in all_rows:
            if not reference_envelope._has_any_thermal(row):
                continue
            cls = reference_envelope._class_for_row(row)
            profile = profiles.get(cls)
            if profile and reference_envelope._profile_matches_row(profile, row):
                reference_proven = True
                break
    except Exception:
        profiles = {}

    vt_filled = []
    for page in pages:
        for row in list(page.get("envelope") or []):
            kind = _text(row.get("kind")).lower()
            if kind not in {"window", "door"} or _row_has_vt(row):
                continue
            # r127 filing policy: preserve an actual VT when supplied; if VT is absent,
            # insert one deterministic project-wide value. Do not branch on tint/reference.
            value = MISSING_VT
            authority = "REVEX_FIXED_MISSING_VT_0_45"
            row["vt"] = round(float(value), 3)
            row["visibleTransmittanceAuthority"] = authority
            vt_filled.append({
                "kind": kind,
                "code": reference_envelope._row_code(row)[0],
                "vt": row["vt"],
                "authority": authority,
            })
            changed = True

    if not changed:
        return request_path

    output_root = Path(output_root).resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    audit = {
        "schema": "liber.revex.energy-final-touchups.v1",
        "version": VERSION,
        "sourceEngineeringRevision": request.get("revision"),
        "roofScheduleTotal": roof_total,
        "vtFallbacks": vt_filled,
        "referenceFamilyCorroborated": reference_proven,
        "sourceEvidenceMutated": False,
        "policy": "NATIVE_SCHEDULE_TOTAL_OVER_REGION_RESUM; ACTUAL_VT_ELSE_FIXED_0_45",
    }
    audit_path = output_root / "ENERGY_FINAL_TOUCHUPS_R125.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")
    facts_out = output_root / "00_PAGE_FACTS_FINAL_TOUCHUPS_R125.json"
    facts_out.write_text(json.dumps(derived, ensure_ascii=True, indent=2), encoding="utf-8")
    derived_request = dict(request)
    derived_request["pageFactsPath"] = str(facts_out)
    derived_request["finalTouchupsR125"] = {
        "schema": audit["schema"], "version": VERSION, "auditFile": audit_path.name,
        "roofScheduleTotalFt2": roof_total["value"] if roof_total else None,
        "vtFallbackCount": len(vt_filled),
    }
    request_out = output_root / "00_PIPELINE_REQUEST_FINAL_TOUCHUPS_R125.json"
    request_out.write_text(json.dumps(derived_request, ensure_ascii=True, indent=2), encoding="utf-8")
    print(json.dumps({
        "stage": "ENERGY_FINAL_TOUCHUPS_R125", "status": "APPLIED",
        "roofScheduleTotalFt2": roof_total["value"] if roof_total else None,
        "vtFallbackCount": len(vt_filled), "sourceEvidenceMutated": False,
    }, ensure_ascii=True), flush=True)
    return request_out


def _compact_label(value: Any, kind: str = "") -> str:
    text = _text(value)
    text = text.replace("Fa?ade", "Facade").replace("Façade", "Facade")
    text = re.sub(r"\[\s*Bldg\.?\s*Use[^\]]*\]", "", text, flags=re.I)
    text = re.sub(r"\bPerf\.?\s*Specs?\.?\s*:\s*", "", text, flags=re.I)
    text = re.sub(r"\bProduct\s+ID\s*[:=]?\s*[^,;]+[,;]?\s*", "", text, flags=re.I)
    text = re.sub(r"\bSHGC\s*[:=]?\s*\d*\.?\d+\s*[,;]?\s*", "", text, flags=re.I)
    text = re.sub(r"\b<?\s*95\s*['’]?\s*above[- ]grade\s*[,;]?\s*", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" ,;-")
    if not text:
        defaults = {"roof": "Roof", "wall": "Exterior wall", "window": "Window", "door": "Exterior door", "floor": "Floor"}
        text = defaults.get(kind, "Envelope assembly")
    if len(text) > 72:
        cut = text[:72].rsplit(" ", 1)[0].rstrip(" ,;-")
        text = cut or text[:72]
    return text


def patch_pipeline(module, reference_envelope=None) -> None:
    if getattr(module, "__revex_r125_patched__", False):
        return

    original_merge = module._merge_diagram_geometry_with_thermal
    def merge(diagram_row, thermal_rows):
        merged, error = original_merge(diagram_row, thermal_rows)
        if merged is None:
            return merged, error
        kind = _text(merged.get("kind")).lower()
        code, base_code = module._comcheck_row_code(diagram_row)
        same = [r for r in thermal_rows if _text(r.get("kind")).lower() == kind]
        exact = [r for r in same if code and module._comcheck_row_code(r)[0] == code]
        base = [r for r in same if base_code and module._comcheck_row_code(r)[1] == base_code]
        for candidate in exact or base or same:
            if candidate.get("vt") not in (None, ""):
                merged["vt"] = candidate.get("vt")
                break
        if code:
            merged["assemblyType"] = code
        merged["description"] = _compact_label(
            merged.get("_diagramDescription") or merged.get("description") or merged.get("assemblyType"), kind
        )
        return merged, error
    module._merge_diagram_geometry_with_thermal = merge

    original_roof = module._merge_roof_geometry_as_one_area
    def roof_merge(diagram_rows, thermal_rows):
        aggregate, errors, audit = original_roof(diagram_rows, thermal_rows)
        if aggregate is None:
            return aggregate, errors, audit
        totals = []
        for row in diagram_rows:
            if _text(row.get("kind")).lower() != "roof":
                continue
            value = _number(row.get("scheduleTotalFt2"))
            if value is not None:
                totals.append(round(value, 6))
        distinct = sorted({round(v, 3) for v in totals})
        if len(distinct) == 1:
            total = float(distinct[0])
            audit["roofRegionArithmeticAreaFt2"] = audit.get("roofAggregateAreaFt2")
            audit["roofAggregateAreaFt2"] = total
            audit["roofAreaAuthority"] = "ACTIVE_REVIT_NATIVE_SCHEDULE_TOTAL"
            audit["roofRegionSumUsedForFiling"] = False
            aggregate["grossAreaFt2"] = total
            aggregate["_geometrySource"] = "ACTIVE_REVIT_NATIVE_SCHEDULE_TOTAL"
            aggregate["_regionArithmeticAreaFt2"] = audit.get("roofRegionArithmeticAreaFt2")
        else:
            audit["roofAreaAuthority"] = "UNIQUE_DIAGRAM_REGION_SUM_FALLBACK"
            audit["roofRegionSumUsedForFiling"] = True
        return aggregate, errors, audit
    module._merge_roof_geometry_as_one_area = roof_merge

    original_canonicalize = module.canonicalize_comcheck_envelope_rows
    def canonicalize(en_pages):
        rows, audit = original_canonicalize(en_pages)
        for row in rows:
            kind = _text(row.get("kind")).lower()
            code = module._comcheck_row_code(row)[0]
            if code:
                row["assemblyType"] = code
            row["description"] = _compact_label(row.get("description") or row.get("assemblyType"), kind)
        audit["r125CompactAssemblyLabels"] = True
        return rows, audit
    module.canonicalize_comcheck_envelope_rows = canonicalize

    original_vt = module._row_visible_transmittance
    def visible_transmittance(row):
        current = original_vt(row)
        if current is not None:
            return current
        # r127 filing policy: all genuinely missing VT resolves to one stable value.
        return MISSING_VT
    module._row_visible_transmittance = visible_transmittance

    original_prepare = module.prepare_project_comcheck
    def prepare(facts, project_identity, filing_dir, log):
        cxl, audit_pdf, audit = original_prepare(facts, project_identity, filing_dir, log)
        if cxl is None or not Path(cxl).is_file():
            return cxl, audit_pdf, audit
        try:
            tree = ET.parse(cxl)
            root = tree.getroot()
            ns_uri = root.tag.split("}")[0].strip("{") if "}" in root.tag else ""
            q = lambda name: f"{{{ns_uri}}}{name}" if ns_uri else name

            def child(node, local):
                return next((c for c in list(node) if c.tag.rsplit("}",1)[-1] == local), None)

            fallbacks = []
            for node in root.iter():
                kind = node.tag.rsplit("}",1)[-1]
                if kind not in {"window", "door"}:
                    continue
                if child(node, "glazingType") is None:
                    continue
                vt_node = child(node, "propVt")
                if vt_node is not None and _text(vt_node.text):
                    continue
                description = _text((child(node, "description").text if child(node, "description") is not None else ""))
                assembly = _text((child(node, "assemblyType").text if child(node, "assemblyType") is not None else ""))
                probe = {"kind": kind, "description": description, "assemblyType": assembly}
                value = visible_transmittance(probe)
                if vt_node is None:
                    vt_node = ET.Element(q("propVt"))
                    # Keep schema ordering adjacent to SHGC/U values instead of appending blindly.
                    children = list(node)
                    insert_at = len(children)
                    for i, existing in enumerate(children):
                        local = existing.tag.rsplit("}",1)[-1]
                        if local in {"propShgc", "propUvalue"}:
                            insert_at = i + 1
                    node.insert(insert_at, vt_node)
                vt_node.text = f"{float(value):.3f}"
                fallbacks.append({"kind": kind, "assemblyType": assembly, "vt": float(value)})

            envelope_node = next((n for n in root.iter() if n.tag.rsplit("}",1)[-1] == "envelope"), None)
            if envelope_node is not None:
                use = child(envelope_node, "useVltDetails")
                if use is not None:
                    use.text = "true"

            if fallbacks:
                if ns_uri:
                    ET.register_namespace("", ns_uri)
                tree.write(cxl, encoding="utf-8", xml_declaration=True)
                audit = dict(audit or {})
                structure = dict(audit.get("cxlStructure") or {})
                structure["vtFallbackApplied"] = fallbacks
                structure["useVltDetails"] = True
                audit["cxlStructure"] = structure
                audit_json = Path(filing_dir) / "COMcheck_INPUT_AUDIT.json"
                if audit_json.is_file():
                    current = json.loads(audit_json.read_text(encoding="utf-8"))
                    current["cxlStructure"] = structure
                    audit_json.write_text(json.dumps(current, indent=2), encoding="utf-8")
                log.write("COMCHECK_VT_R125", "PASSED", fallbackCount=len(fallbacks), fallbacks=fallbacks)
        except Exception as exc:
            log.write("COMCHECK_VT_R125", "FAILED", error=f"{type(exc).__name__}: {exc}")
            raise
        return cxl, audit_pdf, audit
    module.prepare_project_comcheck = prepare
    module.__revex_r125_patched__ = True


def install_guard_touchups(r116_module, reference_envelope) -> None:
    """Patch each dynamically loaded r49 module before r116 uses it for preflight."""
    if getattr(r116_module, "__revex_r125_loader_patched__", False):
        return
    original_loader = r116_module._load_pipeline_module
    def load(impl):
        module = original_loader(impl)
        patch_pipeline(module, reference_envelope)
        return module
    r116_module._load_pipeline_module = load
    r116_module.__revex_r125_loader_patched__ = True


def _used_rows_cols(sheet) -> tuple[set[int], set[int]]:
    rows: set[int] = set()
    cols: set[int] = set()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                rows.add(cell.row)
                cols.add(cell.column)
    return rows, cols


def strict_en1_pdf(xlsx: Path, pdf: Path, output_root: Path) -> dict:
    """Export the original EN-1 workbook appearance; change visibility/fit only in a print copy."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.properties import PageSetupProperties
    from openpyxl.utils import get_column_letter
    from pypdf import PdfReader

    import revex_user_identity_en1 as en1

    source = Path(xlsx)
    source_bytes = source.stat().st_size
    print_copy = source.with_name("EN-1_PRINT_SOURCE_R125.xlsx")
    shutil.copy2(source, print_copy)
    workbook = load_workbook(print_copy)
    missing = [name for name in en1.EN1_PRINT_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError("EN-1 print set is missing filing sheets: " + ", ".join(missing))

    visibility_audit = []
    for sheet in workbook.worksheets:
        if sheet.title in en1.EN1_PRINT_SHEETS:
            sheet.sheet_state = "visible"
            used_rows, used_cols = _used_rows_cols(sheet)
            unhidden_rows = []
            unhidden_cols = []
            for row_index in used_rows:
                dim = sheet.row_dimensions[row_index]
                if dim.hidden:
                    dim.hidden = False
                    unhidden_rows.append(row_index)
            for col_index in used_cols:
                key = get_column_letter(col_index)
                dim = sheet.column_dimensions[key]
                if dim.hidden:
                    dim.hidden = False
                    unhidden_cols.append(key)
            props = sheet.sheet_properties.pageSetUpPr
            if props is None:
                props = PageSetupProperties()
                sheet.sheet_properties.pageSetUpPr = props
            props.fitToPage = True
            props.autoPageBreaks = False
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.scale = None
            # Manual breaks defeat fit-to-one-page in LibreOffice.
            try:
                sheet.row_breaks.brk = []
                sheet.col_breaks.brk = []
            except Exception:
                pass
            visibility_audit.append({
                "sheet": sheet.title,
                "nonEmptyRowsUnhidden": unhidden_rows,
                "nonEmptyColumnsUnhidden": unhidden_cols,
                "printAreaPreserved": str(sheet.print_area or ""),
                "fitToWidth": 1,
                "fitToHeight": 1,
            })
        else:
            sheet.sheet_state = "hidden"

    workbook.active = workbook.sheetnames.index(en1.EN1_PRINT_SHEETS[0])
    workbook.save(print_copy)
    workbook.close()

    # Source workbook must never be mutated by PDF visibility/fit work.
    if source.stat().st_size != source_bytes:
        raise ValueError("EN-1 source workbook changed during print-copy preparation")

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise ValueError("LibreOffice Calc is unavailable for EN-1 PDF export")
    profile = Path(tempfile.gettempdir()) / f"revex-r125-lo-{uuid.uuid4().hex}"
    export_dir = Path(tempfile.mkdtemp(prefix="revex-r125-en1-pdf-"))
    try:
        completed = subprocess.run(
            [soffice, "--headless", f"-env:UserInstallation=file://{profile.as_posix()}",
             "--convert-to", "pdf", "--outdir", str(export_dir), str(print_copy)],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", timeout=180,
        )
        candidate = export_dir / f"{print_copy.stem}.pdf"
        if completed.returncode != 0 or not candidate.is_file() or candidate.stat().st_size < 1024:
            raise ValueError("EN-1 LibreOffice PDF export failed: " + (completed.stdout or "")[-1200:])
        shutil.move(str(candidate), str(pdf))
    finally:
        print_copy.unlink(missing_ok=True)
        shutil.rmtree(export_dir, ignore_errors=True)
        shutil.rmtree(profile, ignore_errors=True)

    reader = PdfReader(str(pdf))
    expected = len(en1.EN1_PRINT_SHEETS)
    if len(reader.pages) != expected:
        raise ValueError(f"EN-1 PDF has {len(reader.pages)} pages; expected exactly {expected} one-page filing sheets")
    page_boxes = []
    for index, page in enumerate(reader.pages, start=1):
        width = float(page.mediabox.width); height = float(page.mediabox.height)
        crop_width = float(page.cropbox.width); crop_height = float(page.cropbox.height)
        if min(width, height, crop_width, crop_height) < 100:
            raise ValueError(f"EN-1 PDF page {index} has an invalid/cut page box")
        if crop_width > width + 1 or crop_height > height + 1:
            raise ValueError(f"EN-1 PDF page {index} crop box exceeds its media box")
        page_boxes.append({
            "page": index, "widthPt": round(width, 2), "heightPt": round(height, 2),
            "cropWidthPt": round(crop_width, 2), "cropHeightPt": round(crop_height, 2),
        })

    audit = {
        "schema": "liber.revex.en1-print-audit.v2",
        "version": VERSION,
        "status": "PASSED",
        "sourceWorkbook": source.name,
        "sourceWorkbookMutatedByPrint": False,
        "appearancePolicy": "PRESERVE_TEMPLATE_STYLES_MERGES_WIDTHS_HEIGHTS_PRINT_AREAS; CHANGE_VISIBILITY_AND_FIT_ONLY",
        "editableValuesOnly": True,
        "pdf": Path(pdf).name,
        "pageCount": len(reader.pages),
        "expectedPageCount": expected,
        "oneFilingSheetPerPdfPage": True,
        "sheetVisibility": visibility_audit,
        "pageBoxes": page_boxes,
    }
    audit_path = Path(output_root) / "EN-1_PRINT_AUDIT.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")
    return audit


def patch_en1_and_resume() -> None:
    """Patch both normal r116 finalization and r124 resumable publication."""
    try:
        import revex_user_identity_en1 as en1
        if not getattr(en1, "__revex_r125_print_patched__", False):
            en1._print_en1_pdf = strict_en1_pdf
            en1.__revex_r125_print_patched__ = True
    except Exception:
        en1 = None

    try:
        import revex_publication_resume_r124 as resume
    except Exception:
        return

    if getattr(resume, "__revex_r125_patched__", False):
        return

    original_pipeline = resume._pipeline
    def pipeline():
        module = original_pipeline()
        try:
            import revex_reference_envelope_projection_r118 as reference_envelope
        except Exception:
            reference_envelope = None
        patch_pipeline(module, reference_envelope)
        return module
    resume._pipeline = pipeline

    # Never use r124's crude whole-workbook PDF fallback. The strict template-preserving
    # exporter above is now the only EN-1 PDF path.
    def no_crude_pdf(xlsx, pdf):
        root = Path(xlsx).resolve().parent.parent
        return strict_en1_pdf(Path(xlsx), Path(pdf), root)
    resume._simple_pdf = no_crude_pdf
    resume.__revex_r125_patched__ = True


def install_worker_touchups() -> None:
    patch_en1_and_resume()
