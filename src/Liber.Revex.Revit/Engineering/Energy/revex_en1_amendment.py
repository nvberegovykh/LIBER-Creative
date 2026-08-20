#!/usr/bin/env python3
"""Publication-only EN-1 applicant/modeler amendment.

This lane is intentionally downstream of an already COMPLETE Energy result. It may
change only EN-1.xlsx, EN-1.pdf and the nine-file release ZIP. GeometryCo,
OpenStudio/EnergyPlus, COMcheck, project identity, signatures and seals are never run
or authored here. Every unchanged artifact remains an exact hash/path reference to
the immutable parent result.
"""
from __future__ import annotations

import datetime as dt
from decimal import Decimal, InvalidOperation
import hashlib
import json
from pathlib import Path
import posixpath
import tempfile
import xml.etree.ElementTree as ET
import zipfile


MODE = "EN1_IDENTITY_AMENDMENT"
SCHEMA = "liber.revex.en1-identity-amendment.v1"
PUBLIC_NAMES = {
    "GEOMETRY.osm", "BASELINE.osm", "PROPOSED.osm", "BASELINE_REPORT.html",
    "PROPOSED_REPORT.html", "EN-1.xlsx", "EN-1.pdf", "COMcheck_BACKSTOP.pdf",
    "PACKAGER_REPORTS.zip",
}
ALLOWED_INFO_CELLS = {
    "B11", "D11", "I11", "B12", "F12", "B13", "F13", "B14", "C14", "E14", "B15", "G15",
    "B19", "D19", "I19", "B20", "F20", "B21", "F21", "B22", "C22", "E22",
}
PROTECTED_ARCHIVE_PREFIXES = ("xl/media/", "xl/drawings/", "_xmlsignatures/")


def _text(value: object) -> str:
    return str(value or "").strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _artifact_key(row: dict) -> str:
    return _text(row.get("reviewName") or row.get("name"))


def validate_parent(data: dict) -> tuple[dict, list[dict]]:
    project_id = _text(data.get("projectId"))
    source_revision = _text(data.get("sourceRevision"))
    parent = dict(data.get("parentResult") or {})
    manifest = dict(parent.get("manifest") or {})
    rows = [dict(row) for row in list(parent.get("artifacts") or []) if isinstance(row, dict)]
    if _text(data.get("mode")) != MODE:
        raise ValueError("unsupported EN-1 amendment mode")
    if manifest.get("schema") != "liber.revex.energy-result.v1" or _text(manifest.get("status")).upper() != "COMPLETE":
        raise ValueError("EN-1 amendment requires an already COMPLETE Energy result")
    if _text(manifest.get("projectId")) != project_id or _text(manifest.get("sourceEngineeringRevision")) != source_revision:
        raise ValueError("EN-1 amendment parent project/Engineering revision mismatch")
    parent_revision = _text(parent.get("revision") or manifest.get("resultRevision"))
    if not parent_revision or parent_revision != _text(manifest.get("resultRevision")):
        raise ValueError("EN-1 amendment parent resultRevision mismatch")
    if manifest.get("revitWriteBack") is not False or manifest.get("pdfInsertion") is not False:
        raise ValueError("EN-1 amendment parent authority boundary is invalid")
    by_public = {_artifact_key(row): row for row in rows if _artifact_key(row) in PUBLIC_NAMES}
    missing = sorted(PUBLIC_NAMES - set(by_public))
    if missing:
        raise ValueError("EN-1 amendment parent is missing public artifacts: " + ", ".join(missing))
    for row in rows:
        if int(row.get("bytes") or 0) <= 0 or len(_text(row.get("sha256"))) != 64 or not _text(row.get("path")):
            raise ValueError(f"EN-1 amendment parent artifact lacks exact transfer evidence: {_artifact_key(row)}")
    if dict(manifest.get("comcheck") or {}).get("officialDoeReport") is not True:
        raise ValueError("EN-1 amendment parent has no verified official COMcheck report")
    return manifest, rows


def workbook_protected_snapshot(path: Path) -> dict:
    """Semantic workbook + signature/seal media snapshot excluding only editable identity cells."""
    from openpyxl import load_workbook

    def stable_value(value):
        if value is None or isinstance(value, (str, bool)):
            return value
        if isinstance(value, (int, float, Decimal)):
            try:
                return format(Decimal(str(value)).normalize(), "f")
            except InvalidOperation:
                return str(value)
        if isinstance(value, (dt.datetime, dt.date, dt.time)):
            return value.isoformat()
        attributes = getattr(value, "__dict__", None)
        if isinstance(attributes, dict):
            return {key: stable_value(item) for key, item in sorted(attributes.items())}
        return str(value)

    workbook = load_workbook(path, data_only=False, read_only=False)
    protected = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if sheet.title == "1,2,3 Information" and cell.coordinate in ALLOWED_INFO_CELLS:
                    continue
                if cell.value is None and not cell.has_style:
                    continue
                protected.append((sheet.title, cell.coordinate, cell.data_type, stable_value(cell.value),
                                  tuple(cell._style) if cell.has_style else ()))
        protected.append((sheet.title, "MERGES", tuple(sorted(str(item) for item in sheet.merged_cells.ranges))))
    workbook.close()
    semantic = hashlib.sha256(json.dumps(protected, ensure_ascii=True, separators=(",", ":")).encode("utf-8")).hexdigest()

    archive_parts: dict[str, str] = {}
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            lower = name.lower()
            if any(lower.startswith(prefix) for prefix in PROTECTED_ARCHIVE_PREFIXES):
                archive_parts[name] = hashlib.sha256(archive.read(name)).hexdigest()
    digital_signatures = [name for name in archive_parts if name.lower().startswith("_xmlsignatures/")]
    if digital_signatures:
        raise ValueError("A digitally signed EN-1 workbook cannot be amended without invalidating its signature")
    return {"semanticSha256": semantic, "protectedArchiveParts": archive_parts}


def assert_only_identity_changed(before: dict, after: dict) -> dict:
    if before.get("semanticSha256") != after.get("semanticSha256"):
        raise ValueError("EN-1 amendment changed workbook content outside Applicant/Modeler cells")
    if before.get("protectedArchiveParts") != after.get("protectedArchiveParts"):
        raise ValueError("EN-1 amendment changed signature/seal media or drawing parts")
    return {
        "protectedWorkbookSemanticSha256": before.get("semanticSha256"),
        "protectedArchivePartCount": len(before.get("protectedArchiveParts") or {}),
        "signatureSealChanged": False,
        "allowedChangedCells": sorted(ALLOWED_INFO_CELLS),
    }


def fill_people_identity_preserving_package(
        xlsx: Path, applicant: dict[str, str], modeler: dict[str, str], en1_module) -> dict:
    """Patch only target worksheet cell XML; copy every other XLSX ZIP part byte-for-byte."""
    mappings = ((en1_module.APPLICANT_FIELDS, applicant), (en1_module.MODELER_FIELDS, modeler))
    values = {
        cell: label + en1_module._clean(person.get(key))
        for mapping, person in mappings
        for cell, (label, key) in mapping.items()
    }
    if set(values) != ALLOWED_INFO_CELLS:
        raise ValueError("EN-1 identity mapping differs from the protected Applicant/Modeler cell contract")

    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", main)
    ET.register_namespace("r", office_rel)
    with zipfile.ZipFile(xlsx, "r") as source:
        infos = source.infolist()
        parts = {info.filename: source.read(info.filename) for info in infos}
    workbook = ET.fromstring(parts["xl/workbook.xml"])
    sheet_rid = ""
    for sheet in workbook.findall(f".//{{{main}}}sheet"):
        if sheet.get("name") == "1,2,3 Information":
            sheet_rid = sheet.get(f"{{{office_rel}}}id") or ""
            break
    if not sheet_rid:
        raise ValueError("EN-1 filing workbook is missing the Information sheet")
    relationships = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    target = ""
    for relationship in relationships.findall(f"{{{package_rel}}}Relationship"):
        if relationship.get("Id") == sheet_rid:
            target = relationship.get("Target") or ""
            break
    if not target:
        raise ValueError("EN-1 Information worksheet relationship is missing")
    sheet_part = posixpath.normpath(posixpath.join("xl", target.lstrip("/")))
    if sheet_part not in parts:
        raise ValueError("EN-1 Information worksheet part is missing")
    sheet_xml = ET.fromstring(parts[sheet_part])
    cells = {cell.get("r"): cell for cell in sheet_xml.findall(f".//{{{main}}}c")}
    missing = sorted(set(values) - set(cells))
    if missing:
        raise ValueError("EN-1 Information worksheet is missing identity cells: " + ", ".join(missing))
    for address, text in values.items():
        cell = cells[address]
        for child in list(cell):
            cell.remove(child)
        cell.set("t", "inlineStr")
        inline = ET.SubElement(cell, f"{{{main}}}is")
        node = ET.SubElement(inline, f"{{{main}}}t")
        node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        node.text = text
    parts[sheet_part] = ET.tostring(sheet_xml, encoding="utf-8", xml_declaration=True)

    temporary = xlsx.with_suffix(xlsx.suffix + ".amendment.tmp")
    try:
        with zipfile.ZipFile(temporary, "w") as destination:
            for info in infos:
                destination.writestr(info, parts[info.filename])
        temporary.replace(xlsx)
    finally:
        temporary.unlink(missing_ok=True)

    # Verify exact values from the package reader without rewriting it.
    from openpyxl import load_workbook
    verify = load_workbook(xlsx, read_only=True, data_only=False)
    info = verify["1,2,3 Information"]
    wrong = [address for address, expected in values.items() if str(info[address].value or "") != expected]
    verify.close()
    if wrong:
        raise ValueError("EN-1 applicant/modeler identity verification failed at: " + ", ".join(wrong))
    return {
        "applicantFields": sorted(key for key, value in applicant.items() if value),
        "modelerFields": sorted(key for key, value in modeler.items() if value),
    }


def _download_parent_rows(bucket, rows: list[dict], root: Path) -> dict[str, Path]:
    local: dict[str, Path] = {}
    for review_name in sorted(PUBLIC_NAMES):
        row = next(item for item in rows if _artifact_key(item) == review_name)
        suffix = Path(_text(row.get("name")) or review_name).suffix
        target = root / (review_name.replace("/", "_") + ("" if Path(review_name).suffix else suffix))
        bucket.blob(_text(row["path"])).download_to_filename(str(target))
        if target.stat().st_size != int(row["bytes"]) or _sha256(target).lower() != _text(row["sha256"]).lower():
            raise ValueError(f"EN-1 amendment parent transfer integrity failed: {review_name}")
        local[review_name] = target
    return local


def _package(path: Path, local: dict[str, Path]) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for review_name in sorted(PUBLIC_NAMES):
            archive.write(local[review_name], review_name)
    with zipfile.ZipFile(path) as archive:
        entries = [name for name in archive.namelist() if not name.endswith("/")]
        if len(entries) != 9 or set(entries) != PUBLIC_NAMES:
            raise ValueError("EN-1 amendment release ZIP is not the exact nine-file contract")


def attempt(data: dict):
    """Return a Flask response for MODE; return None for every normal Energy request."""
    if _text(data.get("mode")) != MODE:
        return None
    from flask import jsonify
    from google.cloud import storage
    import app as worker
    import revex_user_identity_en1 as en1

    manifest, parent_rows = validate_parent(data)
    project_id = _text(data.get("projectId"))
    source_revision = _text(data.get("sourceRevision"))
    prefix = _text(data.get("outputPrefix")).strip("/")
    bucket = storage.Client().bucket(_text(data.get("bucket")))
    if not prefix or not _text(data.get("bucket")):
        raise ValueError("EN-1 amendment output storage binding is incomplete")

    context = dict(data.get("projectSource") or {})
    applicant = en1._sanitize_person(context.get("en1Applicant"), modeler=False)
    modeler = en1._sanitize_person(context.get("en1Modeler"), modeler=True)
    with tempfile.TemporaryDirectory(prefix="revex-en1-amendment-") as temporary:
        root = Path(temporary)
        local = _download_parent_rows(bucket, parent_rows, root)
        xlsx = local["EN-1.xlsx"]
        before = workbook_protected_snapshot(xlsx)
        identity = fill_people_identity_preserving_package(xlsx, applicant, modeler, en1)
        after = workbook_protected_snapshot(xlsx)
        protection = assert_only_identity_changed(before, after)
        pdf = root / "EN-1.pdf"
        print_audit = en1._print_en1_pdf(xlsx, pdf, root)
        local["EN-1.pdf"] = pdf
        package = root / "REVEX_ENERGY_RELEASE_PACKAGE.zip"
        _package(package, local)

        changed = {"EN-1.xlsx": xlsx, "EN-1.pdf": pdf}
        result_rows = [dict(row) for row in parent_rows]
        for review_name, path in changed.items():
            index = next(i for i, row in enumerate(result_rows) if _artifact_key(row) == review_name)
            prior = result_rows[index]
            relative = f"05_FILING/{path.name}"
            uploaded = worker.upload_with_token(bucket, path, f"{prefix}/artifacts/{relative}")
            result_rows[index] = {
                **prior, **uploaded, "path": uploaded["path"], "relativePath": relative,
                "bytes": path.stat().st_size, "sha256": _sha256(path), "amended": True,
            }

        release_relative = package.name
        release_uploaded = worker.upload_with_token(bucket, package, f"{prefix}/artifacts/{release_relative}")
        release_row = {
            "name": package.name, "reviewName": package.name, "kind": "release-package", "userVisible": True,
            "path": release_uploaded["path"], "url": release_uploaded["url"], "relativePath": release_relative,
            "bytes": package.stat().st_size, "sha256": _sha256(package), "amended": True,
        }
        prior_release_index = next((i for i, row in enumerate(result_rows)
                                    if _text(row.get("kind")) == "release-package"), None)
        if prior_release_index is None:
            result_rows.append(release_row)
        else:
            result_rows[prior_release_index] = release_row

        now = dt.datetime.now(dt.timezone.utc)
        result_revision = "energy_" + now.strftime("%Y%m%dT%H%M%S%fZ") + "_en1"
        result = {
            **manifest,
            "schema": "liber.revex.energy-result.v1",
            "pipelineVersion": "0.8.19-r49",
            "sourceCandidate": worker.SOURCE_CANDIDATE or _text(manifest.get("sourceCandidate")),
            "projectId": project_id,
            "sourceEngineeringRevision": source_revision,
            "resultRevision": result_revision,
            "status": "COMPLETE",
            "finishedAt": now.isoformat(),
            "artifacts": result_rows,
            "en1Identity": {
                "schema": "liber.revex.en1-user-identity.v1",
                "authority": "explicit-user-input-publication-only-amendment",
                **identity,
                "transmittedToComcheck": False,
            },
            "en1Pdf": {
                "name": pdf.name,
                "bytes": pdf.stat().st_size,
                "sha256": _sha256(pdf),
                "pageCount": int(print_audit.get("pageCount") or 0),
            },
            "releasePackage": {
                "schema": "liber.revex.energy-release-package.v1",
                "name": package.name,
                "path": release_uploaded["path"],
                "bytes": package.stat().st_size,
                "sha256": _sha256(package),
                "entryCount": 9,
                "source": "EN1_IDENTITY_PUBLICATION_ONLY_AMENDMENT",
            },
            "amendment": {
                "schema": SCHEMA,
                "mode": MODE,
                "parentResultRevision": _text(manifest.get("resultRevision")),
                "sourceEngineeringRevision": source_revision,
                "geometryCoRerun": False,
                "openStudioRerun": False,
                "energyPlusRerun": False,
                "comcheckRerun": False,
                "projectIdentityChanged": False,
                **protection,
                "changedReviewArtifacts": ["EN-1.xlsx", "EN-1.pdf", package.name],
            },
            "revitWriteBack": False,
            "pdfInsertion": False,
            "authorityBoundary": "Publication-only EN-1 identity amendment; simulation, COMcheck, project identity, signature and seal remain unchanged.",
        }
        result_path = root / "energy-result.json"
        result_path.write_text(json.dumps(result, ensure_ascii=True, indent=2), encoding="utf-8")
        manifest_meta = worker.upload_with_token(bucket, result_path, f"{prefix}/energy-result.json", "application/json")
        return jsonify({
            "schema": "liber.revex.energy-server-response.v1",
            "projectId": project_id,
            "sourceRevision": source_revision,
            "resultRevision": result_revision,
            "status": "COMPLETE",
            "error": None,
            "manifest": result,
            "manifestPath": manifest_meta["path"],
            "manifestUrl": manifest_meta["url"],
            "artifacts": result_rows,
        }), 200


def install(app) -> None:
    original = app.view_functions.get("run_energy")
    if original is None or getattr(original, "__revex_en1_amendment__", False):
        return

    def en1_amendment_run_energy():
        from flask import request
        import app as worker
        data = request.get_json(silent=True) or {}
        if worker.TOKEN and request.headers.get("X-REVEX-Runner-Token", "") != worker.TOKEN:
            return original()
        response = attempt(data)
        return response if response is not None else original()

    en1_amendment_run_energy.__name__ = original.__name__
    en1_amendment_run_energy.__revex_en1_amendment__ = True
    app.view_functions["run_energy"] = en1_amendment_run_energy
