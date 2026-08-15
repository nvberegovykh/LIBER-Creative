#!/usr/bin/env python3
"""Offline r49 identity, transformer, and official-engine client regression QA."""

from __future__ import annotations

import importlib.util
import ast
import io
from collections import Counter
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import os
import shutil
import sys
import tempfile
import threading
from pathlib import Path
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
PIPELINE_PATH = HERE / "revex_energy_pipeline.py"
SPEC = importlib.util.spec_from_file_location("revex_energy_pipeline_r49", PIPELINE_PATH)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"Cannot load {PIPELINE_PATH}")
pipeline = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = pipeline
SPEC.loader.exec_module(pipeline)


def local_text(root, name: str) -> str:
    node = next((item for item in root.iter() if item.tag.rsplit("}", 1)[-1] == name), None)
    return str(node.text or "").strip() if node is not None else ""


LOAD_TYPES = (
    "OS:People", "OS:Lights", "OS:ElectricEquipment", "OS:GasEquipment",
    "OS:SteamEquipment", "OS:OtherEquipment", "OS:InternalMass",
    "OS:SpaceInfiltration", "OS:DesignSpecification:OutdoorAir",
)
HVAC_TYPES = (
    "OS:PlantLoop", "OS:AirLoopHVAC", "OS:ZoneHVAC", "OS:Coil", "OS:Fan",
    "OS:Boiler", "OS:Chiller", "OS:Pump", "OS:HeatExchanger", "OS:WaterHeater",
)


def behavior_inventory(compiler, model_path: Path) -> dict:
    counts = Counter(obj.obj_type for obj in compiler.parse_osm(model_path).objects)
    return {
        "schedules": {key: value for key, value in sorted(counts.items()) if key.startswith("OS:Schedule")},
        "loads": {key: value for key, value in sorted(counts.items()) if key.startswith(LOAD_TYPES)},
        "systems": {key: value for key, value in sorted(counts.items()) if key.startswith(HVAC_TYPES)},
    }


class MockComcheckHandler(BaseHTTPRequestHandler):
    pdf_bytes = b""
    index_count = 0
    upload_count = 0
    session_contract_verified = False

    def log_message(self, *_args) -> None:
        return

    def _write(self, body: bytes, content_type: str, status: int = 200, cookies: list[str] | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        for cookie in cookies or []:
            self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        if self.path.endswith("/index.html"):
            type(self).index_count += 1
            jsession = f"qa-http-{type(self).index_count}"
            self._write(b"<html>COMcheck test</html>", "text/html",
                        cookies=[f"JSESSIONID={jsession}; Path=/CheckWeb"])
        elif self.path.endswith("/dwr/engine.js"):
            body = (b'dwr.engine._sessionCookieName = "JSESSIONID";\n'
                    b'dwr.engine._instanceId = -1;\n'
                    b'var g={dwr:{_:[]}};\n'
                    b'dwr.engine._instanceId = g.dwr._.length;\n'
                    b'dwr.engine.transport.setDwrSession=function(dwrsess){};\n'
                    b'// DWRSESSIONID\n')
            self._write(body, "text/javascript")
        elif "/report.html?" in self.path:
            self._write(b'<a href="report/current/pdf">Your compliance report</a>', "text/html")
        elif self.path.endswith("/report/current/pdf"):
            self._write(self.pdf_bytes, "application/pdf")
        else:
            self._write(b"not found", "text/plain", 404)

    def do_POST(self) -> None:
        length = int(self.headers.get("Content-Length") or 0)
        request_body = self.rfile.read(length)
        cookie_header = self.headers.get("Cookie") or ""
        if self.path.endswith("/__System.generateId.dwr"):
            # DWR3 correlates the Java application session through JSESSIONID. Its browser
            # engine does not add the legacy httpSessionId request field during bootstrap.
            assert b"httpSessionId=" not in request_body, request_body[:2000]
            assert b"instanceId=0\n" in request_body, request_body[:2000]
            assert b"instanceId=-1" not in request_body, request_body[:2000]
            assert b"scriptSessionId=\n" in request_body, request_body[:2000]
            # Exercise the real DWR3 response shape: plain calls are prefixed with the
            # script-tag protection statement when allowScriptTagRemoting is false, then
            # wrapped for the initialized engine instance. REVEX must still parse callback.
            body = (b"throw 'allowScriptTagRemoting is false.';\n//#DWR-REPLY\n"
                    b"(function(){if(!window.dwr)return;var dwr=window.dwr._[0];"
                    b'dwr.engine.remote.handleCallback( "0", "0", "REVEXR49DWRSESSION123456789012345" );})();')
            self._write(body, "text/javascript")
            return
        if self.path.endswith("/ProjectService.uploadProject.dwr"):
            type(self).upload_count += 1
            expected_http = f"qa-http-{type(self).index_count}".encode()
            required = [
                b'name="instanceId"\r\n\r\n0',
                b'name="page"\r\n\r\n%2FCheckWeb%2Findex.html',
                b'name="scriptSessionId"\r\n\r\nREVEXR49DWRSESSION123456789012345/',
            ]
            assert all(token in request_body for token in required), request_body[:5000]
            assert b'name="httpSessionId"' not in request_body, request_body[:5000]
            assert f"JSESSIONID=qa-http-{type(self).index_count}" in cookie_header, cookie_header
            assert "DWRSESSIONID=REVEXR49DWRSESSION123456789012345" in cookie_header, cookie_header
            assert b"revex-" not in request_body
            type(self).session_contract_verified = True
            if type(self).upload_count == 1:
                body = (b'<html><script>dwr.engine._executeScript("dwr.engine.remote.handleException('
                        b'\\\"1\\\",\\\"0\\\",{javaClassName:\\\"gov.energycodes.check.common.exception.InvalidSessionException\\\",message:null});");'
                        b'</script></html>')
            else:
                body = b'dwr.engine.remote.handleCallback("1","0",{projectName:"CURRENT TEST PROJECT"});'
            self._write(body, "text/html")
            return
        if self.path.endswith("/ProjectService.calculateEnvelopeCompliance.dwr"):
            body = b'dwr.engine.remote.handleCallback("2","0",{envelopeStatus:{passes:true,complianceIndex:12.3,statusMessage:"Passes"}});'
        else:
            body = b"dwr.engine.remote.handleBatchException({message:'unexpected test endpoint'},'0');"
        self._write(body, "text/javascript")

def mock_comcheck_pdf(path: Path) -> bytes:
    from reportlab.pdfgen import canvas
    pdf = path / "official-test.pdf"
    page = canvas.Canvas(str(pdf))
    page.drawString(72, 740, "COMcheck-Web Compliance Report")
    page.drawString(72, 720, "2020 NYCECC Appendix CA Modeling Envelope Backstop")
    page.drawString(72, 700, "CURRENT TEST PROJECT — 999 CURRENT AVENUE")
    page.save()
    return pdf.read_bytes()


def mock_en1_workbook(path: Path) -> Path:
    from openpyxl import Workbook
    target = path / "EN-1_READY_TO_INSERT.xlsx"
    wb = Workbook()
    first = wb.active
    first.title = pipeline.EN1_PRINT_SHEETS[0]
    for name in pipeline.EN1_PRINT_SHEETS[1:]:
        wb.create_sheet(name)
    wb.save(target)
    return target


def main() -> int:
    required = [
        pipeline.EN1_TEMPLATE, pipeline.COMCHECK_CXL_TEMPLATE,
        pipeline.BASELINE_REFERENCE, pipeline.PROPOSED_REFERENCE, pipeline.GEOMETRYCO,
        pipeline.PACKAGER,
    ]
    missing = [str(path) for path in required if not path.is_file()]
    if missing:
        raise AssertionError("Missing r49 filing structure dependencies: " + ", ".join(missing))

    # Accepted >=80% gbXML must not remain internally marked as failed merely because
    # the stricter read-only proof is below the 95% review target. Exercise the exact
    # pure helper from the Revit exporter without importing its Revit/Python.NET runtime.
    gbxml_engine = HERE.parent / "Gbxml" / "LIBER_gbXML_Preflight_and_Export.py"
    gbxml_graph = HERE.parent / "Gbxml" / "LIBER_gbXML_Preflight_and_Export.dyn"
    engine_source = gbxml_engine.read_text(encoding="utf-8-sig")
    graph = json.loads(gbxml_graph.read_text(encoding="utf-8-sig"))
    python_nodes = [node for node in graph["Nodes"] if node.get("NodeType") == "PythonScriptNode"]
    assert len(python_nodes) == 1 and python_nodes[0]["Code"] == engine_source, "gbXML .dyn/Python identity drift"
    call_marker = "        if acceptable:\n            reconcile_publication_message_severity(messages, publication_threshold_met)\n"
    assert call_marker in engine_source, "accepted gbXML path does not invoke publication severity reconciliation"
    engine_ast = ast.parse(engine_source)
    helper_node = next(
        node for node in engine_ast.body
        if isinstance(node, ast.FunctionDef) and node.name == "reconcile_publication_message_severity"
    )
    helper_module = ast.Module(body=[helper_node], type_ignores=[])
    ast.fix_missing_locations(helper_module)
    helper_ns: dict = {}
    exec(compile(helper_module, str(gbxml_engine), "exec"), helper_ns)
    reconcile_messages = helper_ns["reconcile_publication_message_severity"]
    accepted_messages = [
        {"severity": "ERROR", "code": "REVIT_TO_GBXML_GEOMETRY_INTEGRITY_FAILED", "message": "strict proof"},
        {"severity": "ERROR", "code": "OTHER_FATAL", "message": "must remain fatal"},
    ]
    reconcile_messages(accepted_messages, True)
    assert accepted_messages[0]["severity"] == "WARNING"
    assert accepted_messages[0]["code"] == "REVIT_TO_GBXML_GEOMETRY_INTEGRITY_REVIEW"
    assert accepted_messages[0]["original_code"] == "REVIT_TO_GBXML_GEOMETRY_INTEGRITY_FAILED"
    assert accepted_messages[1]["severity"] == "ERROR"
    blocked_messages = [{"severity": "ERROR", "code": "REVIT_TO_GBXML_GEOMETRY_INTEGRITY_FAILED"}]
    reconcile_messages(blocked_messages, False)
    assert blocked_messages[0]["severity"] == "ERROR", "sub-80 geometry integrity failure was incorrectly downgraded"

    z_project = {
        "title": "CURRENT TEST PROJECT", "address": "999 CURRENT AVENUE",
        "houseNumber": "999", "streetName": "CURRENT AVENUE", "borough": "BROOKLYN",
        "city": "BROOKLYN", "state": "NY", "zip": "11201", "block": "1234",
        "lot": "56", "bin": "9876543", "communityBoard": "2", "jobType": "ALT-CO",
        "architecturalJobNumber": "A0001", "mechanicalJobNumber": "M0002",
        "plumbingJobNumber": "P0003", "energyCode": None,
    }
    facts = {
        "status": "COMPLETE",
        "structuredIdentity": {
            "title": "CURRENT TEST PROJECT", "address": "999 CURRENT AVENUE",
            "city": "BROOKLYN", "state": "NY", "zip": "11201",
            "evidenceDigest": "a" * 64, "evidenceSheets": ["T001 · TITLE", "Z001 · ZONING"],
        },
        "comcheckSemanticVersion": "qa-schedule-semantic",
        "comcheckSemantic": {
            "energyCode": "2020 NYC Energy Conservation Code - NYC Stretch",
            "wholeBuildingType": "Multifamily",
            "floorAreaFt2": 2000,
            "climateZone": "4A",
            "sources": [{"semanticType": "wholeBuildingType", "evidence": "Building Area 1-Multifamily: Residential Floor Area 2000"}],
            "scheduleNamesAuthoritative": False,
        },
        "pages": [
            {"pageType": "T", "confidence": 0.99, "project": {**z_project, "title": None},
             "bulk": {"stories": 4, "buildingHeightFt": 55}},
            {"pageType": "Z", "confidence": 1.0, "project": {**z_project, "address": None, "city": None, "state": None, "zip": None},
             "bulk": {"stories": 4, "buildingHeightFt": 55}},
            {"pageType": "EN", "confidence": 1.0,
             "project": {"title": "FORBIDDEN EN IDENTITY", "address": "DO NOT COPY",
                         "energyCode": "2020 NYC Energy Conservation Code - NYC Stretch"},
             "envelope": [
                 {"kind": "wall", "assemblyType": "steel framed", "description": "CURRENT WALL",
                  "orientation": "N", "grossAreaFt2": 1000, "cavityR": 19, "continuousR": 5, "confidence": 1},
                 {"kind": "roof", "assemblyType": "insulation entirely above deck", "description": "CURRENT ROOF",
                  "grossAreaFt2": 500, "continuousR": 30, "confidence": 1},
                 {"kind": "floor", "assemblyType": "mass", "description": "CURRENT FLOOR",
                  "grossAreaFt2": 500, "continuousR": 10, "confidence": 1},
             ],
             "lighting": {"wholeBuildingType": "Multifamily", "floorAreaFt2": 2000,
                          "lpdWPerFt2": 0.65,
                          "fixtures": [{"description": "CURRENT FIXTURE", "wattage": 20,
                                        "quantity": 65, "confidence": 1}],
                          "exteriorUses": []}},
        ],
    }
    identity = pipeline.current_project_identity(facts)
    assert not identity["missing"], identity["missing"]
    assert identity["title"] == z_project["title"]
    assert identity["address"] == z_project["address"]
    assert identity["evidenceDigest"] == "a" * 64
    assert "FORBIDDEN" not in str(identity)
    same_project_identity = {**identity, "title": "79 WINTHROP STREET", "address": "79 WINTHROP STREET"}
    pipeline.assert_no_reference_identity_text(
        "79 WINTHROP STREET", "same-project publication identity", same_project_identity
    )
    try:
        pipeline.assert_no_reference_identity_text(
            "79 WINTHROP STREET / BEREGOVYKH", "reference-person leak", same_project_identity
        )
        raise AssertionError("Reference applicant/modeler identity was accepted as current-project identity")
    except pipeline.PipelineError:
        pass

    consent = {
        "schema": pipeline.COMCHECK_CONSENT_SCHEMA,
        "projectId": "project_qa",
        "sourceEngineeringRevision": "eng_qa_immutable_001",
        "service": pipeline.COMCHECK_SERVICE,
        "endpoint": pipeline.COMCHECK_ENDPOINT,
        "scope": pipeline.COMCHECK_SCOPE,
        "approved": True,
        "approvedAt": datetime.now(timezone.utc).isoformat(),
        "approvedByUid": "user_qa",
        "immutable": True,
    }
    try:
        pipeline.require_comcheck_consent({}, "project_qa", "eng_qa_immutable_001")
        raise AssertionError("Missing consent was accepted")
    except pipeline.PipelineError:
        pass
    try:
        pipeline.require_comcheck_consent(consent, "project_qa", "eng_different")
        raise AssertionError("Consent for a different immutable revision was accepted")
    except pipeline.PipelineError:
        pass
    assert pipeline.require_comcheck_consent(consent, "project_qa", "eng_qa_immutable_001") == consent

    # COMcheck reconciliation: EN thermal-boundary diagrams own orientation/area while
    # EN/COMcheck rows own U/SHGC/R. A stray filing-table number or dropped area cell must
    # never become building geometry when complete diagram geometry exists.
    reconciliation_pages = [{
        "pageType": "EN", "confidence": 1.0, "envelope": [
            {"kind": "door", "assemblyType": "D4_Door filing row", "grossAreaFt2": 19,
             "uFactor": 0.30, "shgc": 0.30, "confidence": 1.0,
             "evidence": "D4_Door filing row 19 0.300 0.300"},
            {"kind": "window", "assemblyType": "G2.1 Window filing row", "orientation": "WEST",
             "uFactor": 0.30, "shgc": 0.30, "confidence": 1.0,
             "evidence": "G2.1 Window filing row 0.300 0.300"},
            {"kind": "wall", "assemblyType": "W1 filing row", "orientation": "EAST",
             "grossAreaFt2": 999, "cavityR": 13, "continuousR": 8, "confidence": 1.0,
             "evidence": "W1 filing row 999 R13 R8"},
            {"kind": "door", "assemblyType": "Glass Door D4.1", "orientation": "EAST",
             "grossAreaFt2": 74, "confidence": 1.0, "evidence": "D4.1 Glass Door 74 SF"},
            {"kind": "window", "assemblyType": "Window Fixed G2.1", "orientation": "WEST",
             "grossAreaFt2": 146, "confidence": 1.0, "evidence": "G2.1 Window Fixed 146 SF"},
            {"kind": "wall", "assemblyType": "Wall W1.1", "orientation": "EAST",
             "grossAreaFt2": 759, "confidence": 1.0, "evidence": "W1.1 Wall 759 SF"},
        ]
    }]
    reconciled, reconciliation = pipeline.canonicalize_comcheck_envelope_rows(reconciliation_pages)
    assert reconciliation["geometryMode"] == "EN_THERMAL_BOUNDARY_DIAGRAM_PLUS_THERMAL_TABLE"
    assert reconciliation["thermalPropertyMergeErrorCount"] == 0, reconciliation
    assert any(r.get("kind") == "door" and r.get("orientation") == "EAST" and float(r.get("grossAreaFt2")) == 74 for r in reconciled)
    assert any(r.get("kind") == "window" and r.get("orientation") == "WEST" and float(r.get("grossAreaFt2")) == 146 for r in reconciled)
    assert any(r.get("kind") == "wall" and r.get("orientation") == "EAST" and float(r.get("grossAreaFt2")) == 759 for r in reconciled)
    assert not any(float(r.get("grossAreaFt2") or 0) == 19 for r in reconciled), "Spurious filing-table area leaked into authoritative diagram geometry"
    assert not any(float(r.get("grossAreaFt2") or 0) == 999 for r in reconciled), "Aggregated filing-table geometry overrode the EN thermal-boundary diagram"

    # Roof diagrams are geometry regions, not independent COMcheck constructions. R1/R2/R3
    # must become one summed roof area when at least one region proves the single current roof
    # R-value construction. An opaque U-factor-only floor schedule row is not a usable CXL
    # construction and must not outrank the current R-value floor row.
    roof_pages = [{
        "pageType": "EN", "confidence": 1.0, "envelope": [
            {"kind": "roof", "assemblyType": "R1 roof construction", "cavityR": 26.4, "continuousR": 15,
             "confidence": 1.0, "evidence": "R1 roof R26.4 R15"},
            {"kind": "roof", "assemblyType": "Roof region R1", "grossAreaFt2": 1000,
             "confidence": 1.0, "evidence": "R1 roof region 1000 SF"},
            {"kind": "roof", "assemblyType": "Roof region R2", "grossAreaFt2": 700,
             "confidence": 1.0, "evidence": "R2 roof region 700 SF"},
            {"kind": "roof", "assemblyType": "Roof region R3", "grossAreaFt2": 813,
             "confidence": 1.0, "evidence": "R3 roof region 813 SF"},
            {"kind": "floor", "assemblyType": "F1 U-only schedule artifact", "grossAreaFt2": 218,
             "uFactor": 0.10, "confidence": 1.0, "evidence": "F1 floor U-0.10"},
            {"kind": "floor", "assemblyType": "F1 floor construction", "continuousR": 8,
             "confidence": 1.0, "evidence": "F1 floor R8"},
            {"kind": "floor", "assemblyType": "Floor region F1", "grossAreaFt2": 218,
             "confidence": 1.0, "evidence": "F1 floor region 218 SF"},
        ]
    }]
    roof_reconciled, roof_reconciliation = pipeline.canonicalize_comcheck_envelope_rows(roof_pages)
    roofs = [row for row in roof_reconciled if row.get("kind") == "roof"]
    floors = [row for row in roof_reconciled if row.get("kind") == "floor"]
    assert len(roofs) == 1, roofs
    assert abs(float(roofs[0]["grossAreaFt2"]) - 2513.0) < 0.001, roofs[0]
    assert float(roofs[0]["cavityR"]) == 26.4 and float(roofs[0]["continuousR"]) == 15.0
    assert roof_reconciliation["roofGeometryRowsAggregated"] == 3, roof_reconciliation
    assert abs(float(roof_reconciliation["roofAggregateAreaFt2"]) - 2513.0) < 0.001
    assert roof_reconciliation["thermalPropertyMergeErrorCount"] == 0, roof_reconciliation
    assert len(floors) == 1 and float(floors[0]["continuousR"]) == 8.0, floors

    # COMcheck child openings must fit inside same-orientation gross wall hosts. This reproduces
    # the production class where all north fenestration was previously attached to the first wall.
    host_a = ET.Element("wall-a")
    host_b = ET.Element("wall-b")
    opening_plan, host_errors = pipeline._opening_host_plan(
        [(host_a, {"orientation": "NORTH", "grossAreaFt2": 300, "assemblyType": "A"}),
         (host_b, {"orientation": "NORTH", "grossAreaFt2": 1200, "assemblyType": "B"})],
        [{"kind": "window", "orientation": "NORTH", "grossAreaFt2": 700, "assemblyType": "G1"},
         {"kind": "door", "orientation": "NORTH", "grossAreaFt2": 250, "assemblyType": "D1"}],
    )
    assert not host_errors, host_errors
    assert len(opening_plan) == 2
    assert any(host is host_b for _, host in opening_plan), "Large opening was not moved to a wall with enough gross capacity."

    template_tree = ET.parse(pipeline.COMCHECK_CXL_TEMPLATE)
    template_root = template_tree.getroot()
    template_walls = [node for node in template_root.iter() if node.tag.rsplit("}", 1)[-1] == "agWall"]
    assert pipeline._xml_child_text(pipeline._wall_exemplar(template_walls, "Solid Concrete, 12in. Thickness, Furring: Metal"), "wallType") == "CONCRETE_AG_WALL"
    assert pipeline._xml_child_text(pipeline._wall_exemplar(template_walls, "Concrete Block, 8in., Furring: Metal"), "wallType") == "MASONRY_AG_WALL"
    assert pipeline._xml_child_text(pipeline._wall_exemplar(template_walls, "Steel-Framed, 16in. o.c."), "wallType") == "METAL_FRAME_16_AG_WALL"

    from comcheck_backstop import ComcheckClient, _dwr_error
    assert _dwr_error('dwr.engine.remote.handleException("0","0",{message:"schema invalid: test"});') == "schema invalid: test"
    assert _dwr_error("dwr.engine.remote.handleBatchException({message:'upload failed'},'0');") == "upload failed"
    # Match DWR 3 engine.js tokenify exactly: least-significant 6-bit digits are emitted first.
    assert ComcheckClient._dwr_tokenify(1) == "2"
    assert ComcheckClient._dwr_tokenify(64) == "12"
    assert ComcheckClient._parse_engine_contract('dwr.engine._instanceId = -1;')[1] == "0"

    workbook = load_workbook(pipeline.EN1_TEMPLATE)
    info = workbook["1,2,3 Information"]
    pipeline.blank_en1_identity_fields(info)
    pipeline.apply_en1_project_identity(info, identity)
    pipeline.assert_en1_identity_fields_blank(info)
    pipeline.assert_en1_project_identity(info, identity)
    pipeline.assert_no_reference_identity_workbook(workbook)
    assert "999" in str(info["B5"].value)
    assert "CURRENT AVENUE" in str(info["C5"].value)
    with tempfile.TemporaryDirectory(prefix="revex-r49-en1-qa-") as temp:
        transformed = Path(temp) / "EN-1_CURRENT_PROJECT.xlsx"
        workbook.save(transformed)
        pipeline.assert_no_reference_identity_xlsx(transformed)

    with tempfile.TemporaryDirectory(prefix="revex-r49-energy-qa-") as temp:
        folder = Path(temp)
        log = pipeline.RunLog(folder, "r49-offline-qa", "release QA")
        console_bytes = io.BytesIO()
        console_probe = io.TextIOWrapper(console_bytes, encoding="cp1252", errors="strict")
        prior_stdout = sys.stdout
        try:
            sys.stdout = console_probe
            log.write("WINDOWS_CONSOLE_ENCODING", "PASSED", replacement="\ufffd", ellipsis="…")
            console_probe.flush()
        finally:
            sys.stdout = prior_stdout
        assert b"\\ufffd" in console_bytes.getvalue()
        subprocess_environment = pipeline.utf8_subprocess_environment()
        assert subprocess_environment["PYTHONUTF8"] == "1"
        assert subprocess_environment["PYTHONIOENCODING"] == "utf-8"
        cxl, audit_pdf, audit = pipeline.prepare_project_comcheck(facts, identity, folder, log)
        assert audit["status"] == "INPUT_READY", audit
        assert audit["officialDoeReport"] is None, "Input preparation must not record a failed official report."
        assert audit["officialDoeReportStatus"] == "NOT_RUN"
        assert cxl and cxl.is_file()
        assert audit_pdf.is_file()
        root = ET.parse(cxl).getroot()
        assert local_text(root, "projectTitle") == z_project["title"]
        assert local_text(root, "projectAddress") == z_project["address"]
        assert local_text(root, "projectCity") == z_project["city"]
        assert local_text(root, "projectState") == z_project["state"]
        assert local_text(root, "projectZipCode") == z_project["zip"]
        assert local_text(root, "floorArea") == "2000.000"
        assert {local_text(node, "grossArea") for node in root.iter() if node.tag.rsplit("}", 1)[-1] in ("agWall", "roof", "floor")} == {"1000.000", "500.000"}
        assert "1991.000" not in cxl.read_text(encoding="utf-8"), "Template model areas must not survive into the current-project CXL."
        # powerDensity is code-derived whole-building-use metadata, not the project's proposed LPD.
        # Proposed lighting is represented by current fixture rows; never overwrite the code enum
        # metadata with an unrelated scan value.
        assert local_text(root, "powerDensity") == "0.68"
        assert local_text(root, "wholeBldgType") == "WHOLE_BUILDING_MULTIFAMILY"
        assert local_text(root, "state") == "New York", "COMcheck climate-location enum must not be postal NY."
        location = next(node for node in root.iter() if node.tag.rsplit("}", 1)[-1] == "location")
        assert local_text(location, "city") == "New York", "COMcheck climate-location enum must not be postal Brooklyn."
        assert next((node for node in root.iter() if node.tag.rsplit("}", 1)[-1] == "lighting"), None) is not None
        assert next((node for node in root.iter() if node.tag.rsplit("}", 1)[-1] == "wholeBldgUse"), None) is not None
        assert not pipeline._validate_comcheck_cxl_structure(root)
        assert audit["cxlStructure"]["scheduleSemanticVersion"] == "qa-schedule-semantic"
        assert audit["cxlStructure"]["scheduleNamesAuthoritative"] is False
        assert audit["cxlStructure"]["currentFloorAreaFt2"] == 2000
        slab = next(node for node in root.iter() if node.tag.rsplit("}", 1)[-1] == "floor")
        assert local_text(slab, "depthOfInsulation"), "Slab structural insulation depth was stripped."
        assert local_text(slab, "insulationPosition"), "Slab structural insulation position was stripped."
        assert "FORBIDDEN EN IDENTITY" not in cxl.read_text(encoding="utf-8")

        stamped_models = []
        for role, source in (("BASELINE", pipeline.BASELINE_REFERENCE), ("PROPOSED", pipeline.PROPOSED_REFERENCE)):
            stamped = folder / f"{role}_CURRENT_PROJECT.osm"
            shutil.copy2(source, stamped)
            pipeline.stamp_compiled_project_identity(stamped, identity, role, log)
            stamped_text = stamped.read_text(encoding="utf-8")
            assert z_project["title"] in stamped_text
            pipeline.assert_no_reference_identity_text(stamped_text, f"{role} QA model")
            stamped_models.append(stamped)

        # OSM fields are comma/semicolon delimited, so arbitrary project identity text
        # must survive escaped delimiter characters without corrupting object boundaries.
        punctuation_identity = {
            **identity,
            "title": "CURRENT & TEST, PROJECT; PHASE A",
            "address": "999 CURRENT, AVENUE; SUITE A",
        }
        punctuation_model = folder / "BASELINE_PUNCTUATION_IDENTITY.osm"
        shutil.copy2(pipeline.BASELINE_REFERENCE, punctuation_model)
        pipeline.stamp_compiled_project_identity(
            punctuation_model, punctuation_identity, "BASELINE_PUNCTUATION", log
        )
        punctuation_text = punctuation_model.read_text(encoding="utf-8")
        assert "CURRENT & TEST&#44 PROJECT&#59 PHASE A" in punctuation_text
        assert "999 CURRENT&#44 AVENUE&#59 SUITE A" in punctuation_text
        punctuation_compiler = pipeline.load_module(
            pipeline.GEOMETRYCO, "revex_geometryco_punctuation_identity_qa"
        )
        punctuation_reparsed = punctuation_compiler.parse_osm(punctuation_model)
        assert len(punctuation_reparsed.by_handle) == len(punctuation_reparsed.objects)
        actual_duplicate = folder / "ACTUAL_DUPLICATE_HANDLE.osm"
        actual_duplicate.write_text(
            "OS:Facility,\n  {00000000-0000-0000-0000-000000000001};\n\n"
            "OS:Building,\n  {00000000-0000-0000-0000-000000000001},\n  X;\n",
            encoding="utf-8",
        )
        try:
            punctuation_compiler.parse_osm(actual_duplicate)
        except punctuation_compiler.CompileError as exc:
            assert "Duplicate handle" in str(exc)
        else:
            raise AssertionError("GeometryCo parser must still reject genuine duplicate object handles.")

        MockComcheckHandler.pdf_bytes = mock_comcheck_pdf(folder)
        MockComcheckHandler.index_count = 0
        MockComcheckHandler.upload_count = 0
        MockComcheckHandler.session_contract_verified = False
        server = ThreadingHTTPServer(("127.0.0.1", 0), MockComcheckHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        os.environ["REVEX_ALLOW_COMCHECK_TEST_ENDPOINT"] = "true"
        try:
            from comcheck_backstop import run_official_backstop
            report, response, official = run_official_backstop(
                cxl, folder, identity,
                lambda status, **detail: log.write("COMCHECK_BACKSTOP_QA", status, **detail),
                base_url=f"http://127.0.0.1:{server.server_port}/CheckWeb/"
            )
        finally:
            server.shutdown()
            server.server_close()
            os.environ.pop("REVEX_ALLOW_COMCHECK_TEST_ENDPOINT", None)
        assert report.is_file() and report.read_bytes().startswith(b"%PDF-")
        assert response.is_file()
        assert official["officialDoeReport"] is True
        assert official["code"] == "2020 NYCECC Appendix CA Modeling Envelope Backstop"
        assert MockComcheckHandler.session_contract_verified
        assert MockComcheckHandler.upload_count == 2, "InvalidSessionException must trigger exactly one fresh session bootstrap."
        assert MockComcheckHandler.index_count == 2, "Fresh session retry must discard and rebuild HTTP/DWR cookies."
        compiler = pipeline.load_module(pipeline.GEOMETRYCO, "revex_geometryco_record_contract_qa")
        vectors, local_ai_provider, local_ai_error = compiler._local_ai_embeddings([
            "Architectural bedroom adjacent to corridor on an upper level.",
            "Unconditioned mechanical shaft with exterior exposure.",
        ])
        assert local_ai_error is None, local_ai_error
        assert vectors and len(vectors) == 2 and all(len(vector) == 384 for vector in vectors)
        assert local_ai_provider in {
            "DmlExecutionProvider", "CUDAExecutionProvider", "OpenVINOExecutionProvider",
            "CoreMLExecutionProvider", "CPUExecutionProvider",
        }
        log.dependency(
            "GeometryCo local-AI inference", True,
            provider=local_ai_provider, model="onnx-community/all-MiniLM-L6-v2-ONNX",
            vectors=len(vectors), dimensions=len(vectors[0]),
        )
        compiled = folder / "02_COMPILED_MODELS"
        geometry_fixture = folder / "APPROVED_GEOMETRY_FIXTURE.osm"
        shutil.copy2(pipeline.BASELINE_REFERENCE, geometry_fixture)
        compilation = compiler.compile_baseline_proposed_pair(
            geometry_fixture,
            pipeline.BASELINE_REFERENCE,
            pipeline.PROPOSED_REFERENCE,
            compiled,
            require_native_check=False,
        )
        assert compilation["success"] is True
        expected_geometry = {"spaces": 159, "surfaces": 1930, "subsurfaces": 294}
        expected_schedules = {"baseline": 243, "proposed": 89}
        expected_systems = {"baseline": 13, "proposed": 18}
        reference_paths = {"baseline": pipeline.BASELINE_REFERENCE, "proposed": pipeline.PROPOSED_REFERENCE}
        output_paths = {
            "baseline": compiled / "BASELINE_UPDATED_GEOMETRY.osm",
            "proposed": compiled / "PROPOSED_UPDATED_GEOMETRY.osm",
        }
        for role in ("baseline", "proposed"):
            role_report = compilation["reports"][role]
            assert role_report["exact_geometry_lock"]["passed"] is True
            assert role_report["exact_geometry_lock"]["new_space_identity_passed"] is True
            assert {key: role_report["exact_geometry_lock"][key] for key in expected_geometry} == expected_geometry
            assert role_report["space_mapping"]["count"] == 159
            assert role_report["space_mapping"]["ambiguous_count"] == 0
            assert role_report["schedule_lock"]["passed"] is True
            assert role_report["schedule_lock"]["changed_schedule_objects"] == 0
            assert role_report["schedule_lock"]["changed_protected_schedule_references"] == 0
            assert role_report["serialized_roundtrip_validation"]["passed"] is True
            repairs = role_report.get("energyplus_compatibility_repairs") or {"count": 0, "lossless": True}
            assert repairs.get("lossless") is True
            before = behavior_inventory(compiler, reference_paths[role])
            after = behavior_inventory(compiler, output_paths[role])
            assert after == before, f"{role} schedules, loads, or HVAC/system inventory changed"
            compiled_model = compiler.parse_osm(output_paths[role])
            invalid_special_ground = []
            for surface in compiled_model.by_type.get("OS:Surface", []):
                if len(surface.fields) <= 5 or surface.fields[5] != "GroundFCfactorMethod":
                    continue
                handle = surface.fields[3] if len(surface.fields) > 3 else ""
                construction = compiled_model.by_handle.get(handle) if handle else None
                if construction is None or construction.obj_type not in {
                    "OS:Construction:FfactorGroundFloor", "OS:Construction:CfactorUndergroundWall",
                }:
                    invalid_special_ground.append(surface.name)
            assert not invalid_special_ground, (
                f"{role} retained GroundFCfactorMethod without an explicit F/C-factor "
                f"construction: {invalid_special_ground[:10]}"
            )
            assert sum(after["schedules"].values()) == expected_schedules[role]
            assert sum(after["systems"].values()) == expected_systems[role]
            pipeline.stamp_compiled_project_identity(output_paths[role], identity, role.upper(), log)

        baseline_run_dir = folder / "03_SIMULATION" / "BASELINE"
        proposed_run_dir = folder / "03_SIMULATION" / "PROPOSED"
        baseline_run_dir.mkdir(parents=True)
        proposed_run_dir.mkdir(parents=True)
        baseline_osm = output_paths["baseline"]
        proposed_osm = output_paths["proposed"]
        runs = []
        for run_dir in (baseline_run_dir, proposed_run_dir):
            html = run_dir / "eplustbl.html"
            idf = run_dir / "in.idf"
            html.write_text("<html><head><title>CURRENT TEST PROJECT</title></head><body><p>Building: CURRENT TEST PROJECT</p>" + ("current simulation " * 40) + "</body></html>", encoding="utf-8")
            idf.write_text("Version,24.2;\n" + ("!- current model\n" * 40), encoding="utf-8")
            (run_dir / "eplusout.err").write_text("Program Version,EnergyPlus, Version 24.2\n************* EnergyPlus Completed Successfully-- 0 Warning; 0 Severe Errors; Elapsed Time=00hr 00min 01.00sec", encoding="utf-8")
            (run_dir / "eplusout.end").write_text("EnergyPlus Completed Successfully-- 0 Warning; 0 Severe Errors;", encoding="utf-8")
            runs.append({"folder": run_dir, "html": html, "idf": idf})

        packager = pipeline.load_module(pipeline.PACKAGER, "revex_record_review_packager_qa")
        review_dir = folder / "04_REVIEW_PACKAGE"
        review_zip = Path(packager.generate_package(
            str(runs[0]["html"]), str(runs[1]["html"]), str(review_dir),
            z_project["title"], True, standard_version="NYCECC 2020",
            baseline_model_file=str(runs[0]["idf"]), proposed_model_file=str(runs[1]["idf"]),
        ))
        expected_review_names = {f"{z_project['title']} - {label}.pdf" for label in pipeline.REVIEW_PACKAGE_PDF_LABELS}
        with zipfile.ZipFile(review_zip) as package:
            assert {Path(name).name for name in package.namelist() if not name.endswith("/")} == expected_review_names

        en1_xlsx = mock_en1_workbook(folder)
        completion = pipeline.validate_completion_outputs(
            baseline_osm, proposed_osm, runs[0], runs[1], review_zip, z_project["title"],
            en1_xlsx, cxl, report, official
        )
        assert completion["compiledOsmCount"] == 2
        assert completion["officialDoeReport"] is True
        assert completion["reviewPackagePdfCount"] == 9

        approved_metrics = {
            "modeledSquareFeet": pipeline.APPROVED_RUN_PROFILE["modeledSquareFeet"],
            "conditionedSquareFeet": pipeline.APPROVED_RUN_PROFILE["conditionedSquareFeet"],
        }
        for role in ("baseline", "proposed"):
            expected = pipeline.APPROVED_RUN_PROFILE["roles"][role]
            approved_metrics[role] = {
                "siteKbtu": expected["siteKbtu"],
                "siteEuiKbtuPerFt2": expected["siteEuiKbtuPerFt2"],
                "electricKwh": expected["electricKwh"],
                "gasTherm": expected["gasTherm"],
                "unmetHours": expected["unmetHours"],
                "cost": 40402.0 if role == "baseline" else 25556.0,
                "endUses": {label: {"sharePct": value} for label, value in expected["endUseSharePct"].items()},
            }
        approved_comparison = pipeline.compare_approved_run_profile(
            approved_metrics, compiled / "COMPILATION_AUDIT.json", baseline_osm
        )
        assert approved_comparison["status"] == "PASSED", approved_comparison
        assert approved_comparison["iterationSelection"] == "BEST_WORKING_ITERATION"
        regressed_metrics = json.loads(json.dumps(approved_metrics))
        regressed_metrics["proposed"]["siteKbtu"] *= 1.20
        regressed_metrics["proposed"]["siteEuiKbtuPerFt2"] *= 1.20
        regression = pipeline.compare_approved_run_profile(
            regressed_metrics, compiled / "COMPILATION_AUDIT.json", baseline_osm
        )
        assert regression["status"] == "REGRESSION", regression
        assert regression["reviewEligible"] is False

        geometry_dir = folder / "01_ORIGINAL_MODELS"
        geometry_dir.mkdir()
        geometry_osm = geometry_dir / "REVIT_GEOMETRY_ORIGINAL.osm"
        shutil.copy2(baseline_osm, geometry_osm)
        en1_xlsx = folder / "EN-1_READY_TO_INSERT.xlsx"
        review_workbook = load_workbook(pipeline.EN1_TEMPLATE)
        review_info = review_workbook["1,2,3 Information"]
        pipeline.blank_en1_identity_fields(review_info)
        pipeline.apply_en1_project_identity(review_info, identity)
        review_workbook.save(en1_xlsx)
        review_workbook.close()
        pipeline.assert_no_reference_identity_xlsx(en1_xlsx)
        review_artifacts = []
        for path, kind, review_name in (
            (geometry_osm, "geometry-osm", "GEOMETRY.osm"),
            (baseline_osm, "baseline-osm", "BASELINE.osm"),
            (proposed_osm, "proposed-osm", "PROPOSED.osm"),
            (runs[0]["html"], "baseline-html", "BASELINE_REPORT.html"),
            (runs[1]["html"], "proposed-html", "PROPOSED_REPORT.html"),
            (en1_xlsx, "en1-spreadsheet", "EN-1.xlsx"),
            (report, "official-comcheck-pdf", "COMcheck_BACKSTOP.pdf"),
            (review_zip, "packager-reports-archive", "PACKAGER_REPORTS.zip"),
        ):
            artifact = pipeline.relative_artifact(path, folder, kind)
            artifact["reviewName"] = review_name
            review_artifacts.append(artifact)
        assert tuple(artifact["kind"] for artifact in review_artifacts) == pipeline.VALID_ENERGY_REVIEW_PACKAGE
        manual_package = pipeline.create_manual_review_package(folder, review_artifacts, {
            "schema": "liber.revex.energy-manual-review-package.v1",
            "projectName": z_project["title"],
            "referenceTemplatesIncluded": False,
            "referenceIdentityExcluded": True,
            "files": review_artifacts,
        })
        with zipfile.ZipFile(manual_package) as package:
            names = {name for name in package.namelist() if not name.endswith("/")}
            assert len(names) == 8
            assert names == {artifact["reviewName"] for artifact in review_artifacts}
            assert all("/" not in name and "\\" not in name for name in names)
            index = json.loads(package.comment)
            assert index["referenceTemplatesIncluded"] is False
            assert index["referenceIdentityExcluded"] is True

    print({
        "activeRevitTzIdentityOnly": True,
        "applicantBlank": True,
        "modelerBlank": True,
        "referenceIdentityExcluded": True,
        "currentEnTechnicalFacts": True,
        "compiledModelIdentityFromActiveRevitTzEvidence": True,
        "comcheckInputReady": True,
        "officialBackstopClient": True,
        "missingConsentRejected": True,
        "wrongRevisionConsentRejected": True,
        "exactRevisionConsentAccepted": True,
        "compiledOsmCompletionGate": True,
        "approvedRecordGeometryCounts": {"spaces": 159, "surfaces": 1930, "subsurfaces": 294, "thermalZones": 2},
        "baselineScheduleObjects": 243,
        "proposedScheduleObjects": 89,
        "baselineHvacSystemObjects": 13,
        "proposedHvacSystemObjects": 18,
        "loadsAndSystemsInventoryPreserved": True,
        "approvedNinePdfReviewFormat": True,
        "maskedApprovedRunComparison": True,
        "referenceRegressionWithholdsReviewCandidate": True,
        "userReviewContractItems": 8,
        "productionUserArtifactsExactlySevenFilesPlusOneArchive": True,
        "manualReviewPackageBesideRun": True,
        "manualReviewPackageExcludesProtectedReferences": True,
        "filingSheets": len(pipeline.EN1_PRINT_SHEETS),
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
