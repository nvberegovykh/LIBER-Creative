#!/usr/bin/env python3
"""REVEX Companion-side energy package orchestrator.

This process is deliberately outside Revit. It consumes an approved Engineering
Sync revision and never calls Revit, changes sheets, or modifies a printing set.
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import glob
import hashlib
import importlib.metadata
import importlib.util
import json
import os
from pathlib import Path
import platform
import re
import shutil
import subprocess
import sys
import time
import traceback
import uuid
import zipfile
import xml.etree.ElementTree as ET

from comcheck_backstop import ComcheckBackstopError, run_official_backstop


SCHEMA = "liber.revex.energy-result.v1"
PIPELINE_VERSION = "0.8.19-r49"
COMCHECK_CONSENT_SCHEMA = "liber.revex.comcheck-consent.v1"
COMCHECK_SERVICE = "PNNL_COMCHECK_BACKSTOP"
COMCHECK_ENDPOINT = "https://legacy-comcheck.energycode.pnl.gov/CheckWeb/"
COMCHECK_SCOPE = "GENERATED_CURRENT_PROJECT_CXL_ONLY"
ROOT = Path(__file__).resolve().parent
REFERENCES = Path(os.environ.get("REVEX_ENERGY_REFERENCES", str(ROOT / "References"))).resolve()
GEOMETRYCO = ROOT / "GeometryCo" / "OpenStudio_Energy_Model_Geometry_Compiler.py"
PACKAGER = ROOT / "Packager" / "EnergyPlusReviewPackager.py"
GBXML_TO_OSM = ROOT / "gbxml_to_osm.rb"
EN1_TEMPLATE = REFERENCES / "EN-1_79_WINTHROP_AMENDMENT.xlsx"
COMCHECK_CXL_TEMPLATE = REFERENCES / "COMcheck_250_MIDWOOD_STRUCTURE_REFERENCE.cxl"
BASELINE_REFERENCE = REFERENCES / "79_WINTHROP_APPROVED_BASELINE.osm"
PROPOSED_REFERENCE = REFERENCES / "79_WINTHROP_APPROVED_PROPOSED.osm"
EN1_PRINT_SHEETS = (
    "1,2,3 Information", "3d. LL97 GHG Summary", "4. Compliance",
    "5a. Baseline Rotations", "5b. Usage Summary", "6a. Ext. Wall Areas",
    "6b. Fenestration", "6c1. Wall Types", "6c2. Wall Types - Addl Rows",
    "6d.1 Interior LPD-Space Method", "6d.2 Interior LPD-Bldg  Method",
    "6e. Ext LPD", "6f. Process Equip.", "6g. Service HW", "HVAC_Cover",
    "HVAC Air-side",
)
EN1_IDENTITY_FIELDS = {
    # Section 2a - Applicant Information. Keep the form label; clear the value.
    "B11": "Last Name ", "D11": "First Name ", "I11": "Middle Initial ",
    "B12": "Business Name ", "F12": "Business Email ",
    "B13": "Business Address ", "F13": "Business Telephone ",
    "B14": "City  ", "C14": "State ", "E14": "Zip  ",
    "B15": "Email ", "G15": "License Number ",
    # Section 2b - Modeling Firm / lead modeler.
    "B19": "Last Name (lead modeler) ", "D19": "First Name (lead modeler) ",
    "I19": "Middle Initial ", "B20": "Business Name ",
    "F20": "Email (lead modeler) ", "B21": "Business Address ",
    "F21": "Telephone (lead modeler) ", "B22": "City  ",
    "C22": "State ", "E22": "Zip  ",
}
EN1_PROJECT_FIELDS = {
    "B5": ("House No(s)   ", "houseNumber"),
    "C5": ("Street Name ", "streetName"),
    "B6": ("Borough  ", "borough"),
    "C6": ("Block ", "block"),
    "D6": ("Lot ", "lot"),
    "E6": ("BIN  ", "bin"),
    "G6": ("CB No. ", "communityBoard"),
    "C7": ("", "jobType"),
    "C38": ("", "architecturalJobNumber"),
    "C39": ("", "mechanicalJobNumber"),
    "C40": ("", "plumbingJobNumber"),
}
REQUIRED_PROJECT_IDENTITY = ("title", "address", "city", "state", "zip")
OPTIONAL_PROJECT_IDENTITY = (
    "houseNumber", "streetName", "borough", "block", "lot", "bin",
    "communityBoard", "jobType", "architecturalJobNumber",
    "mechanicalJobNumber", "plumbingJobNumber",
)
REFERENCE_IDENTITY_TOKENS = (
    "WINTHROP", "FAYBYSHENKO", "CHOSEN MEP", "BEREGOVYKH",
    "2306 OCEAN", "B01304513", "31-00 47TH", "091045",
)
REVIEW_PACKAGE_PDF_LABELS = (
    "Document Index",
    "Envelope Performance_Windows",
    "Envelope Performance (Opaque)",
    "Exterior Lighting Calculations",
    "Interior Lighting Calculations",
    "Proposed HVAC",
    "Baseline HVAC",
    "Proposed Reports",
    "Baseline Reports",
)
VALID_ENERGY_REVIEW_PACKAGE = (
    "geometry-osm",
    "baseline-osm",
    "proposed-osm",
    "baseline-html",
    "proposed-html",
    "en1-spreadsheet",
    "official-comcheck-pdf",
    "packager-reports-archive",
)

# Identity-free numerical fingerprint of the approved EnergyPlus review record.
# The structure templates remain protected worker inputs; only these verification
# facts are exposed in run evidence and the manual-review package.
APPROVED_RUN_PROFILE = {
    "schema": "liber.revex.approved-energy-profile.v1",
    "cohort": "MASKED_APPROVED_COHORT_A",
    "referenceIdentity": "MASKED",
    "topology": {"spaces": 159, "surfaces": 1930, "subsurfaces": 294, "thermalZones": 2},
    "modeledSquareFeet": 11690.791,
    "conditionedSquareFeet": 11169.925,
    "roles": {
        "baseline": {
            "siteKbtu": 942537.659,
            "siteEuiKbtuPerFt2": 80.623,
            "electricKwh": 133325.0,
            "gasTherm": 4876.0,
            "unmetHours": 28.33,
            "endUseSharePct": {
                "Space Heat": 37.1, "Space Cool": 7.5, "Interior Lighting": 10.8,
                "Exterior Lighting": 0.1, "Misc. Equip. Unregulated": 1.0,
                "Vent Fans": 26.8, "Pumps & Misc": 2.0, "Dom. Hot Water": 14.7,
            },
        },
        "proposed": {
            "siteKbtu": 348806.134,
            "siteEuiKbtuPerFt2": 29.836,
            "electricKwh": 102225.0,
            "gasTherm": 0.0,
            "unmetHours": 70.33,
            "endUseSharePct": {
                "Space Heat": 24.7, "Space Cool": 6.6, "Interior Lighting": 3.3,
                "Exterior Lighting": 0.3, "Misc. Equip. Unregulated": 2.8,
                "Vent Fans": 17.3, "Pumps & Misc": 5.2, "Dom. Hot Water": 39.8,
            },
        },
    },
    "tolerances": {
        "modeledAreaRelativePct": 0.25,
        "energyRelativePct": 5.0,
        "endUseSharePercentagePoints": 3.0,
        "unmetHoursRelativePct": 25.0,
        "unmetHoursAbsolute": 5.0,
        "nearZeroGasTherm": 5.0,
    },
}


class PipelineError(RuntimeError):
    pass


class RunLog:
    def __init__(self, folder: Path, correlation_id: str, initiator: str):
        self.path = folder / "REVEX-ENERGY-PIPELINE.jsonl"
        self.events: list[dict] = []
        self.dependencies: list[dict] = []
        self.correlation_id = correlation_id
        self.initiator = initiator
        self.started = time.monotonic()
        self.current_stage = "BOOTSTRAP"
        self._event_index = 0

    def write(self, stage: str, status: str, **detail) -> None:
        self.current_stage = stage
        self._event_index += 1
        row = {
            "eventId": f"{self.correlation_id}:{self._event_index:04d}",
            "at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "elapsedMs": round((time.monotonic() - self.started) * 1000),
            "correlationId": self.correlation_id,
            "initiator": self.initiator,
            "stage": stage,
            "status": status,
            **detail,
        }
        self.events.append(row)
        with self.path.open("a", encoding="utf-8") as stream:
            stream.write(json.dumps(row, ensure_ascii=False) + "\n")
        print(json.dumps(row, ensure_ascii=False), flush=True)

    def dependency(self, name: str, available: bool, *, required: bool = True, **detail) -> None:
        row = {"name": name, "available": available, "required": required, **detail}
        self.dependencies.append(row)
        self.write("DEPENDENCY", "AVAILABLE" if available else ("MISSING" if required else "UNAVAILABLE_OPTIONAL"), **row)


def exception_chain(ex: BaseException) -> list[dict]:
    output: list[dict] = []
    current: BaseException | None = ex
    seen: set[int] = set()
    while current is not None and id(current) not in seen and len(output) < 16:
        seen.add(id(current))
        output.append({"depth": len(output), "type": type(current).__name__, "message": str(current)})
        current = current.__cause__ or current.__context__
    return output


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_name(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip()).strip("._")
    return text or "REVEX_Energy"


def run_command(command: list[str], cwd: Path, log_path: Path, log: RunLog, stage: str) -> None:
    started = time.monotonic()
    log.write(stage, "STARTED", command=command, cwd=str(cwd), log=str(log_path))
    process = subprocess.Popen(command, cwd=str(cwd), text=True, encoding="utf-8", errors="replace",
                               stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    log.write(stage, "PROCESS_STARTED", pid=process.pid)
    stdout, _ = process.communicate()
    log_path.write_text(stdout or "", encoding="utf-8")
    elapsed_ms = round((time.monotonic() - started) * 1000)
    tail = (stdout or "").splitlines()[-24:]
    if process.returncode != 0:
        log.write(stage, "FAILED", pid=process.pid, exitCode=process.returncode, elapsedMs=elapsed_ms,
                  log=str(log_path), outputTail=tail)
        raise PipelineError(f"{stage} failed with exit code {process.returncode}; see {log_path.name}")
    log.write(stage, "PASSED", pid=process.pid, exitCode=process.returncode, elapsedMs=elapsed_ms,
              log=str(log_path), outputTail=tail[-4:])


def _openstudio_cli_version(path: Path) -> str:
    try:
        probe = subprocess.run([str(path), "--version"], capture_output=True, text=True,
                               encoding="utf-8", errors="replace", timeout=20)
        text = ((probe.stdout or "") + "\n" + (probe.stderr or "")).strip()
        match = re.search(r"(?<!\d)(\d+\.\d+(?:\.\d+)?)(?!\d)", text)
        return match.group(1) if match else "unknown"
    except Exception:
        return "unknown"


def find_openstudio(requested: str = "") -> Path:
    """Resolve the OpenStudio SDK 3.10 CLI used by the approved PRM templates.

    OpenStudio Application 1.10 commonly nests openstudio-3.10.x under the
    application directory, so discovery must not assume the older flat SDK path.
    An explicit user override is still honored, but a known wrong SDK is rejected.
    """
    candidates: list[Path] = []
    for value in [requested, os.environ.get("OPENSTUDIO_CLI", ""),
                  shutil.which("openstudio") or "", shutil.which("openstudio.exe") or ""]:
        if value:
            candidates.append(Path(value))

    roots = [
        Path(os.environ.get("ProgramFiles", r"C:\Program Files")),
        Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")),
        Path(os.environ.get("LOCALAPPDATA", "")) if os.environ.get("LOCALAPPDATA") else None,
        Path("C:/"),
    ]
    patterns = [
        "OpenStudio*/bin/openstudio.exe",
        "openstudio*/bin/openstudio.exe",
        "OpenStudio*/openstudio.exe",
        "openstudio*/openstudio.exe",
        "OpenStudioApplication*/bin/openstudio.exe",
        "openstudioapplication*/bin/openstudio.exe",
        "OpenStudioApplication*/openstudio-*/bin/openstudio.exe",
        "openstudioapplication*/openstudio-*/bin/openstudio.exe",
        "Programs/OpenStudio*/bin/openstudio.exe",
        "Programs/OpenStudioApplication*/openstudio-*/bin/openstudio.exe",
    ]
    for root in [r for r in roots if r is not None]:
        try:
            for pattern in patterns:
                candidates.extend(root.glob(pattern))
        except Exception:
            pass

    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        try:
            path = candidate.resolve()
        except Exception:
            path = candidate
        if not path.is_file():
            continue
        key = os.path.normcase(str(path))
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)

    inspected = [(path, _openstudio_cli_version(path)) for path in unique]
    exact = [(path, version) for path, version in inspected if version.startswith("3.10.") or version == "3.10"]
    if exact:
        # Explicit path wins when it is the correct SDK; otherwise prefer a path
        # whose installation name also advertises 3.10.
        if requested:
            req = os.path.normcase(str(Path(requested).resolve()))
            for path, _ in exact:
                if os.path.normcase(str(path)) == req:
                    return path
        exact.sort(key=lambda item: ("3.10" in str(item[0]), str(item[0])), reverse=True)
        return exact[0][0]

    if inspected:
        found = "; ".join(f"{version} at {path}" for path, version in inspected)
        raise PipelineError(
            "OpenStudio SDK 3.10 was not found. The approved Baseline/Proposed templates require OpenStudio 3.10.0; "
            f"detected: {found}. Install OpenStudio Application 1.10 / SDK 3.10 or set OPENSTUDIO_CLI to its openstudio.exe."
        )
    raise PipelineError(
        "OpenStudio SDK 3.10 CLI was not found. Install OpenStudio Application 1.10 / SDK 3.10 or set OPENSTUDIO_CLI to its openstudio.exe."
    )


def validate_epw(weather: Path) -> dict:
    try:
        parts = [part.strip() for part in weather.open("r", encoding="utf-8", errors="ignore").readline().strip().split(",")]
    except OSError as ex:
        raise PipelineError(f"The selected EPW could not be read: {ex}")
    if len(parts) < 10 or parts[0].upper() != "LOCATION" or not parts[1]:
        raise PipelineError("Weather file is not a valid EPW with a LOCATION header.")
    try:
        latitude, longitude, timezone, elevation = map(float, parts[6:10])
    except Exception as ex:
        raise PipelineError("Weather EPW LOCATION coordinates/time-zone/elevation are invalid.") from ex
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180 and -14 <= timezone <= 14):
        raise PipelineError("Weather EPW LOCATION metadata is out of range.")
    return {"city":parts[1],"stateProvince":parts[2],"country":parts[3],"dataSource":parts[4],"wmo":parts[5],"latitude":latitude,"longitude":longitude,"timeZone":timezone,"elevation":elevation}


def load_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise PipelineError(f"Could not load {path}")
    module = importlib.util.module_from_spec(spec)
    # Dataclasses and other runtime type resolvers require the module to be
    # registered while its class bodies execute. Without this, GeometryCo can
    # fail before compilation with a misleading dataclass AttributeError.
    sys.modules[name] = module
    try:
        spec.loader.exec_module(module)
    except Exception:
        sys.modules.pop(name, None)
        raise
    return module


def convert_gbxml(gbxml: Path, geometry_osm: Path, cli: Path, folder: Path, log: RunLog) -> None:
    run_command([str(cli), str(GBXML_TO_OSM), str(gbxml), str(geometry_osm)], folder,
                folder / "01_GBXML_TO_OSM.log", log, "GBXML_TO_OSM")
    if not geometry_osm.is_file() or geometry_osm.stat().st_size < 1024:
        raise PipelineError("OpenStudio returned without a usable geometry OSM.")


def compile_models(geometry_osm: Path, cli: Path, folder: Path, log: RunLog) -> tuple[Path, Path]:
    outdir = folder / "02_COMPILED_MODELS"
    outdir.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable, str(GEOMETRYCO),
        "--geometry", str(geometry_osm),
        "--baseline", str(BASELINE_REFERENCE),
        "--proposed", str(PROPOSED_REFERENCE),
        "--outdir", str(outdir),
        "--openstudio-cli", str(cli),
        "--require-native-check",
    ]
    run_command(command, folder, folder / "02_GEOMETRYCO.log", log, "GEOMETRYCO_4_3_1")
    baseline = outdir / "BASELINE_UPDATED_GEOMETRY.osm"
    proposed = outdir / "PROPOSED_UPDATED_GEOMETRY.osm"
    if not baseline.is_file() or not proposed.is_file():
        raise PipelineError("GeometryCo did not atomically commit both compiled OSMs.")
    return baseline, proposed


def simulate(role: str, model: Path, weather: Path, cli: Path, folder: Path, log: RunLog) -> dict:
    role_dir = folder / "03_SIMULATION" / role.upper()
    role_dir.mkdir(parents=True, exist_ok=True)
    osw = role_dir / "workflow.osw"
    osw.write_text(json.dumps({
        "seed_file": str(model.resolve()),
        "weather_file": str(weather.resolve()),
        "run_options": {"skip_zip_results": True},
        "steps": [],
    }, indent=2), encoding="utf-8")
    run_command([str(cli), "run", "-w", str(osw)], role_dir, role_dir / "REVEX_OPENSTUDIO_RUN.log",
                log, f"SIMULATE_{role.upper()}")
    html = next(iter(role_dir.glob("**/eplustbl.html")), None)
    idf = next(iter(role_dir.glob("**/in.idf")), None)
    err = next(iter(role_dir.glob("**/eplusout.err")), None)
    if html is None or idf is None:
        raise PipelineError(f"{role.title()} simulation completed without eplustbl.html and in.idf.")
    return {"html": html, "idf": idf, "err": err, "folder": role_dir}


def end_use_rows(packager, soup) -> list[tuple[str, float, float, float]]:
    rows = []
    for label, elec_gj, gas_gj, other_gj in packager._end_uses_rows_gj(soup):
        rows.append((label, elec_gj * packager.GJ_TO_KWH,
                     gas_gj * packager.GJ_TO_THERM, other_gj * packager.GJ_TO_KBTU))
    return rows


def normalize_end_use(label: str) -> str:
    value = re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()
    rules = [
        ("interior lighting", "Interior Lighting"),
        ("interior equipment", "Misc. Equip. Unregulated"),
        ("heating", "Space Heat"),
        ("cooling", "Space Cool"),
        ("heat rejection", "Heat Rejection"),
        ("pumps", "Pumps & Misc"),
        ("fans", "Vent Fans"),
        ("water systems", "Dom. Hot Water"),
        ("exterior lighting", "Exterior Lighting"),
        ("exterior equipment", "Exterior Misc."),
    ]
    for token, target in rules:
        if token in value:
            return target
    return "Pumps & Misc"


def aggregate_end_uses(rows: list[tuple[str, float, float, float]]) -> dict[str, list[float]]:
    output: dict[str, list[float]] = {}
    for label, electric, gas, other in rows:
        key = normalize_end_use(label)
        target = output.setdefault(key, [0.0, 0.0, 0.0])
        target[0] += electric
        target[1] += gas
        target[2] += other
    return output


def summarize_end_uses(rows: dict[str, list[float]]) -> dict[str, dict[str, float]]:
    """Return review-safe fuel totals and site-energy shares for each end use."""
    raw = {}
    total_site_kbtu = 0.0
    for label, values in rows.items():
        electric_kwh, gas_therm, other_kbtu = (float(value or 0) for value in values)
        site_kbtu = electric_kwh * 3.412141633 + gas_therm * 100.0 + other_kbtu
        raw[label] = {
            "electricKwh": electric_kwh,
            "gasTherm": gas_therm,
            "otherKbtu": other_kbtu,
            "siteKbtu": site_kbtu,
        }
        total_site_kbtu += site_kbtu
    for values in raw.values():
        values["sharePct"] = 100.0 * values["siteKbtu"] / total_site_kbtu if total_site_kbtu else 0.0
    return raw


def _relative_delta_pct(actual: float, expected: float) -> float:
    if expected == 0:
        return 0.0 if actual == 0 else float("inf")
    return abs(actual - expected) * 100.0 / abs(expected)


def compare_approved_run_profile(metrics: dict, compilation_audit: Path, compiled_baseline_model: Path) -> dict:
    """Compare a matching model cohort with the masked approved EnergyPlus run.

    Topology and model area identify the cohort. Numerical simulation results are
    compared with bounded tolerances so harmless engine-version variance does not
    obscure meaningful regressions. No source-project identity enters the result.
    """
    profile = APPROVED_RUN_PROFILE
    tolerances = profile["tolerances"]
    audit = json.loads(compilation_audit.read_text(encoding="utf-8"))
    baseline_report = (audit.get("reports") or {}).get("baseline") or {}
    geometry = baseline_report.get("exact_geometry_lock") or {}
    compiler = load_module(GEOMETRYCO, "revex_geometryco_approved_profile_compare")
    compiled_model = compiler.parse_osm(compiled_baseline_model)
    actual_topology = {
        "spaces": int(geometry.get("spaces") or -1),
        "surfaces": int(geometry.get("surfaces") or -1),
        "subsurfaces": int(geometry.get("subsurfaces") or -1),
        "thermalZones": len(compiled_model.by_type.get("OS:ThermalZone", [])),
    }
    topology_matches = actual_topology == profile["topology"]
    modeled_area = float(metrics.get("modeledSquareFeet") or 0)
    area_delta = _relative_delta_pct(modeled_area, float(profile["modeledSquareFeet"]))
    cohort_matches = topology_matches and area_delta <= float(tolerances["modeledAreaRelativePct"])
    profile_digest = hashlib.sha256(
        json.dumps(profile, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    result = {
        "schema": "liber.revex.approved-run-comparison.v1",
        "cohort": profile["cohort"],
        "referenceIdentity": "MASKED",
        "profileSha256": profile_digest,
        "cohortMatches": cohort_matches,
        "topology": {"expected": profile["topology"], "actual": actual_topology, "passed": topology_matches},
        "modeledArea": {
            "expectedSquareFeet": profile["modeledSquareFeet"],
            "actualSquareFeet": modeled_area,
            "relativeDeltaPct": area_delta,
            "tolerancePct": tolerances["modeledAreaRelativePct"],
            "passed": area_delta <= float(tolerances["modeledAreaRelativePct"]),
        },
        "checks": [],
        "iterationSelection": "UNBENCHMARKED_DIFFERENT_COHORT",
        "reviewEligible": True,
    }
    if not cohort_matches:
        result["status"] = "NOT_APPLICABLE_DIFFERENT_COHORT"
        return result

    def add_relative(role: str, field: str, actual: float, expected: float, tolerance: float) -> None:
        delta = _relative_delta_pct(actual, expected)
        result["checks"].append({
            "role": role, "metric": field, "actual": actual, "expected": expected,
            "delta": delta, "tolerance": tolerance, "units": "relative-percent", "passed": delta <= tolerance,
        })

    def add_absolute(role: str, field: str, actual: float, expected: float, tolerance: float, units: str) -> None:
        delta = abs(actual - expected)
        result["checks"].append({
            "role": role, "metric": field, "actual": actual, "expected": expected,
            "delta": delta, "tolerance": tolerance, "units": units, "passed": delta <= tolerance,
        })

    for role in ("baseline", "proposed"):
        actual_role = metrics.get(role) or {}
        expected_role = profile["roles"][role]
        add_relative(role, "siteKbtu", float(actual_role.get("siteKbtu") or 0),
                     float(expected_role["siteKbtu"]), float(tolerances["energyRelativePct"]))
        add_relative(role, "siteEuiKbtuPerFt2", float(actual_role.get("siteEuiKbtuPerFt2") or 0),
                     float(expected_role["siteEuiKbtuPerFt2"]), float(tolerances["energyRelativePct"]))
        add_relative(role, "electricKwh", float(actual_role.get("electricKwh") or 0),
                     float(expected_role["electricKwh"]), float(tolerances["energyRelativePct"]))
        if float(expected_role["gasTherm"]) == 0:
            add_absolute(role, "gasTherm", float(actual_role.get("gasTherm") or 0), 0.0,
                         float(tolerances["nearZeroGasTherm"]), "therm")
        else:
            add_relative(role, "gasTherm", float(actual_role.get("gasTherm") or 0),
                         float(expected_role["gasTherm"]), float(tolerances["energyRelativePct"]))
        unmet_tolerance = max(float(tolerances["unmetHoursAbsolute"]),
                              float(expected_role["unmetHours"]) * float(tolerances["unmetHoursRelativePct"]) / 100.0)
        add_absolute(role, "unmetHours", float(actual_role.get("unmetHours") or 0),
                     float(expected_role["unmetHours"]), unmet_tolerance, "hours")
        actual_end_uses = actual_role.get("endUses") or {}
        for label, expected_share in expected_role["endUseSharePct"].items():
            actual_share = float((actual_end_uses.get(label) or {}).get("sharePct") or 0)
            add_absolute(role, f"endUseShare:{label}", actual_share, float(expected_share),
                         float(tolerances["endUseSharePercentagePoints"]), "percentage-points")

    baseline = metrics.get("baseline") or {}
    proposed = metrics.get("proposed") or {}
    directional = [
        {
            "role": "pair", "metric": "proposedSiteEnergyBelowBaseline",
            "actual": float(proposed.get("siteKbtu") or 0), "expected": f"< {float(baseline.get('siteKbtu') or 0)}",
            "passed": 0 < float(proposed.get("siteKbtu") or 0) < float(baseline.get("siteKbtu") or 0),
        },
        {
            "role": "pair", "metric": "proposedVirtualCostBelowBaseline",
            "actual": float(proposed.get("cost") or 0), "expected": f"< {float(baseline.get('cost') or 0)}",
            "passed": 0 < float(proposed.get("cost") or 0) < float(baseline.get("cost") or 0),
        },
    ]
    result["checks"].extend(directional)
    passed = all(check.get("passed") is True for check in result["checks"])
    finite_ratios = [
        min(float(check["delta"]) / float(check["tolerance"]), 10.0)
        for check in result["checks"]
        if isinstance(check.get("delta"), (int, float)) and check.get("tolerance") not in (None, 0)
    ]
    result["normalizedRegressionScore"] = (
        sum(finite_ratios) / len(finite_ratios) if finite_ratios else (0.0 if passed else 10.0)
    )
    result["passedChecks"] = sum(1 for check in result["checks"] if check.get("passed") is True)
    result["totalChecks"] = len(result["checks"])
    result["status"] = "PASSED" if passed else "REGRESSION"
    result["iterationSelection"] = "BEST_WORKING_ITERATION" if passed else "WITHHELD_REFERENCE_REGRESSION"
    result["reviewEligible"] = passed
    return result


def find_area_ft2(packager, soup, labels: tuple[str, ...]) -> float | None:
    hits = packager.find_tables_by_keywords(soup, ["building area"])
    for _, table in hits:
        for row in packager.table_to_matrix(table):
            if not row:
                continue
            label = packager.norm(row[0]).lower()
            if not any(token in label for token in labels):
                continue
            for raw in row[1:]:
                value = packager.parse_number(raw)
                if value is not None and value > 0:
                    # EnergyPlus Building Area is normally m2.
                    return float(value) * 10.7639104167
    return None


def find_unmet_hours(packager, soup) -> float | None:
    for table in soup.find_all("table"):
        for row in packager.table_to_matrix(table):
            label = " ".join(packager.norm(cell) for cell in row).lower()
            if "time setpoint not met during occupied" not in label and "facility any zone hours" not in label:
                continue
            values = [packager.parse_number(cell) for cell in row[1:]]
            nums = [float(value) for value in values if value is not None]
            if nums:
                return max(nums)
    return None


def blank_en1_identity_fields(info) -> None:
    """Clear prior-project applicant/modeler values in a run copy, preserving labels/style."""
    for cell, blank_label in EN1_IDENTITY_FIELDS.items():
        info[cell] = blank_label


def assert_en1_identity_fields_blank(info) -> None:
    leaked = [cell for cell, blank_label in EN1_IDENTITY_FIELDS.items() if info[cell].value != blank_label]
    if leaked:
        raise PipelineError("EN-1 identity fields were not blank: " + ", ".join(leaked))


def load_page_facts(page_facts_path: Path | None) -> dict:
    if not page_facts_path or not page_facts_path.is_file():
        return {"status": "MISSING", "pages": [], "errors": ["Revit T/Z/EN page facts are unavailable"]}
    try:
        value = json.loads(page_facts_path.read_text(encoding="utf-8"))
        return value if isinstance(value, dict) else {"status": "FAILED", "pages": [], "errors": ["Page facts root is not an object"]}
    except Exception as exc:
        return {"status": "FAILED", "pages": [], "errors": [str(exc)]}


def _best_page_value(pages: list[dict], section: str, key: str):
    candidates = []
    for page in pages:
        value = (page.get(section) or {}).get(key)
        if value not in (None, ""):
            candidates.append((float(page.get("confidence") or 0), str(value).strip()))
    return sorted(candidates, key=lambda row: row[0], reverse=True)[0][1] if candidates else None


def current_project_identity(facts: dict) -> dict:
    """Resolve filing identity from the bound active document and its T/Z pages.

    Visible Z facts take precedence, T/title-sheet facts fill gaps, and the
    deterministic native Revit parameter capture fills only remaining gaps.
    Browser state, prior revisions, filenames and reference models are never read.
    """
    pages = list(facts.get("pages") or [])
    z_pages = [page for page in pages if str(page.get("pageType") or "").upper() == "Z"]
    t_pages = [page for page in pages if str(page.get("pageType") or "").upper() == "T"]
    native = dict(facts.get("structuredIdentity") or {})
    keys = (
        "title", "address", "city", "state", "zip", "houseNumber", "streetName", "borough",
        "block", "lot", "bin", "communityBoard", "jobType", "architecturalJobNumber",
        "mechanicalJobNumber", "plumbingJobNumber",
    )
    identity = {
        key: (_best_page_value(z_pages, "project", key)
              or _best_page_value(t_pages, "project", key)
              or native.get(key))
        for key in keys
    }
    address = str(identity.get("address") or "").strip()
    if address and (not identity.get("houseNumber") or not identity.get("streetName")):
        match = re.match(r"^\s*([0-9]+(?:-[0-9]+)?)\s+(.+?)\s*$", address)
        if match:
            identity["houseNumber"] = identity.get("houseNumber") or match.group(1)
            identity["streetName"] = identity.get("streetName") or match.group(2)
    if not address and (identity.get("houseNumber") or identity.get("streetName")):
        identity["address"] = " ".join(str(identity.get(key) or "").strip() for key in ("houseNumber", "streetName")).strip()
    if not identity.get("title"):
        identity["title"] = identity.get("address")
    if not identity.get("city"):
        identity["city"] = identity.get("borough")
    identity["source"] = "bound active Revit document plus immutable T/Z pages"
    identity["evidenceDigest"] = native.get("evidenceDigest")
    identity["evidenceSheets"] = list(native.get("evidenceSheets") or [])
    identity["missing"] = [key for key in REQUIRED_PROJECT_IDENTITY if not str(identity.get(key) or "").strip()]
    identity["optionalMissing"] = [key for key in OPTIONAL_PROJECT_IDENTITY if not str(identity.get(key) or "").strip()]
    return identity


def apply_en1_project_identity(info, identity: dict) -> None:
    """Clear template values, then populate only current active-document facts."""
    for cell, (label, key) in EN1_PROJECT_FIELDS.items():
        info[cell] = label + str(identity.get(key) or "").strip()


def assert_en1_project_identity(info, identity: dict) -> None:
    wrong = []
    for cell, (label, key) in EN1_PROJECT_FIELDS.items():
        expected = label + str(identity.get(key) or "").strip()
        if str(info[cell].value or "") != expected:
            wrong.append(cell)
    if wrong:
        raise PipelineError("EN-1 current-project identity verification failed: " + ", ".join(wrong))


def assert_no_reference_identity_text(value: str, artifact: str) -> None:
    upper = str(value or "").upper()
    leaked = sorted({token for token in REFERENCE_IDENTITY_TOKENS if token in upper})
    if leaked:
        raise PipelineError(f"{artifact} retained structure-template identity: " + ", ".join(leaked))


def stamp_compiled_project_identity(model_path: Path, project_identity: dict, role: str, log: RunLog) -> None:
    """Replace template project labels before native simulation.

    Approved OSMs contribute schedules, constructions, loads, and systems only.
    Their building/display labels must not survive into a current-project OSM,
    IDF, EnergyPlus report, or user-visible artifact.
    """
    missing = list(project_identity.get("missing") or [])
    if missing:
        raise PipelineError(
            "Current active-Revit project identity is incomplete; compiled models cannot be stamped: "
            + ", ".join(missing)
        )

    def osm_text(value: object) -> str:
        text = re.sub(r"[\r\n]+", " ", str(value or "")).strip()
        return text.replace("&", "&amp;").replace(",", "&#44;").replace(";", "&#59;")

    title = osm_text(project_identity.get("title") or project_identity.get("address"))
    address = osm_text(project_identity.get("address") or project_identity.get("title"))
    city = osm_text(project_identity.get("city"))
    state = osm_text(project_identity.get("state"))
    postal = osm_text(project_identity.get("zip"))
    display_name = "&#44; ".join(part for part in (address, city, state, postal) if part)

    compiler = load_module(GEOMETRYCO, "revex_geometryco_identity_stamp")
    model = compiler.parse_osm(model_path)
    buildings = [obj for obj in model.by_type.get("OS:Building", []) if obj.handle]
    if len(buildings) != 1:
        raise PipelineError(f"{role} compiled OSM must contain exactly one OS:Building for current-project identity stamping.")
    building = buildings[0]
    while len(building.fields) < 2:
        building.fields.append("")
    building.fields[1] = title

    display_fields = 0
    for props in model.by_type.get("OS:AdditionalProperties", []):
        if len(props.fields) < 2 or props.fields[1] != building.handle:
            continue
        for index in range(2, len(props.fields) - 2, 3):
            if str(props.fields[index] or "").strip().lower() == "displayname":
                props.fields[index + 2] = display_name
                display_fields += 1

    compiler.write_osm(model_path, model.objects)
    reparsed = compiler.parse_osm(model_path)
    current_building = reparsed.by_type.get("OS:Building", [None])[0]
    if current_building is None or len(current_building.fields) < 2 or current_building.fields[1] != title:
        raise PipelineError(f"{role} compiled OSM current-project identity stamp did not survive serialization.")
    assert_no_reference_identity_text(model_path.read_text(encoding="utf-8"), f"{role} compiled OSM")
    log.write(
        "COMPILED_MODEL_IDENTITY", "PASSED", role=role,
        source="bound active Revit document plus immutable T/Z pages", buildingName=title,
        displayNameFields=display_fields, model=model_path.name,
    )


def assert_no_reference_identity_workbook(workbook) -> None:
    hits = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                upper = cell.value.upper()
                if any(token in upper for token in REFERENCE_IDENTITY_TOKENS):
                    hits.append(f"{sheet.title}!{cell.coordinate}")
    if hits:
        raise PipelineError("EN-1 retained structure-template identity at: " + ", ".join(hits[:24]))


def assert_no_reference_identity_xlsx(path: Path) -> None:
    leaked = []
    with zipfile.ZipFile(path) as package:
        for name in package.namelist():
            raw = package.read(name).upper()
            for token in REFERENCE_IDENTITY_TOKENS:
                if token.encode("utf-8") in raw:
                    leaked.append(f"{name}:{token}")
    if leaked:
        raise PipelineError("EN-1 package retained structure-template identity: " + ", ".join(leaked[:24]))


def prepare_en1(packager, baseline_html: Path, proposed_html: Path,
                folder: Path, log: RunLog, weather_meta: dict, project_identity: dict) -> tuple[Path, Path | None, dict]:
    from openpyxl import load_workbook
    from openpyxl.workbook.views import BookView

    baseline_soup = packager.load_soup(str(baseline_html))
    proposed_soup = packager.load_soup(str(proposed_html))
    baseline_rows = aggregate_end_uses(end_use_rows(packager, baseline_soup))
    proposed_rows = aggregate_end_uses(end_use_rows(packager, proposed_soup))
    workbook_path = folder / "EN-1_READY_TO_INSERT.xlsx"
    shutil.copy2(EN1_TEMPLATE, workbook_path)
    workbook = load_workbook(workbook_path)
    info = workbook["1,2,3 Information"]
    blank_en1_identity_fields(info)
    apply_en1_project_identity(info, project_identity)
    assert_en1_identity_fields_blank(info)
    assert_en1_project_identity(info, project_identity)
    assert_no_reference_identity_workbook(workbook)

    baseline_total = [sum(values[i] for values in baseline_rows.values()) for i in range(3)]
    proposed_total = [sum(values[i] for values in proposed_rows.values()) for i in range(3)]
    baseline_cost = baseline_total[0] * packager.CONED_VIRTUAL_ELEC_RATE_PER_KWH + baseline_total[1] * packager.CONED_VIRTUAL_GAS_RATE_PER_THERM
    proposed_cost = proposed_total[0] * packager.CONED_VIRTUAL_ELEC_RATE_PER_KWH + proposed_total[1] * packager.CONED_VIRTUAL_GAS_RATE_PER_THERM

    rotations = workbook["5a. Baseline Rotations"]
    usage = workbook["5b. Usage Summary"]
    row_map = {
        "Interior Lighting": 9, "Misc. Equip. Unregulated": 10, "Misc. Equip. Regulated": 11,
        "Space Heat": 12, "Space Cool": 13, "Heat Rejection": 14, "Pumps & Misc": 15,
        "Vent Fans": 16, "Dom. Hot Water": 17, "Exterior Lighting": 18, "Exterior Misc.": 19,
    }
    usage_map = {key: row - 3 for key, row in row_map.items()}
    for key, row in row_map.items():
        values = baseline_rows.get(key, [0.0, 0.0, 0.0])
        rotations.cell(row, 3, round(values[0], 3))
        rotations.cell(row, 4, round(values[1], 3))
        rotations.cell(row, 5, round(values[2], 3))
    rotations["C23"] = round(baseline_total[0] * packager.CONED_VIRTUAL_ELEC_RATE_PER_KWH, 2)
    rotations["C24"] = round(baseline_total[1] * packager.CONED_VIRTUAL_GAS_RATE_PER_THERM, 2)
    for key, row in usage_map.items():
        values = proposed_rows.get(key, [0.0, 0.0, 0.0])
        usage.cell(row, 7, round(values[0], 3))
        usage.cell(row, 8, round(values[1], 3))
        usage.cell(row, 9, round(values[2], 3))

    compliance = workbook["4. Compliance"]
    compliance["G8"] = round(proposed_total[0] * packager.CONED_VIRTUAL_ELEC_RATE_PER_KWH, 2)
    compliance["G9"] = round(proposed_total[1] * packager.CONED_VIRTUAL_GAS_RATE_PER_THERM, 2)
    modeled = find_area_ft2(packager, proposed_soup, ("total building area", "gross floor area"))
    conditioned = find_area_ft2(packager, proposed_soup, ("net conditioned building area", "conditioned building area"))
    unconditioned = find_area_ft2(packager, proposed_soup, ("unconditioned building area",))
    if modeled:
        info["C31"] = round(modeled, 3)
        info["C32"] = round(modeled, 3)
        compliance["C29"] = round(modeled, 3)
    if conditioned:
        info["C33"] = round(conditioned, 3)
    if unconditioned is not None:
        info["C35"] = round(unconditioned, 3)
    baseline_unmet = find_unmet_hours(packager, baseline_soup)
    proposed_unmet = find_unmet_hours(packager, proposed_soup)
    if baseline_unmet is not None:
        info["C30"] = round(baseline_unmet, 2)
    if proposed_unmet is not None:
        info["C29"] = round(proposed_unmet, 2)
    info["C27"] = "EnergyPlus (native REVEX run; exact version in attached report)"
    info["C28"] = " ".join(str(weather_meta.get(k) or "").strip() for k in ("city","stateProvince","country","dataSource") if str(weather_meta.get(k) or "").strip())
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(workbook_path)
    assert_no_reference_identity_xlsx(workbook_path)
    verification_workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        assert_en1_identity_fields_blank(verification_workbook["1,2,3 Information"])
        assert_en1_project_identity(verification_workbook["1,2,3 Information"], project_identity)
        assert_no_reference_identity_workbook(verification_workbook)
    finally:
        verification_workbook.close()

    # The approved filing PDF is the 16-form-sheet set, not every visible helper,
    # instruction, or unused system sheet in the source workbook. Export through a
    # disposable print copy so the delivered workbook retains the template topology.
    target_dir = folder / "en1-export"
    target_dir.mkdir(exist_ok=True)
    print_workbook_path = target_dir / "EN-1_READY_TO_INSERT_PRINT.xlsx"
    shutil.copy2(workbook_path, print_workbook_path)
    print_workbook = load_workbook(print_workbook_path)
    printable = set(EN1_PRINT_SHEETS)
    missing_print_sheets = [name for name in EN1_PRINT_SHEETS if name not in print_workbook.sheetnames]
    if missing_print_sheets:
        raise PipelineError("EN-1 template is missing filing sheet(s): " + ", ".join(missing_print_sheets))
    for sheet in print_workbook.worksheets:
        sheet.sheet_state = "visible" if sheet.title in printable else "hidden"
        if sheet.title in printable:
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.scale = None
    active_index = print_workbook.sheetnames.index("1,2,3 Information")
    print_workbook.active = active_index
    print_workbook.views = [BookView(activeTab=active_index, firstSheet=active_index)]
    print_workbook.save(print_workbook_path)

    baseline_site_kbtu = baseline_total[0] * 3.412141633 + baseline_total[1] * 100.0 + baseline_total[2]
    proposed_site_kbtu = proposed_total[0] * 3.412141633 + proposed_total[1] * 100.0 + proposed_total[2]
    metrics = {
        "baseline": {
            "electricKwh": baseline_total[0], "gasTherm": baseline_total[1], "otherKbtu": baseline_total[2],
            "siteKbtu": baseline_site_kbtu,
            "siteEuiKbtuPerFt2": baseline_site_kbtu / modeled if modeled else None,
            "cost": baseline_cost, "unmetHours": baseline_unmet,
            "endUses": summarize_end_uses(baseline_rows),
        },
        "proposed": {
            "electricKwh": proposed_total[0], "gasTherm": proposed_total[1], "otherKbtu": proposed_total[2],
            "siteKbtu": proposed_site_kbtu,
            "siteEuiKbtuPerFt2": proposed_site_kbtu / modeled if modeled else None,
            "cost": proposed_cost, "unmetHours": proposed_unmet,
            "endUses": summarize_end_uses(proposed_rows),
        },
        "modeledSquareFeet": modeled, "conditionedSquareFeet": conditioned, "unconditionedSquareFeet": unconditioned,
        "identityFields": {
            "project": "ACTIVE_REVIT_T_Z_EVIDENCE", "applicant": "BLANK", "leadModeler": "BLANK",
            "missingProjectFields": list(project_identity.get("missing") or []),
        },
    }
    (folder / "EN-1_DATA_AUDIT.json").write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    pdf_path = folder / "EN-1_READY_TO_INSERT.pdf"
    export_error = None
    if os.name == "nt":
        powershell = shutil.which("powershell.exe") or shutil.which("powershell")
        log.dependency("Excel PDF PowerShell bridge", bool(powershell), required=False,
                       executable=powershell or "not found")
        if powershell:
            export_started = time.monotonic()
            result = subprocess.run([powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
                                     str(ROOT / "export_en1_pdf.ps1"), "-WorkbookPath", str(print_workbook_path),
                                     "-PdfPath", str(pdf_path)], capture_output=True, text=True)
            log.write("PREPARE_EN1_EXCEL_PDF", "PASSED" if result.returncode == 0 and pdf_path.is_file() else "FAILED",
                      executable=powershell, exitCode=result.returncode,
                      elapsedMs=round((time.monotonic() - export_started) * 1000),
                      outputTail=((result.stdout or "") + "\n" + (result.stderr or "")).splitlines()[-24:])
            if result.returncode != 0:
                export_error = (result.stdout or "") + "\n" + (result.stderr or "")
    if not pdf_path.is_file():
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        log.dependency("LibreOffice PDF fallback", bool(soffice), required=False,
                       executable=soffice or "not found")
        if soffice:
            profile_dir = folder / "libreoffice-profile"
            profile_dir.mkdir(exist_ok=True)
            export_started = time.monotonic()
            result = subprocess.run([soffice, "--headless", f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                                     "--convert-to", "pdf", "--outdir", str(target_dir), str(print_workbook_path)],
                                    capture_output=True, text=True)
            candidate = target_dir / (print_workbook_path.stem + ".pdf")
            log.write("PREPARE_EN1_LIBREOFFICE_PDF", "PASSED" if result.returncode == 0 and candidate.is_file() else "FAILED",
                      executable=soffice, exitCode=result.returncode,
                      elapsedMs=round((time.monotonic() - export_started) * 1000),
                      outputTail=((result.stdout or "") + "\n" + (result.stderr or "")).splitlines()[-24:])
            if result.returncode == 0 and candidate.is_file():
                shutil.move(candidate, pdf_path)
            else:
                export_error = (result.stdout or "") + "\n" + (result.stderr or "")
    if not pdf_path.is_file():
        log.write("PREPARE_EN1", "WORKBOOK_READY_PDF_BLOCKED", error=export_error or "Excel/LibreOffice unavailable")
        return workbook_path, None, metrics
    from pypdf import PdfReader
    page_count = len(PdfReader(str(pdf_path)).pages)
    if page_count != len(EN1_PRINT_SHEETS):
        pdf_path.unlink(missing_ok=True)
        raise PipelineError(
            f"EN-1 filing PDF must contain exactly {len(EN1_PRINT_SHEETS)} pages; export produced {page_count}."
        )
    metrics["filingPdfPages"] = page_count
    log.write("PREPARE_EN1", "PASSED", workbook=str(workbook_path), pdf=str(pdf_path), pages=page_count)
    return workbook_path, pdf_path, metrics


def _fact_value(pages: list[dict], section: str, key: str):
    return _best_page_value(pages, section, key)


def _set_xml(parent, local_name: str, value) -> None:
    if value is None: return
    child = next((c for c in list(parent) if c.tag.rsplit('}',1)[-1] == local_name), None)
    if child is not None:
        child.text = str(value)


def _replace_xml(parent, local_name: str, value) -> None:
    """Replace one current-project fact or remove the template fact entirely."""
    child = next((c for c in list(parent) if c.tag.rsplit('}',1)[-1] == local_name), None)
    if child is None:
        return
    if value is None or str(value).strip() == "":
        parent.remove(child)
    else:
        child.text = str(value)


def _orientation(value: str | None) -> str:
    token = re.sub(r"[^A-Z]", "", str(value or "").upper())
    mapping = {"N":"NORTH","NORTH":"NORTH","S":"SOUTH","SOUTH":"SOUTH","E":"EAST","EAST":"EAST","W":"WEST","WEST":"WEST"}
    return mapping.get(token, "NORTH")


def _wall_exemplar(walls: list, description: str | None):
    text = str(description or "").lower()
    wanted = "METAL_FRAME" if any(t in text for t in ("steel", "metal", "stud")) else "CONCRETE" if "concrete" in text else "MASONRY" if any(t in text for t in ("masonry","cmu","block")) else ""
    if wanted:
        for e in walls:
            wt = next((c.text or "" for c in list(e) if c.tag.rsplit('}',1)[-1] == "wallType"), "")
            if wanted in wt: return e
    return walls[0] if walls else None


def _clear_children_by_local(parent, local_name: str) -> None:
    for child in list(parent):
        if child.tag.rsplit('}',1)[-1] == local_name:
            parent.remove(child)


def _make_comcheck_audit_pdf(path: Path, project_name: str, facts: dict, status: str, missing: list[str], cxl_name: str | None) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(path), pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = [Paragraph("REVEX COMcheck Input Audit", styles["Title"]),
             Paragraph(project_name, styles["Heading2"]),
             Paragraph("This is a REVEX input/audit report generated from bound active-document identity and immutable Revit T/Z/EN sheets. It is not an official DOE COMcheck compliance report.", styles["BodyText"]), Spacer(1, 10)]
    rows = [["Status", status], ["CXL", cxl_name or "Not generated"], ["Missing/blocked fields", ", ".join(missing) if missing else "None"],
            ["AI scope", "Revit T/Z/EN sheet scanning only"], ["Geometry authority", "Revit/gbXML evidence graph; AI has none"]]
    table = Table(rows, colWidths=[150, 350])
    table.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.grey), ("VALIGN",(0,0),(-1,-1),"TOP")]))
    story += [table, Spacer(1, 12)]
    pages = list(facts.get("pages") or [])
    for page in pages:
        story.append(Paragraph(f"{page.get('pageType','')} {page.get('sheetNumber','')} — {page.get('sheetName','')}", styles["Heading3"]))
        story.append(Paragraph(f"Page confidence: {float(page.get('confidence') or 0):.0%}; source: {page.get('sourceFile','')}", styles["BodyText"]))
        envelope = [row for row in list(page.get("envelope") or []) if float(row.get("confidence") or 0) >= 0.90]
        if envelope:
            data = [["Kind","Assembly","Orient.","Area ft²","U","SHGC","R cavity","R cont."]]
            for row in envelope[:40]:
                data.append([row.get("kind"), row.get("assemblyType") or "", row.get("orientation") or "", row.get("grossAreaFt2") or "", row.get("uFactor") or "", row.get("shgc") or "", row.get("cavityR") or "", row.get("continuousR") or ""])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.35,colors.grey), ("FONTSIZE",(0,0),(-1,-1),7), ("VALIGN",(0,0),(-1,-1),"TOP")]))
            story.append(t)
        story.append(Spacer(1, 8))
    doc.build(story)


def prepare_project_comcheck(facts: dict, project_identity: dict, filing_dir: Path, log: RunLog) -> tuple[Path | None, Path, dict]:
    filing_dir.mkdir(parents=True, exist_ok=True)
    pages = list(facts.get("pages") or [])
    en_pages = [p for p in pages if str(p.get("pageType") or "").upper() == "EN"]
    identity_pages = [p for p in pages if str(p.get("pageType") or "").upper() in ("T", "Z")]
    envelope = []
    for page in en_pages:
        for row in list(page.get("envelope") or []):
            if float(row.get("confidence") or 0) >= 0.90:
                envelope.append(row)
    project_title = project_identity.get("title")
    address = project_identity.get("address")
    city = project_identity.get("city")
    state = project_identity.get("state")
    zip_code = project_identity.get("zip")
    energy_code = _fact_value(en_pages, "project", "energyCode")
    stories = _fact_value(identity_pages or pages, "bulk", "stories")
    height = _fact_value(identity_pages or pages, "bulk", "buildingHeightFt")
    floor_area = _fact_value(en_pages or pages, "lighting", "floorAreaFt2") or _fact_value(identity_pages or pages, "bulk", "conditionedFloorAreaFt2") or _fact_value(identity_pages or pages, "bulk", "grossFloorAreaFt2")
    lpd = next(((page.get("lighting") or {}).get("lpdWPerFt2") for page in en_pages if (page.get("lighting") or {}).get("lpdWPerFt2") is not None), None)
    interior_fixtures = []
    exterior_uses = []
    for page in en_pages:
        interior_fixtures.extend([row for row in list((page.get("lighting") or {}).get("fixtures") or []) if float(row.get("confidence") or 0) >= 0.90])
        exterior_uses.extend([row for row in list((page.get("lighting") or {}).get("exteriorUses") or []) if float(row.get("confidence") or 0) >= 0.90])

    missing = []
    for label, value in (("T/Z project title", project_title), ("T/Z project address", address),
                         ("T/Z project city", city), ("T/Z project state", state), ("T/Z project ZIP", zip_code)):
        if value in (None, ""): missing.append(label)
    for label, value in (("EN energy code",energy_code),("T/Z stories",stories),("T/Z building height",height),("EN envelope rows",envelope)):
        if value in (None, "", []): missing.append(label)
    for index, row in enumerate(envelope, start=1):
        kind = str(row.get("kind") or "").lower()
        required = ["grossAreaFt2"]
        if kind in ("wall", "window", "door"):
            required.append("orientation")
        if kind in ("window", "door"):
            required.append("uFactor")
        if kind == "window":
            required.append("shgc")
        if kind in ("wall", "roof", "floor") and row.get("cavityR") in (None, "") and row.get("continuousR") in (None, ""):
            missing.append(f"EN {kind} row {index} has no current R-value")
        if not row.get("assemblyType") and not row.get("description"):
            missing.append(f"EN {kind or 'envelope'} row {index} has no current assembly label")
        absent = [key for key in required if row.get(key) in (None, "")]
        if absent:
            missing.append(f"EN {kind or 'envelope'} row {index} missing " + "/".join(absent))
    if not COMCHECK_CXL_TEMPLATE.is_file(): missing.append("COMcheck CXL structure template")

    # The approved CXL is a schema/enum structure only. Every project and model
    # value below comes from current immutable T/Z/EN evidence.
    cxl = None
    if not missing:
        tree = ET.parse(COMCHECK_CXL_TEMPLATE)
        root = tree.getroot()
        ns_uri = root.tag.split('}')[0].strip('{') if '}' in root.tag else ""
        q = lambda name: f"{{{ns_uri}}}{name}" if ns_uri else name
        _set_xml(root, "feetBldgHeight", f"{float(height):.3f}")
        _set_xml(root, "numberOfStories", int(stories))
        project_node = root.find(q("project"))
        if project_node is not None:
            for key, value in (("projectTitle",project_title),("projectAddress",address),("projectCity",city),("projectState",state),("projectZipCode",zip_code)):
                _set_xml(project_node, key, value)
        location_node = root.find(q("location"))
        if location_node is not None:
            _set_xml(location_node, "state", state)
            _set_xml(location_node, "city", city)
        # Preserve the known CXL enum only when the visible EN code text is consistent with the template's NYC Stretch structure.
        control = root.find(q("control"))
        visible_code = str(energy_code or "")
        if control is not None:
            visible_code_lower = visible_code.lower()
            backstop_evidence = (
                "stretch" in visible_code_lower or
                "appendix ca" in visible_code_lower or
                "backstop" in visible_code_lower or
                ("2020" in visible_code_lower and ("nycecc" in visible_code_lower or "nyc energy" in visible_code_lower))
            )
            if not backstop_evidence:
                missing.append("EN code does not visibly support the 2020 NYCECC Appendix CA Backstop mapping")
        if not missing:
            envelope_node = root.find(q("envelope"))
            above = envelope_node.find(q("aboveGroundWalls")) if envelope_node is not None else None
            roofs = envelope_node.find(q("roofs")) if envelope_node is not None else None
            floors = envelope_node.find(q("floors")) if envelope_node is not None else None
            old_walls = list(above.findall(q("agWall"))) if above is not None else []
            old_roofs = list(roofs.findall(q("roof"))) if roofs is not None else []
            old_floors = list(floors.findall(q("floor"))) if floors is not None else []
            win_example = root.find(f".//{q('window')}")
            door_example = root.find(f".//{q('door')}")
            if above is not None: _clear_children_by_local(above, "agWall")
            if roofs is not None: _clear_children_by_local(roofs, "roof")
            if floors is not None: _clear_children_by_local(floors, "floor")
            opaque_walls = [r for r in envelope if r.get("kind") == "wall" and r.get("grossAreaFt2")]
            windows = [r for r in envelope if r.get("kind") == "window" and r.get("grossAreaFt2")]
            doors = [r for r in envelope if r.get("kind") == "door" and r.get("grossAreaFt2")]
            roofs_facts = [r for r in envelope if r.get("kind") == "roof" and r.get("grossAreaFt2")]
            floors_facts = [r for r in envelope if r.get("kind") == "floor" and r.get("grossAreaFt2")]
            pos = 1
            created_walls = []
            for row in opaque_walls:
                ex = _wall_exemplar(old_walls, row.get("description"))
                if ex is None: continue
                item = copy.deepcopy(ex); pos += 1
                _clear_children_by_local(item, "windows"); _clear_children_by_local(item, "doors")
                for key in ("concreteThickness", "concreteDensity"):
                    _replace_xml(item, key, None)
                for key,val in (("assemblyType",row.get("assemblyType")),("description",row.get("description")),("orientation",_orientation(row.get("orientation"))),
                                ("grossArea",f"{float(row.get('grossAreaFt2')):.3f}"),("cavityRvalue",row.get("cavityR")),("continuousRvalue",row.get("continuousR")),("listPosition",pos)):
                    _replace_xml(item,key,val)
                above.append(item); created_walls.append((item,row))
            def host_for(row):
                parent = str(row.get("parentAssemblyType") or "").strip().lower()
                orient = _orientation(row.get("orientation"))
                for item, src in created_walls:
                    if parent and str(src.get("assemblyType") or "").strip().lower() == parent: return item
                for item, src in created_walls:
                    if _orientation(src.get("orientation")) == orient: return item
                return created_walls[0][0] if created_walls else None
            for kind, rows, exemplar, container_name in (("window",windows,win_example,"windows"),("door",doors,door_example,"doors")):
                if exemplar is None: continue
                for row in rows:
                    host = host_for(row)
                    if host is None: continue
                    container = next((c for c in list(host) if c.tag.rsplit('}',1)[-1] == container_name), None)
                    if container is None:
                        container = ET.Element(q(container_name)); host.append(container)
                    item = copy.deepcopy(exemplar); pos += 1
                    for key in ("propVt", "feetAg", "propProjectionFactor"):
                        _replace_xml(item, key, None)
                    for key,val in (("assemblyType",row.get("assemblyType")),("description",row.get("description")),("orientation",_orientation(row.get("orientation"))),
                                    ("grossArea",f"{float(row.get('grossAreaFt2')):.3f}"),("propUvalue",row.get("uFactor")),("propShgc",row.get("shgc")),
                                    ("productId",row.get("product")),("listPosition",pos)):
                        _replace_xml(item,key,val)
                    container.append(item)
            for row in roofs_facts:
                if not old_roofs or roofs is None: continue
                item=copy.deepcopy(old_roofs[0]); pos+=1
                for key in ("solarReflectance", "solarReflectanceIndex", "thermalEmittance"):
                    _replace_xml(item, key, None)
                for key,val in (("assemblyType",row.get("assemblyType")),("description",row.get("description")),("grossArea",f"{float(row.get('grossAreaFt2')):.3f}"),
                                ("cavityRvalue",row.get("cavityR")),("continuousRvalue",row.get("continuousR")),("listPosition",pos)):
                    _replace_xml(item,key,val)
                roofs.append(item)
            for row in floors_facts:
                if not old_floors or floors is None: continue
                item=copy.deepcopy(old_floors[0]); pos+=1
                for key in ("depthOfInsulation", "slabFullInsulBelowMinRValue"):
                    _replace_xml(item, key, None)
                for key,val in (("assemblyType",row.get("assemblyType")),("description",row.get("description")),("grossArea",f"{float(row.get('grossAreaFt2')):.3f}"),
                                ("continuousRvalue",row.get("continuousR")),("listPosition",pos)):
                    _replace_xml(item,key,val)
                floors.append(item)
            lighting = root.find(q("lighting"))
            has_current_lighting = bool(floor_area and (lpd is not None or interior_fixtures or exterior_uses))
            if lighting is not None and not has_current_lighting:
                root.remove(lighting)
            elif lighting is not None:
                use = lighting.find(f".//{q('wholeBldgUse')}")
                if use is not None:
                    _set_xml(use, "floorArea", f"{float(floor_area):.3f}")
                    _replace_xml(use, "powerDensity", f"{float(lpd):.6f}" if lpd is not None else None)
                    fixture_container = use.find(f".//{q('fixtures')}")
                    fixture_example = fixture_container.find(q("fixture")) if fixture_container is not None else None
                    if fixture_container is not None:
                        _clear_children_by_local(fixture_container, "fixture")
                        if fixture_example is not None:
                            for fixture in interior_fixtures:
                                item = copy.deepcopy(fixture_example); pos += 1
                                _set_xml(item, "description", fixture.get("description"))
                                _set_xml(item, "fixtureWattage", f"{float(fixture.get('wattage') or 0):.6f}")
                                _set_xml(item, "quantity", int(round(float(fixture.get("quantity") or 0))))
                                _set_xml(item, "listPosition", pos)
                                fixture_container.append(item)
                # Never carry prior-project exterior-lighting values into a current project.
                # Only visible EN-page facts may repopulate this section.
                exterior_container = lighting.find(q("exteriorUses"))
                if exterior_container is not None:
                    old_ext = list(exterior_container.findall(q("exteriorUse")))
                    ext_example = old_ext[0] if old_ext else None
                    _clear_children_by_local(exterior_container, "exteriorUse")
                    if exterior_uses and ext_example is not None:
                        for ext in exterior_uses:
                            item = copy.deepcopy(ext_example); pos += 1
                            _set_xml(item, "description", ext.get("description"))
                            _set_xml(item, "useQuantity", ext.get("quantity"))
                            _set_xml(item, "quantityUnits", ext.get("quantityUnits"))
                            _set_xml(item, "listPosition", pos)
                            ext_space = item.find(q("exteriorLightingSpace"))
                            fx = ext_space.find(q("fixtures")) if ext_space is not None else None
                            fx_example = fx.find(q("fixture")) if fx is not None else None
                            if fx is not None:
                                _clear_children_by_local(fx, "fixture")
                                if fx_example is not None and ext.get("fixtureWattage") is not None and ext.get("fixtureQuantity") is not None:
                                    fixture = copy.deepcopy(fx_example)
                                    _set_xml(fixture, "fixtureWattage", f"{float(ext.get('fixtureWattage')):.6f}")
                                    _set_xml(fixture, "quantity", int(round(float(ext.get("fixtureQuantity")))))
                                    fx.append(fixture)
                            exterior_container.append(item)
            if ns_uri: ET.register_namespace("", ns_uri)
            cxl = filing_dir / "COMcheck_PROJECT_INPUT_READY.cxl"
            tree.write(cxl, encoding="utf-8", xml_declaration=True)
            assert_no_reference_identity_text(cxl.read_text(encoding="utf-8"), "COMcheck CXL")

    status = "INPUT_READY" if cxl and not missing else "INPUT_INCOMPLETE"
    audit_json = filing_dir / "COMcheck_INPUT_AUDIT.json"
    audit = {
        "schema":"liber.revex.comcheck-input-audit.v1", "status":status, "project":project_title,
        "source":"bound active Revit evidence and immutable T/Z/EN page facts", "aiScope":"REVIT_T_Z_EN_PAGE_SCAN_ONLY",
        "geometryAuthority":False, "requiredIntegrityGate":0.80, "qualityTarget":0.95, "missing":missing,
        "cxl":cxl.name if cxl else None, "projectFilingReady": bool(cxl and not missing),
        "officialDoeReport": None, "officialDoeReportStatus": "NOT_RUN",
        "projectIdentitySource":"bound active Revit document plus immutable T/Z pages"
    }
    audit_json.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    audit_pdf = filing_dir / "COMcheck_INPUT_AUDIT.pdf"
    _make_comcheck_audit_pdf(audit_pdf, str(project_title or address or "Current REVEX Project"), facts, status, missing, cxl.name if cxl else None)
    log.write("PREPARE_PROJECT_COMCHECK", "PASSED" if cxl else "BLOCKED", comcheckStatus=status, cxl=str(cxl) if cxl else "", audit=str(audit_pdf), missing=missing)
    return cxl, audit_pdf, audit


def require_comcheck_consent(consent: dict, project_id: str, source_revision: str) -> dict:
    consent = dict(consent or {})
    valid = (
        consent.get("schema") == COMCHECK_CONSENT_SCHEMA
        and consent.get("approved") is True
        and str(consent.get("projectId") or "") == str(project_id or "")
        and str(consent.get("sourceEngineeringRevision") or "") == str(source_revision or "")
        and str(consent.get("approvedByUid") or "").strip()
        and str(consent.get("approvedAt") or "").strip()
        and consent.get("service") == COMCHECK_SERVICE
        and consent.get("endpoint") == COMCHECK_ENDPOINT
        and consent.get("scope") == COMCHECK_SCOPE
    )
    if not valid:
        raise PipelineError(
            "Official COMcheck processing requires explicit authorization bound to this project and immutable Engineering revision. No current-project CXL was transmitted."
        )
    return consent


def run_project_backstop(cxl: Path, filing_dir: Path, project_identity: dict, log: RunLog,
                         consent: dict, project_id: str, source_revision: str) -> tuple[Path, Path, Path, dict]:
    consent = require_comcheck_consent(consent, project_id, source_revision)

    def event(status: str, **detail) -> None:
        log.write("COMCHECK_BACKSTOP", status, **detail)

    log.write(
        "COMCHECK_BACKSTOP", "STARTED", cxl=cxl.name,
        engine="PNNL Legacy COMcheck-Web",
        code="2020 NYCECC Appendix CA Modeling Envelope Backstop",
        identitySource="bound active Revit document plus immutable T/Z pages",
        consentScope=consent.get("scope"),
        approvedAt=consent.get("approvedAt"),
        approvedByUid=consent.get("approvedByUid"),
    )
    try:
        report, response_evidence, summary = run_official_backstop(
            cxl, filing_dir, project_identity, event,
            base_url=os.environ.get("REVEX_COMCHECK_BASE_URL", "https://legacy-comcheck.energycode.pnl.gov/CheckWeb/")
        )
    except ComcheckBackstopError:
        raise
    except Exception as exc:
        raise ComcheckBackstopError(f"Unexpected COMcheck Backstop failure: {exc}") from exc
    result_json = filing_dir / "COMcheck_BACKSTOP_RESULT.json"
    from pypdf import PdfReader
    report_text = "\n".join((page.extract_text() or "") for page in PdfReader(str(report)).pages)
    assert_no_reference_identity_text(report_text, "official COMcheck Backstop report")
    log.write(
        "COMCHECK_BACKSTOP", "PASSED", report=report.name,
        result=result_json.name, officialDoeReport=True,
        passes=summary.get("passes"), complianceIndex=summary.get("complianceIndex"),
    )
    return report, response_evidence, result_json, summary


def validate_completion_outputs(
    baseline_model: Path, proposed_model: Path,
    baseline_run: dict, proposed_run: dict,
    review_zip: Path, project_name: str,
    en1_pdf: Path | None, comcheck_cxl: Path | None,
    comcheck_report: Path | None, comcheck_summary: dict | None,
) -> dict:
    failures: list[str] = []
    compiled_models = [baseline_model, proposed_model]
    if len({path.name for path in compiled_models}) != 2:
        failures.append("compiled Baseline/Proposed OSM filenames are not distinct")
    for path in compiled_models:
        if not path.is_file() or path.stat().st_size < 4096:
            failures.append(f"compiled OSM is missing or empty: {path.name}")
        elif "OS:Building" not in path.read_text(encoding="utf-8", errors="ignore"):
            failures.append(f"compiled OSM is not an OpenStudio model: {path.name}")
        else:
            try:
                assert_no_reference_identity_text(path.read_text(encoding="utf-8", errors="ignore"), path.name)
            except PipelineError as exc:
                failures.append(str(exc))

    audit_path = baseline_model.parent / "COMPILATION_AUDIT.json"
    geometry_hashes = []
    if not audit_path.is_file():
        failures.append("GeometryCo compilation audit is missing")
    else:
        try:
            audit = json.loads(audit_path.read_text(encoding="utf-8"))
            if audit.get("success") is not True:
                failures.append("GeometryCo did not atomically commit a successful Baseline/Proposed pair")
            for role in ("baseline", "proposed"):
                report = (audit.get("reports") or {}).get(role) or {}
                geometry = report.get("exact_geometry_lock") or {}
                mapping = report.get("space_mapping") or {}
                schedules = report.get("schedule_lock") or {}
                serialized = report.get("serialized_roundtrip_validation") or {}
                repairs = report.get("energyplus_compatibility_repairs") or {"count": 0, "lossless": True}
                if geometry.get("passed") is not True or geometry.get("new_space_identity_passed") is not True:
                    failures.append(f"{role.title()} GeometryCo exact coordinate/space identity lock did not pass")
                if int(mapping.get("count") or -1) != int(geometry.get("spaces") or -2) or int(mapping.get("ambiguous_count") or 0) != 0:
                    failures.append(f"{role.title()} GeometryCo space mapping is incomplete or ambiguous")
                if schedules.get("passed") is not True or int(schedules.get("changed_schedule_objects") or 0) != 0 or int(schedules.get("changed_protected_schedule_references") or 0) != 0:
                    failures.append(f"{role.title()} GeometryCo schedule/load reference lock did not pass")
                if serialized.get("passed") is not True:
                    failures.append(f"{role.title()} serialized OSM round-trip validation did not pass")
                if int(repairs.get("count") or 0) > 0 and repairs.get("lossless") is not True:
                    failures.append(f"{role.title()} EnergyPlus compatibility repair was not lossless")
                if geometry.get("coordinate_sha256"):
                    geometry_hashes.append(str(geometry["coordinate_sha256"]))
            if len(set(geometry_hashes)) != 1:
                failures.append("Baseline and Proposed compiled OSMs do not share the exact current-project geometry lock")
        except (OSError, ValueError, TypeError) as exc:
            failures.append(f"GeometryCo compilation audit is unreadable: {exc}")

    for role, run in (("Baseline", baseline_run), ("Proposed", proposed_run)):
        for key in ("html", "idf"):
            path = run.get(key)
            if not isinstance(path, Path) or not path.is_file() or path.stat().st_size < 512:
                failures.append(f"{role} simulation {key.upper()} is missing or empty")
            elif isinstance(path, Path):
                try:
                    assert_no_reference_identity_text(path.read_text(encoding="utf-8", errors="ignore"), f"{role} simulation {key.upper()}")
                except PipelineError as exc:
                    failures.append(str(exc))
        end_file = next(iter(Path(run["folder"]).glob("**/eplusout.end")), None)
        if end_file is None or "EnergyPlus Completed Successfully" not in end_file.read_text(encoding="utf-8", errors="ignore"):
            failures.append(f"{role} EnergyPlus run did not record successful completion")
        err_file = next(iter(Path(run["folder"]).glob("**/eplusout.err")), None)
        if err_file is None:
            failures.append(f"{role} EnergyPlus error/warning report is missing")
        else:
            err_text = err_file.read_text(encoding="utf-8", errors="ignore")
            if "**  Fatal  **" in err_text or "EnergyPlus Terminated--Fatal Error Detected" in err_text:
                failures.append(f"{role} EnergyPlus error report contains a fatal termination")

    if not review_zip.is_file() or not zipfile.is_zipfile(review_zip):
        failures.append("EnergyPlus reviewer ZIP is missing or invalid")
    else:
        expected_names = {f"{project_name} - {label}.pdf" for label in REVIEW_PACKAGE_PDF_LABELS}
        with zipfile.ZipFile(review_zip) as package:
            actual_names = {Path(name).name for name in package.namelist() if not name.endswith("/")}
            if actual_names != expected_names:
                failures.append(
                    "EnergyPlus reviewer ZIP does not match the approved nine-PDF record format "
                    f"(missing={sorted(expected_names - actual_names)}, unexpected={sorted(actual_names - expected_names)})"
                )
            for name in sorted(actual_names):
                try:
                    assert_no_reference_identity_text(name, f"review package entry {name}")
                except PipelineError as exc:
                    failures.append(str(exc))
                if not package.read(name).startswith(b"%PDF-"):
                    failures.append(f"review package entry is not a PDF: {name}")
    if en1_pdf is None or not en1_pdf.is_file():
        failures.append("16-page EN-1 filing PDF is missing")
    else:
        from pypdf import PdfReader
        en1_reader = PdfReader(str(en1_pdf))
        if len(en1_reader.pages) != len(EN1_PRINT_SHEETS):
            failures.append(f"EN-1 filing PDF has {len(en1_reader.pages)} pages instead of {len(EN1_PRINT_SHEETS)}")
        try:
            assert_no_reference_identity_text("\n".join((page.extract_text() or "") for page in en1_reader.pages), "EN-1 filing PDF")
        except PipelineError as exc:
            failures.append(str(exc))
    if comcheck_cxl is None or not comcheck_cxl.is_file():
        failures.append("current-project COMcheck CXL is missing")
    if comcheck_report is None or not comcheck_report.is_file() or not comcheck_report.read_bytes().startswith(b"%PDF-"):
        failures.append("official COMcheck Backstop PDF is missing")
    if not comcheck_summary or comcheck_summary.get("officialDoeReport") is not True:
        failures.append("official COMcheck Backstop engine result is not verified")
    if failures:
        raise PipelineError("Energy completion gate failed: " + "; ".join(failures))
    return {
        "compiledOsmCount": 2,
        "simulationCount": 2,
        "geometrySpaceIdentityLock": True,
        "scheduleAndLoadReferenceLock": True,
        "losslessCompatibilityRepair": True,
        "reviewPackagePdfCount": len(REVIEW_PACKAGE_PDF_LABELS),
        "en1Pdf": en1_pdf.name,
        "officialComcheckReport": comcheck_report.name,
        "officialDoeReport": True,
    }


def archive_source_evidence(request: dict, output_root: Path, log: RunLog) -> list[Path]:
    target = output_root / "00_SOURCE_EVIDENCE"
    target.mkdir(parents=True, exist_ok=True)
    copied = []
    seen = set()
    for raw in list(request.get("sourceArtifacts") or []):
        src = Path(str(raw or ""))
        if not src.is_file(): continue
        name = src.name
        key = name.lower()
        if key in seen: continue
        seen.add(key)
        dst = target / name
        shutil.copy2(src, dst)
        copied.append(dst)
    index = target / "SOURCE-EVIDENCE-MANIFEST.json"
    index.write_text(json.dumps({"schema":"liber.revex.source-evidence-package.v1","files":[{"name":p.name,"bytes":p.stat().st_size,"sha256":sha256(p)} for p in copied]}, indent=2), encoding="utf-8")
    copied.append(index)
    log.write("ARCHIVE_SOURCE_EVIDENCE", "PASSED", files=len(copied), folder=str(target))
    return copied


def relative_artifact(path: Path, root: Path, kind: str) -> dict:
    return {"name": path.name, "path": path.relative_to(root).as_posix(), "kind": kind,
            "bytes": path.stat().st_size, "sha256": sha256(path)}


def create_manual_review_package(output_root: Path, artifacts: list[dict], index: dict) -> Path:
    package_path = output_root / "REVEX_ENERGY_MANUAL_REVIEW_PACKAGE.zip"
    package_path.unlink(missing_ok=True)
    review_names: set[str] = set()
    with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as package:
        # Keep the archive itself at exactly eight user-visible entries. The same
        # immutable index is stored as the ZIP comment and in energy-result.json.
        package.comment = json.dumps(index, sort_keys=True, separators=(",", ":")).encode("utf-8")
        for artifact in artifacts:
            relative = Path(str(artifact.get("path") or ""))
            source = (output_root / relative).resolve()
            try:
                source.relative_to(output_root.resolve())
            except ValueError as exc:
                raise PipelineError(f"Manual-review artifact escaped the Energy output folder: {relative}") from exc
            if not source.is_file():
                raise PipelineError(f"Manual-review artifact is missing: {relative}")
            if source == package_path.resolve():
                continue
            if source.stat().st_size != int(artifact.get("bytes") or -1) or sha256(source) != str(artifact.get("sha256") or ""):
                raise PipelineError(f"Manual-review artifact failed its immutable hash/byte contract: {relative}")
            review_name = str(artifact.get("reviewName") or source.name).strip()
            if not review_name or Path(review_name).name != review_name or review_name in review_names:
                raise PipelineError(f"Manual-review artifact has an invalid or duplicate flat name: {review_name}")
            review_names.add(review_name)
            package.write(source, review_name)
    with zipfile.ZipFile(package_path) as verification:
        entries = [name for name in verification.namelist() if not name.endswith("/")]
        if len(entries) != len(VALID_ENERGY_REVIEW_PACKAGE):
            raise PipelineError(
                f"Manual-review archive contains {len(entries)} items instead of "
                f"{len(VALID_ENERGY_REVIEW_PACKAGE)}."
            )
    return package_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--request", type=Path, required=True)
    args = parser.parse_args()
    request = json.loads(args.request.read_text(encoding="utf-8"))
    publication_qa = (
        request.get("publicationQa") is True
        or str(os.environ.get("REVEX_PUBLICATION_QA") or "").strip().lower() in {"1", "true", "yes"}
    )
    output_root = Path(request["outputFolder"]).resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    correlation_id = str(request.get("correlationId") or f"energy-{uuid.uuid4().hex[:12]}")
    initiator = str(request.get("initiator") or "REVEX Energy worker")
    log = RunLog(output_root, correlation_id, initiator)
    started = dt.datetime.now(dt.timezone.utc)
    result_path = output_root / "energy-result.json"
    status = "FAILED"
    error = None
    metrics = None
    artifacts: list[dict] = []
    deliverables: list[Path] = []
    failure_context = None
    approved_run_comparison = {
        "schema": "liber.revex.approved-run-comparison.v1",
        "status": "NOT_RUN",
        "iterationSelection": "WITHHELD_INCOMPLETE_RUN",
        "reviewEligible": False,
        "referenceIdentity": "MASKED",
    }
    page_facts = {"status": "NOT_READ", "pages": []}
    project_identity = {"source": "bound active Revit document plus immutable T/Z pages", "missing": list(REQUIRED_PROJECT_IDENTITY)}
    current_project_name = "Current REVEX Project"
    comcheck_summary = {
        "status": "NOT_RUN",
        "officialDoeReport": None,
        "officialDoeReportStatus": "NOT_RUN",
    }
    try:
        gbxml = Path(request["gbxmlPath"]).resolve()
        weather = Path(request.get("weatherFile") or "").resolve()
        log.write("PIPELINE", "STARTED", version=PIPELINE_VERSION,
                  parentCorrelationId=request.get("parentCorrelationId"), requestManifest=str(args.request),
                  sourceEngineeringRevision=request.get("revision"), standardVersion=request.get("standardVersion"))
        log.dependency("Python runtime", True, executable=sys.executable, version=sys.version.replace("\n", " "),
                       platform=platform.platform())
        for distribution, module_name in (
            ("beautifulsoup4", "bs4"), ("openpyxl", "openpyxl"), ("pypdf", "pypdf"),
            ("reportlab", "reportlab"), ("shapely", "shapely"), ("Pillow", "PIL"),
        ):
            try:
                version = importlib.metadata.version(distribution)
                available = importlib.util.find_spec(module_name) is not None
            except importlib.metadata.PackageNotFoundError:
                version = "not installed"
                available = False
            log.dependency(f"Python module {module_name}", available, version=version)
        for name, path, internal_template in (
            ("Worker request manifest", args.request, False),
            ("Engineering Sync gbXML", gbxml, False),
            ("Weather file (.EPW)", weather, False),
            ("gbXML-to-OSM translator", GBXML_TO_OSM, False),
            ("GeometryCo 4.3.1", GEOMETRYCO, True),
            ("EnergyPlus review packager", PACKAGER, True),
            ("Approved Baseline template", BASELINE_REFERENCE, True),
            ("Approved Proposed template", PROPOSED_REFERENCE, True),
            ("Approved EN-1 structure template", EN1_TEMPLATE, True),
            ("Approved COMcheck CXL structure template", COMCHECK_CXL_TEMPLATE, True),
        ):
            available = path.is_file()
            detail = {"internalTemplate": True} if internal_template else {"path": str(path)}
            if available:
                detail.update(bytes=path.stat().st_size, sha256=sha256(path))
            log.dependency(name, available, **detail)
        log.write("INPUT_VALIDATE", "STARTED", gbxml=str(gbxml), weather=str(weather))
        if not gbxml.is_file():
            raise PipelineError("The Engineering Sync gbXML is missing.")
        if not weather.is_file() or weather.suffix.lower() != ".epw":
            raise PipelineError("Weather file (.EPW) was not resolved by the native REVEX Engineering host.")
        weather_meta = validate_epw(weather)
        log.write("WEATHER_VALIDATE", "PASSED", path=str(weather), **weather_meta)
        log.write("OPENSTUDIO_RESOLVE", "STARTED",
                  requested=str(request.get("openStudioCli") or ""),
                  environmentConfigured=bool(os.environ.get("OPENSTUDIO_CLI")))
        try:
            cli = find_openstudio(str(request.get("openStudioCli") or ""))
        except Exception as openstudio_error:
            log.dependency("OpenStudio CLI", False, requested=str(request.get("openStudioCli") or ""),
                           environmentConfigured=bool(os.environ.get("OPENSTUDIO_CLI")),
                           error=str(openstudio_error))
            log.current_stage = "OPENSTUDIO_RESOLVE"
            raise
        log.dependency("OpenStudio CLI", True, path=str(cli), version=_openstudio_cli_version(cli))
        source_evidence = archive_source_evidence(request, output_root, log)
        page_facts_path = Path(str(request.get("pageFactsPath") or "")) if request.get("pageFactsPath") else None
        page_facts = load_page_facts(page_facts_path)
        project_identity = current_project_identity(page_facts)
        current_project_name = str(project_identity.get("title") or project_identity.get("address") or "Current REVEX Project")
        log.write(
            "PROJECT_IDENTITY", "PASSED" if not project_identity.get("missing") else "INCOMPLETE",
            source=project_identity.get("source"), missing=project_identity.get("missing"),
            tPages=sum(1 for page in page_facts.get("pages") or [] if str(page.get("pageType") or "").upper() == "T"),
            zPages=sum(1 for page in page_facts.get("pages") or [] if str(page.get("pageType") or "").upper() == "Z"),
            optionalMissing=project_identity.get("optionalMissing") or [],
        )
        if project_identity.get("missing"):
            raise PipelineError(
                "Current project identity is incomplete in active Revit T/Z evidence: "
                + ", ".join(project_identity["missing"])
            )
        original_dir = output_root / "01_ORIGINAL_MODELS"
        original_dir.mkdir(parents=True, exist_ok=True)
        geometry_osm = original_dir / "REVIT_GEOMETRY_ORIGINAL.osm"
        convert_gbxml(gbxml, geometry_osm, cli, output_root, log)
        baseline_model, proposed_model = compile_models(geometry_osm, cli, output_root, log)
        stamp_compiled_project_identity(baseline_model, project_identity, "BASELINE", log)
        stamp_compiled_project_identity(proposed_model, project_identity, "PROPOSED", log)
        baseline_run = simulate("baseline", baseline_model, weather, cli, output_root, log)
        proposed_run = simulate("proposed", proposed_model, weather, cli, output_root, log)
        packager = load_module(PACKAGER, "revex_energyplus_packager")
        review_dir = output_root / "04_REVIEW_PACKAGE"
        review_zip = Path(packager.generate_package(
            str(baseline_run["html"]), str(proposed_run["html"]), str(review_dir),
            current_project_name, True,
            standard_version=str(request.get("standardVersion") or "NYCECC 2020"),
            baseline_model_file=str(baseline_run["idf"]), proposed_model_file=str(proposed_run["idf"]),
        ))
        log.write("REVIEW_PACKAGER", "PASSED", zip=str(review_zip))
        filing_dir = output_root / "05_FILING"
        filing_dir.mkdir(parents=True, exist_ok=True)
        en1_xlsx, en1_pdf, metrics = prepare_en1(
            packager, baseline_run["html"], proposed_run["html"], filing_dir, log, weather_meta, project_identity
        )
        comcheck_cxl, comcheck_audit_pdf, comcheck_audit = prepare_project_comcheck(
            page_facts, project_identity, filing_dir, log
        )
        comcheck_report = comcheck_response = comcheck_result_json = None
        if comcheck_cxl and en1_pdf and not project_identity.get("missing"):
            comcheck_report, comcheck_response, comcheck_result_json, comcheck_summary = run_project_backstop(
                comcheck_cxl, filing_dir, project_identity, log,
                request.get("externalProcessingConsent") or {},
                str(request.get("projectId") or ""), str(request.get("revision") or "")
            )
        deliverables = [*source_evidence, geometry_osm, baseline_model, proposed_model, baseline_run["html"], proposed_run["html"],
                        baseline_run["idf"], proposed_run["idf"], review_zip, en1_xlsx, comcheck_audit_pdf,
                        filing_dir / "COMcheck_INPUT_AUDIT.json"]
        if comcheck_cxl: deliverables.append(comcheck_cxl)
        if en1_pdf:
            deliverables.append(en1_pdf)
        for path in (comcheck_report, comcheck_response, comcheck_result_json):
            if path:
                deliverables.append(path)
        # Preserve native simulation/report evidence, not only the summarized review package.
        # Companion groups these separately so the original OpenStudio/EnergyPlus outputs remain auditable.
        for role in (baseline_run, proposed_run):
            deliverables.extend(path for path in role["folder"].glob("**/*") if path.is_file())
        deliverables.extend(path for path in review_dir.glob("**/*") if path.is_file())
        deliverables.extend(path for path in (output_root / "02_COMPILED_MODELS").glob("**/*") if path.is_file())
        if not en1_pdf:
            status = "BLOCKED_EN1_PDF_EXPORT"
            error = "The model/report package and EN-1 workbook are complete, but Excel or LibreOffice did not produce the EN-1 PDF."
            failure_context = {"failedStage": "PREPARE_EN1", "type": "PipelineBlocked", "message": error}
        elif project_identity.get("missing"):
            status = "BLOCKED_PROJECT_IDENTITY"
            error = "Current project identity is incomplete in active Revit T/Z evidence: " + ", ".join(project_identity["missing"])
            failure_context = {"failedStage": "PROJECT_IDENTITY", "type": "PipelineBlocked", "message": error}
        elif not comcheck_cxl:
            status = "BLOCKED_COMCHECK_INPUT"
            error = "Current T/Z/EN evidence is insufficient to form a project-specific COMcheck CXL: " + ", ".join(comcheck_audit.get("missing") or [])
            failure_context = {"failedStage": "PREPARE_PROJECT_COMCHECK", "type": "PipelineBlocked", "message": error}
        else:
            completion = validate_completion_outputs(
                baseline_model, proposed_model, baseline_run, proposed_run,
                review_zip, current_project_name,
                en1_pdf, comcheck_cxl, comcheck_report, comcheck_summary,
            )
            log.write("COMPLETION_GATE", "PASSED", **completion)
            approved_run_comparison = compare_approved_run_profile(
                metrics, baseline_model.parent / "COMPILATION_AUDIT.json", baseline_model
            )
            comparison_path = output_root / "APPROVED_RUN_COMPARISON_MASKED.json"
            comparison_path.write_text(json.dumps(approved_run_comparison, indent=2), encoding="utf-8")
            deliverables.append(comparison_path)
            if approved_run_comparison.get("status") == "REGRESSION":
                status = "BLOCKED_APPROVED_RUN_REGRESSION"
                error = (
                    "The completed simulation diverged from the masked approved-run profile; "
                    "the manual-review candidate was withheld."
                )
                failure_context = {
                    "failedStage": "APPROVED_RUN_COMPARISON",
                    "type": "PipelineBlocked",
                    "message": error,
                }
                log.write(
                    "APPROVED_RUN_COMPARISON", "REGRESSION",
                    passedChecks=approved_run_comparison.get("passedChecks"),
                    totalChecks=approved_run_comparison.get("totalChecks"),
                    normalizedRegressionScore=approved_run_comparison.get("normalizedRegressionScore"),
                    referenceIdentity="MASKED",
                )
            else:
                log.write(
                    "APPROVED_RUN_COMPARISON", approved_run_comparison.get("status") or "PASSED",
                    iterationSelection=approved_run_comparison.get("iterationSelection"),
                    normalizedRegressionScore=approved_run_comparison.get("normalizedRegressionScore"),
                    referenceIdentity="MASKED",
                )
                status = "COMPLETE"
        log.write("PIPELINE", status)
    except Exception as ex:
        failed_stage = log.current_stage
        error = str(ex)
        failure_context = {
            "failedStage": failed_stage,
            "type": type(ex).__name__,
            "message": error,
            "causes": exception_chain(ex),
            "traceback": traceback.format_exc(),
        }
        log.write("PIPELINE", "FAILED", error=error, failedStage=failed_stage,
                  causes=failure_context["causes"], traceback=failure_context["traceback"])

    # Publish lightweight execution evidence even when a later stage fails. Hash only
    # after the final log line so the manifest always describes the attached bytes.
    deliverables.extend(output_root.glob("*.log"))
    deliverables.extend(output_root.glob("*.jsonl"))
    deliverables.extend(output_root.glob("*_AUDIT.json"))
    deliverables.extend(output_root.glob("03_SIMULATION/*/REVEX_OPENSTUDIO_RUN.log"))
    for pattern in ("03_SIMULATION/**/eplustbl.html", "03_SIMULATION/**/eplusout.err", "03_SIMULATION/**/eplusout.sql",
                    "03_SIMULATION/**/eplusout.csv", "03_SIMULATION/**/eplusout.end", "03_SIMULATION/**/in.idf", "03_SIMULATION/**/workflow.osw"):
        deliverables.extend(output_root.glob(pattern))
    for path in sorted({path.resolve() for path in deliverables if path and path.is_file()}):
        parts = set(path.parts)
        if "05_FILING" in parts:
            if path.name in {"COMcheck_PROJECT_INPUT_READY.cxl", "EN-1_READY_TO_INSERT.xlsx"}:
                kind = "filing-input"
            elif path.suffix.lower() == ".pdf" and (path.name.startswith("EN-1_") or path.name.startswith("COMcheck_OFFICIAL_")):
                kind = "filing-output"
            elif path.name.startswith("COMcheck_BACKSTOP_"):
                kind = "engine-evidence"
            else:
                kind = "review-report"
        elif "04_REVIEW_PACKAGE" in parts:
            kind = "review-report"
        elif "03_SIMULATION" in parts:
            kind = "simulation-output"
        elif "02_COMPILED_MODELS" in parts:
            kind = "compiled-model"
        elif "01_ORIGINAL_MODELS" in parts:
            kind = "original-model"
        elif "00_SOURCE_EVIDENCE" in parts:
            kind = "source-evidence"
        else:
            kind = "diagnostic"
        artifacts.append(relative_artifact(path, output_root, kind))

    # Put one human-reviewable bundle beside the run folder only for an eligible
    # completed iteration. The review contract is exactly seven files plus the
    # Packager reports archive; diagnostics and protected references stay outside.
    manual_review_index = {
        "schema": "liber.revex.energy-manual-review-package.v1",
        "pipelineVersion": PIPELINE_VERSION,
        "projectId": request.get("projectId"),
        "projectName": current_project_name,
        "sourceEngineeringRevision": request.get("revision"),
        "status": status,
        "iterationSelection": approved_run_comparison.get("iterationSelection"),
        "approvedRunComparisonStatus": approved_run_comparison.get("status"),
        "approvedRunProfileSha256": approved_run_comparison.get("profileSha256"),
        "referenceTemplatesIncluded": False,
        "referenceIdentityExcluded": True,
        "recordFormat": {
            "topLevelFiles": 7,
            "topLevelArchives": 1,
            "contract": list(VALID_ENERGY_REVIEW_PACKAGE),
            "packagerArchivePdfCount": len(REVIEW_PACKAGE_PDF_LABELS),
        },
        "files": [],
    }
    manual_review_package = None
    review_artifacts: list[dict] = []
    if status == "COMPLETE" and approved_run_comparison.get("reviewEligible") is True:
        review_contract_paths = [
            (geometry_osm, "geometry-osm", "GEOMETRY.osm"),
            (baseline_model, "baseline-osm", "BASELINE.osm"),
            (proposed_model, "proposed-osm", "PROPOSED.osm"),
            (baseline_run["html"], "baseline-html", "BASELINE_REPORT.html"),
            (proposed_run["html"], "proposed-html", "PROPOSED_REPORT.html"),
            (en1_xlsx, "en1-spreadsheet", "EN-1.xlsx"),
            (comcheck_report, "official-comcheck-pdf", "COMcheck_BACKSTOP.pdf"),
            (review_zip, "packager-reports-archive", "PACKAGER_REPORTS.zip"),
        ]
        review_artifacts = []
        for path, kind, review_name in review_contract_paths:
            artifact = relative_artifact(path, output_root, kind)
            artifact["reviewName"] = review_name
            review_artifacts.append(artifact)
        if tuple(artifact["kind"] for artifact in review_artifacts) != VALID_ENERGY_REVIEW_PACKAGE:
            raise PipelineError("Manual-review package did not match the seven-files-plus-one-archive contract.")
        manual_review_index["files"] = [
            {key: artifact[key] for key in ("path", "reviewName", "kind", "bytes", "sha256")}
            for artifact in review_artifacts
        ]
        manual_review_package = create_manual_review_package(output_root, review_artifacts, manual_review_index)
        artifacts.append(relative_artifact(manual_review_package, output_root, "manual-review-package"))

    finished = dt.datetime.now(dt.timezone.utc)
    result = {
        "schema": SCHEMA,
        "pipelineVersion": PIPELINE_VERSION,
        "correlationId": correlation_id,
        "parentCorrelationId": request.get("parentCorrelationId"),
        "initiator": initiator,
        "projectId": request.get("projectId"),
        "projectName": current_project_name,
        "projectIdentity": project_identity,
        "sourceEngineeringRevision": request.get("revision"),
        "resultRevision": "energy_" + finished.strftime("%Y%m%dT%H%M%SZ"),
        "status": status,
        "startedAt": started.isoformat(),
        "finishedAt": finished.isoformat(),
        "error": error,
        "failureContext": failure_context,
        "dependencies": log.dependencies,
        "diagnosticLog": log.path.name,
        "metrics": metrics,
        "packageLayout": [
            {"folder":"00_SOURCE_EVIDENCE", "kind":"source-evidence", "label":"Immutable source evidence"},
            {"folder":"01_ORIGINAL_MODELS", "kind":"original-model", "label":"Original Revit-derived OSM"},
            {"folder":"02_COMPILED_MODELS", "kind":"compiled-model", "label":"Compiled Baseline / Proposed OSM models"},
            {"folder":"03_SIMULATION", "kind":"simulation-output", "label":"Original OpenStudio / EnergyPlus reports"},
            {"folder":"04_REVIEW_PACKAGE", "kind":"review-report", "label":"Organized review reports"},
            {"folder":"05_FILING", "kind":"filing-output", "label":"Filing-ready outputs and current-project inputs"},
            {"folder":"REVEX_ENERGY_MANUAL_REVIEW_PACKAGE.zip", "kind":"manual-review-package", "label":"Seven files plus one Packager archive; eligible completed iteration only"},
        ],
        "approvedRunComparison": approved_run_comparison,
        "manualReviewPackage": ({
            "status": "CREATED",
            "name": manual_review_package.name,
            "path": manual_review_package.name,
            "bytes": manual_review_package.stat().st_size,
            "sha256": sha256(manual_review_package),
            "topLevelFiles": 7,
            "topLevelArchives": 1,
            "referenceTemplatesIncluded": False,
            "referenceIdentityExcluded": True,
        } if manual_review_package else {
            "status": "WITHHELD",
            "reason": approved_run_comparison.get("iterationSelection") or status,
            "referenceTemplatesIncluded": False,
            "referenceIdentityExcluded": True,
        }),
        # Production consumers receive exactly the seven-files-plus-one-archive
        # review contract. Extra run evidence is exported only by publication QA.
        "artifacts": review_artifacts,
        "debugArtifacts": artifacts if publication_qa else [],
        "publicationQa": publication_qa,
        "comcheck": {
            **(comcheck_audit if 'comcheck_audit' in locals() else {"status":"NOT_RUN"}),
            **comcheck_summary,
            "projectInputReady": bool('comcheck_cxl' in locals() and comcheck_cxl),
            # Do not record `false` when the official engine was not reached or failed.
            # `true` is reserved for a validated PNNL PDF; otherwise the value remains unknown.
            "officialDoeReport": True if comcheck_summary.get("officialDoeReport") is True else None,
            "officialDoeReportStatus": "VERIFIED" if comcheck_summary.get("officialDoeReport") is True else str(comcheck_summary.get("status") or "NOT_RUN"),
        },
        "revitWriteBack": False,
        "pdfInsertion": False,
        "authorityBoundary": "Revit writes ended after Space/EADM/EN-Energy tag synchronization.",
    }
    result_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return 0 if status == "COMPLETE" else 2


if __name__ == "__main__":
    raise SystemExit(main())
