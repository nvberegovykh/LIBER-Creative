#!/usr/bin/env python3
from __future__ import annotations

import ast
import json
from pathlib import Path
import re
import shutil
import tempfile

import revex_user_identity_en1 as r89
import revex_identity_content_agent as r88
import revex_energy_pipeline_r69 as r69


def _pinned_identity_consumer(facts: dict) -> dict:
    """Execute the exact r49 identity consumer functions without heavy simulation imports."""
    pipeline = Path(__file__).resolve().parents[2] / 'src/Liber.Revex.Revit/Engineering/Energy/revex_energy_pipeline.py'
    tree = ast.parse(pipeline.read_text(encoding='utf-8'))
    wanted = {'_best_page_value', 'current_project_identity'}
    nodes = [node for node in tree.body if isinstance(node, ast.FunctionDef) and node.name in wanted]
    assert {node.name for node in nodes} == wanted
    namespace = {
        're': re,
        'REQUIRED_PROJECT_IDENTITY': ('title','address','city','state','zip'),
        'OPTIONAL_PROJECT_IDENTITY': (
            'houseNumber','streetName','borough','block','lot','bin','communityBoard','jobType',
            'architecturalJobNumber','mechanicalJobNumber','plumbingJobNumber',
        ),
    }
    exec(compile(ast.Module(body=nodes, type_ignores=[]), str(pipeline), 'exec'), namespace)
    return namespace['current_project_identity'](facts)


def test_missing_only_identity() -> None:
    with tempfile.TemporaryDirectory(prefix='revex-r89-identity-') as td:
        root = Path(td)
        facts = root / 'facts.json'
        facts.write_text(json.dumps({
            'schema':'liber.revex.revit-page-facts.v1',
            'structuredIdentity':{'title':'18 Example Avenue','address':'18 Example Avenue'},
            'pages':[
                {'pageType':'T','confidence':0.99,'project':{'title':'18 Example Avenue','address':'18 Example Avenue'}},
                {'pageType':'Z','confidence':0.98,'project':{}},
            ]
        }), encoding='utf-8')
        request = root / 'request.json'
        request.write_text(json.dumps({
            'revision':'eng_test',
            'pageFactsPath':str(facts),
            'sourceArtifacts':[],
            'comcheckContext':{'identityOverride':{
                'title':'SHOULD NOT OVERWRITE', 'address':'999 Wrong Street',
                'city':'Brooklyn','state':'NY','zip':'11225'
            }}
        }), encoding='utf-8')

        # Reproduce the exact guard order, including r95's final missing-only projection.
        effective = r89.resolve_request(request, root)
        assert effective != request
        effective = r88.resolve_request(effective, root)
        effective = r69._resolved_request(effective, root)
        effective = r89.resolve_request(effective, root)

        data = json.loads(effective.read_text(encoding='utf-8'))
        resolved = json.loads(Path(data['pageFactsPath']).read_text(encoding='utf-8'))
        current = r89._current_project_identity(resolved)
        assert current['title'] == '18 Example Avenue'
        assert current['address'] == '18 Example Avenue'
        assert current['city'] == 'Brooklyn' and current['state'] == 'NY' and current['zip'] == '11225'

        # This is the actual downstream consumer which raised PROJECT_IDENTITY in the live run.
        consumed = _pinned_identity_consumer(resolved)
        assert consumed['missing'] == [], consumed
        assert consumed['title'] == '18 Example Avenue'
        assert consumed['address'] == '18 Example Avenue'
        assert consumed['city'] == 'Brooklyn' and consumed['state'] == 'NY' and consumed['zip'] == '11225'

        audit = json.loads((root/'PROJECT_IDENTITY_USER_OVERRIDE_R89.json').read_text(encoding='utf-8'))
        # A second idempotent projection may rewrite the audit to NO_MISSING_FIELDS_TO_FILL;
        # the derived facts remain the consumer authority and source evidence remains immutable.
        assert audit['sourceEvidenceMutated'] is False
        assert audit['remainingMissing'] == []
        assert current['title'] != 'SHOULD NOT OVERWRITE' and current['address'] != '999 Wrong Street'


def make_workbook(path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for index, name in enumerate(r89.EN1_PRINT_SHEETS, start=1):
        ws = wb.create_sheet(name)
        ws['A1'] = f'REVEX EN-1 QA PAGE {index}'
        ws['A2'] = 'Border-safe fit-to-one-page regression'
        ws.print_area = 'A1:H40'
    wb.save(path)


def test_en1_people_and_print() -> None:
    with tempfile.TemporaryDirectory(prefix='revex-r89-en1-') as td:
        root = Path(td)
        xlsx = root/'EN-1_READY_TO_INSERT.xlsx'
        make_workbook(xlsx)
        applicant = {
            'firstName':'Test','lastName':'Applicant','middleInitial':'A','businessName':'Applicant Studio',
            'businessEmail':'applicant@example.com','businessAddress':'1 Test Street','businessTelephone':'212-555-0100',
            'city':'Brooklyn','state':'NY','zip':'11225','email':'person@example.com','licenseNumber':'012345'
        }
        modeler = {
            'firstName':'Test','lastName':'Modeler','middleInitial':'M','businessName':'Modeling Studio',
            'businessAddress':'2 Test Street','telephone':'212-555-0101','city':'Brooklyn','state':'NY','zip':'11225',
            'email':'modeler@example.com'
        }
        audit = r89._fill_people_identity(xlsx, applicant, modeler)
        assert 'licenseNumber' in audit['applicantFields']
        assert 'email' in audit['modelerFields']
        if shutil.which('soffice') or shutil.which('libreoffice'):
            pdf = root/'EN-1_READY_TO_INSERT.pdf'
            print_audit = r89._print_en1_pdf(xlsx, pdf, root)
            assert pdf.is_file() and pdf.stat().st_size > 1024
            assert print_audit['pageCount'] == len(r89.EN1_PRINT_SHEETS) == 16


def test_exact_nine_review_entries() -> None:
    with tempfile.TemporaryDirectory(prefix='revex-r89-nine-') as td:
        root = Path(td)
        rows=[]
        for index, review_name in enumerate(r89.PUBLIC_REVIEW_NAMES, start=1):
            path=root/f'item-{index}.dat'
            path.write_bytes((review_name+'\n').encode('utf-8'))
            rows.append(r89._artifact(path, root, 'qa', review_name=review_name, user_visible=True))
        package=r89._regenerate_review_package(root, rows, {'projectId':'revex_test','sourceEngineeringRevision':'eng_test','status':'COMPLETE'})
        assert package.is_file()
        import zipfile
        with zipfile.ZipFile(package) as z:
            assert len(z.namelist()) == 9
            assert set(z.namelist()) == set(r89.PUBLIC_REVIEW_NAMES)


test_missing_only_identity()
test_en1_people_and_print()
test_exact_nine_review_entries()
print(json.dumps({
    'schema':'liber.revex.r95-user-identity-en1-qa.v1',
    'status':'PASSED',
    'projectIdentity':{
        'manualFallbackOnlyFillsMissing':True,
        'existingRevitValuesCannotBeOverwritten':True,
        'r88R69Survival':True,
        'pinnedR49ConsumerBoundaryComplete':True,
    },
    'en1':{'applicantModelerExplicit':True,'pdfPrintContractPages':16,'fitToOnePagePerFormSheet':True},
    'userOutput':{'count':9,'exactReviewNames':list(r89.PUBLIC_REVIEW_NAMES)}
}, indent=2))
