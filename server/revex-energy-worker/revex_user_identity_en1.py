#!/usr/bin/env python3
"""REVEX r89 user identity + EN-1 finalization.

This module is deliberately narrow:
- explicit project identity entered during revision-scoped COMcheck authorization may
  fill ONLY fields still missing from immutable active-Revit T/Z evidence;
- applicant / lead-modeler identity is used only for the EN-1 filing workbook/PDF and
  is never projected into COMcheck;
- immutable Revit evidence is never edited;
- the clean human-review contract is exactly nine user-visible entries, while the CXL
  and COMcheck engine JSON may remain hidden engine evidence for integrity checks.
"""
from __future__ import annotations

import copy
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile

REQUIRED_PROJECT_IDENTITY = ("title", "address", "city", "state", "zip")
EN1_PRINT_SHEETS = (
    "1,2,3 Information", "3d. LL97 GHG Summary", "4. Compliance",
    "5a. Baseline Rotations", "5b. Usage Summary", "6a. Ext. Wall Areas",
    "6b. Fenestration", "6c1. Wall Types", "6c2. Wall Types - Addl Rows",
    "6d.1 Interior LPD-Space Method", "6d.2 Interior LPD-Bldg  Method",
    "6e. Ext LPD", "6f. Process Equip.", "6g. Service HW", "HVAC_Cover",
    "HVAC Air-side",
)
APPLICANT_FIELDS = {
    "B11": ("Last Name ", "lastName"),
    "D11": ("First Name ", "firstName"),
    "I11": ("Middle Initial ", "middleInitial"),
    "B12": ("Business Name ", "businessName"),
    "F12": ("Business Email ", "businessEmail"),
    "B13": ("Business Address ", "businessAddress"),
    "F13": ("Business Telephone ", "businessTelephone"),
    "B14": ("City  ", "city"),
    "C14": ("State ", "state"),
    "E14": ("Zip  ", "zip"),
    "B15": ("Email ", "email"),
    "G15": ("License Number ", "licenseNumber"),
}
MODELER_FIELDS = {
    "B19": ("Last Name (lead modeler) ", "lastName"),
    "D19": ("First Name (lead modeler) ", "firstName"),
    "I19": ("Middle Initial ", "middleInitial"),
    "B20": ("Business Name ", "businessName"),
    "F20": ("Email (lead modeler) ", "email"),
    "B21": ("Business Address ", "businessAddress"),
    "F21": ("Telephone (lead modeler) ", "telephone"),
    "B22": ("City  ", "city"),
    "C22": ("State ", "state"),
    "E22": ("Zip  ", "zip"),
}
PUBLIC_REVIEW_NAMES = (
    "GEOMETRY.osm",
    "BASELINE.osm",
    "PROPOSED.osm",
    "BASELINE_REPORT.html",
    "PROPOSED_REPORT.html",
    "EN-1.xlsx",
    "EN-1.pdf",
    "COMcheck_BACKSTOP.pdf",
    "PACKAGER_REPORTS.zip",
)


def _clean(value: object, limit: int = 200) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())[:limit]


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def _context(request: dict) -> dict:
    context = dict(request.get("comcheckContext") or {})
    consent = dict(request.get("externalProcessingConsent") or {})
    # Broker sends these in projectSource/comcheckContext; consent copies are accepted as
    # a defensive fallback so a future transport refactor does not silently lose user input.
    for key in ("identityOverride", "en1Applicant", "en1Modeler"):
        if not isinstance(context.get(key), dict) and isinstance(consent.get(key), dict):
            context[key] = consent.get(key)
    return context


def _sanitize_project_identity(raw: object) -> dict[str, str]:
    if not isinstance(raw, dict):
        return {}
    out: dict[str, str] = {}
    for key in REQUIRED_PROJECT_IDENTITY:
        value = _clean(raw.get(key))
        if not value:
            continue
        if key == "state":
            value = value.upper()
            if not re.fullmatch(r"[A-Z][A-Z .-]{1,31}", value):
                raise ValueError("manual project identity state is invalid")
        if key == "zip" and not re.fullmatch(r"\d{5}(?:-\d{4})?", value):
            raise ValueError("manual project identity ZIP must be 5 digits or ZIP+4")
        out[key] = value
    return out


def _sanitize_person(raw: object, *, modeler: bool = False) -> dict[str, str]:
    if not isinstance(raw, dict):
        return {}
    allowed = {key for _, key in (MODELER_FIELDS if modeler else APPLICANT_FIELDS).values()}
    out: dict[str, str] = {}
    for key in allowed:
        value = _clean(raw.get(key))
        if not value:
            continue
        if key == "middleInitial":
            value = value[:1].upper()
        elif key == "state":
            value = value.upper()
            if not re.fullmatch(r"[A-Z][A-Z .-]{1,31}", value):
                raise ValueError(f"EN-1 {'modeler' if modeler else 'applicant'} state is invalid")
        elif key == "zip" and not re.fullmatch(r"\d{5}(?:-\d{4})?", value):
            raise ValueError(f"EN-1 {'modeler' if modeler else 'applicant'} ZIP must be 5 digits or ZIP+4")
        out[key] = value
    return out


def _best_page_value(pages: list[dict], page_type: str, key: str) -> str:
    candidates: list[tuple[float, str]] = []
    for page in pages:
        if _clean(page.get("pageType")).upper() != page_type:
            continue
        value = _clean((page.get("project") or {}).get(key))
        if value:
            candidates.append((float(page.get("confidence") or 0), value))
    candidates.sort(reverse=True)
    return candidates[0][1] if candidates else ""


def _current_project_identity(facts: dict) -> dict[str, str]:
    pages = list(facts.get("pages") or [])
    native = dict(facts.get("structuredIdentity") or {})
    return {
        key: _best_page_value(pages, "Z", key) or _best_page_value(pages, "T", key) or _clean(native.get(key))
        for key in REQUIRED_PROJECT_IDENTITY
    }


def resolve_request(request_path: Path, output_root: Path) -> Path:
    """Fill only missing project identity fields using explicit revision-scoped input."""
    request_path = Path(request_path).resolve()
    request = json.loads(request_path.read_text(encoding="utf-8"))
    supplied = _sanitize_project_identity(_context(request).get("identityOverride"))
    if not supplied:
        return request_path

    page_path = Path(_clean(request.get("pageFactsPath"), 1000))
    if not page_path.is_file():
        raise ValueError("manual project identity fallback was supplied but pageFactsPath is unavailable")
    facts = json.loads(page_path.read_text(encoding="utf-8"))
    before = _current_project_identity(facts)
    accepted = {key: value for key, value in supplied.items() if not before.get(key)}
    ignored = {key: value for key, value in supplied.items() if before.get(key)}

    audit = {
        "schema": "liber.revex.project-identity-user-override.v1",
        "authority": "explicit-user-input-during-revision-scoped-comcheck-authorization",
        "sourceEngineeringRevision": _clean(request.get("revision")),
        "before": before,
        "suppliedFields": sorted(supplied),
        "acceptedMissingFields": accepted,
        "ignoredAlreadyEstablishedFields": sorted(ignored),
        "sourceEvidenceMutated": False,
    }
    audit_path = output_root / "PROJECT_IDENTITY_USER_OVERRIDE_R89.json"

    if not accepted:
        audit.update(status="NO_MISSING_FIELDS_TO_FILL", after=before, remainingMissing=[key for key in REQUIRED_PROJECT_IDENTITY if not before.get(key)])
        audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")
        print(json.dumps({"stage":"PROJECT_IDENTITY_USER_OVERRIDE","status":audit["status"],"ignored":sorted(ignored)}, ensure_ascii=True), flush=True)
        return request_path

    derived = copy.deepcopy(facts)
    structured = dict(derived.get("structuredIdentity") or {})
    for key, value in accepted.items():
        structured[key] = value
    derived["structuredIdentity"] = structured
    # r49 gives Z/T page fields precedence. Project accepted values into blank T/Z slots only.
    for page in list(derived.get("pages") or []):
        if _clean(page.get("pageType")).upper() not in {"T", "Z"}:
            continue
        project = dict(page.get("project") or {})
        for key, value in accepted.items():
            if not _clean(project.get(key)):
                project[key] = value
        page["project"] = project
    after = _current_project_identity(derived)
    audit.update(status="APPLIED", after=after, remainingMissing=[key for key in REQUIRED_PROJECT_IDENTITY if not after.get(key)])
    audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")

    facts_path = output_root / "00_PAGE_FACTS_USER_IDENTITY_R89.json"
    facts_path.write_text(json.dumps(derived, ensure_ascii=True, indent=2), encoding="utf-8")
    derived_request = dict(request)
    derived_request["pageFactsPath"] = str(facts_path)
    derived_request["identityUserOverride"] = {
        "schema": audit["schema"],
        "authority": audit["authority"],
        "acceptedMissingFields": sorted(accepted),
        "auditFile": audit_path.name,
    }
    request_copy = output_root / "00_PIPELINE_REQUEST_USER_IDENTITY_R89.json"
    request_copy.write_text(json.dumps(derived_request, ensure_ascii=True, indent=2), encoding="utf-8")
    print(json.dumps({
        "stage":"PROJECT_IDENTITY_USER_OVERRIDE","status":"APPLIED",
        "accepted":sorted(accepted),"ignored":sorted(ignored),"remainingMissing":audit["remainingMissing"]
    }, ensure_ascii=True), flush=True)
    return request_copy


def _fill_people_identity(xlsx: Path, applicant: dict[str, str], modeler: dict[str, str]) -> dict:
    from openpyxl import load_workbook
    workbook = load_workbook(xlsx)
    if "1,2,3 Information" not in workbook.sheetnames:
        raise ValueError("EN-1 filing workbook is missing the Information sheet")
    info = workbook["1,2,3 Information"]
    for mapping, values in ((APPLICANT_FIELDS, applicant), (MODELER_FIELDS, modeler)):
        for cell, (label, key) in mapping.items():
            info[cell] = label + _clean(values.get(key))
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(xlsx)

    verify = load_workbook(xlsx, read_only=True, data_only=False)
    info = verify["1,2,3 Information"]
    wrong = []
    for mapping, values in ((APPLICANT_FIELDS, applicant), (MODELER_FIELDS, modeler)):
        for cell, (label, key) in mapping.items():
            expected = label + _clean(values.get(key))
            if str(info[cell].value or "") != expected:
                wrong.append(cell)
    verify.close()
    if wrong:
        raise ValueError("EN-1 applicant/modeler identity verification failed at: " + ", ".join(wrong))
    return {"applicantFields": sorted(applicant), "modelerFields": sorted(modeler)}


def _print_en1_pdf(xlsx: Path, pdf: Path, output_root: Path) -> dict:
    from openpyxl import load_workbook
    from openpyxl.worksheet.properties import PageSetupProperties
    from pypdf import PdfReader

    print_copy = xlsx.with_name("EN-1_PRINT_SOURCE.xlsx")
    shutil.copy2(xlsx, print_copy)
    workbook = load_workbook(print_copy)
    missing = [name for name in EN1_PRINT_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError("EN-1 print set is missing filing sheets: " + ", ".join(missing))
    for sheet in workbook.worksheets:
        if sheet.title in EN1_PRINT_SHEETS:
            sheet.sheet_state = "visible"
            props = sheet.sheet_properties.pageSetUpPr
            if props is None:
                props = PageSetupProperties()
                sheet.sheet_properties.pageSetUpPr = props
            props.fitToPage = True
            props.autoPageBreaks = False
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
        else:
            sheet.sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(EN1_PRINT_SHEETS[0])
    workbook.save(print_copy)

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise ValueError("LibreOffice Calc is unavailable for EN-1 PDF export")
    profile = Path(tempfile.gettempdir()) / f"revex-lo-{uuid.uuid4().hex}"
    export_dir = Path(tempfile.mkdtemp(prefix="revex-en1-pdf-"))
    try:
        command = [
            soffice, "--headless", f"-env:UserInstallation=file://{profile.as_posix()}",
            "--convert-to", "pdf", "--outdir", str(export_dir), str(print_copy)
        ]
        completed = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                   text=True, encoding="utf-8", errors="replace", timeout=180)
        candidate = export_dir / (print_copy.stem + ".pdf")
        if completed.returncode != 0 or not candidate.is_file() or candidate.stat().st_size < 1024:
            raise ValueError("EN-1 LibreOffice PDF export failed: " + (completed.stdout or "").strip()[-1200:])
        shutil.move(str(candidate), str(pdf))
    finally:
        print_copy.unlink(missing_ok=True)
        shutil.rmtree(export_dir, ignore_errors=True)
        shutil.rmtree(profile, ignore_errors=True)

    reader = PdfReader(str(pdf))
    if len(reader.pages) != len(EN1_PRINT_SHEETS):
        raise ValueError(f"EN-1 PDF has {len(reader.pages)} pages; expected exactly {len(EN1_PRINT_SHEETS)} filing sheets")
    page_boxes = []
    for index, page in enumerate(reader.pages, start=1):
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        crop_width = float(page.cropbox.width)
        crop_height = float(page.cropbox.height)
        if min(width, height, crop_width, crop_height) < 100:
            raise ValueError(f"EN-1 PDF page {index} has an invalid/cut page box")
        if crop_width > width + 1 or crop_height > height + 1:
            raise ValueError(f"EN-1 PDF page {index} crop box exceeds its media box")
        page_boxes.append({"page": index, "widthPt": round(width, 2), "heightPt": round(height, 2),
                           "cropWidthPt": round(crop_width, 2), "cropHeightPt": round(crop_height, 2)})
    audit = {
        "schema": "liber.revex.en1-print-audit.v1",
        "status": "PASSED",
        "sourceWorkbook": xlsx.name,
        "pdf": pdf.name,
        "pageCount": len(reader.pages),
        "expectedPageCount": len(EN1_PRINT_SHEETS),
        "fitToWidth": 1,
        "fitToHeight": 1,
        "hiddenNonFilingSheetsInPrintCopy": True,
        "sourceWorkbookSheetVisibilityPreserved": True,
        "pageBoxes": page_boxes,
    }
    audit_path = output_root / "EN-1_PRINT_AUDIT.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")
    return audit


def _artifact(path: Path, output_root: Path, kind: str, *, review_name: str | None = None,
              user_visible: bool = True) -> dict:
    row = {
        "name": path.name,
        "path": path.resolve().relative_to(output_root.resolve()).as_posix(),
        "kind": kind,
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
        "userVisible": bool(user_visible),
    }
    if review_name:
        row["reviewName"] = review_name
    return row


def _regenerate_review_package(output_root: Path, public_rows: list[dict], result: dict) -> Path:
    if len(public_rows) != 9:
        raise ValueError(f"clean Energy review output has {len(public_rows)} user-visible files instead of 9")
    names = [str(row.get("reviewName") or row.get("name") or "") for row in public_rows]
    if set(names) != set(PUBLIC_REVIEW_NAMES) or len(set(names)) != 9:
        raise ValueError("clean Energy review output does not match the exact nine-file contract")
    package_path = output_root / "REVEX_ENERGY_MANUAL_REVIEW_PACKAGE.zip"
    package_path.unlink(missing_ok=True)
    index = {
        "schema": "liber.revex.energy-manual-review-package.v2",
        "projectId": result.get("projectId"),
        "sourceEngineeringRevision": result.get("sourceEngineeringRevision"),
        "status": result.get("status"),
        "entryCount": 9,
        "files": [
            {key: row.get(key) for key in ("path", "reviewName", "kind", "bytes", "sha256")}
            for row in public_rows
        ],
    }
    with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as package:
        package.comment = json.dumps(index, sort_keys=True, separators=(",", ":")).encode("utf-8")
        for row in public_rows:
            source = (output_root / str(row["path"])).resolve()
            source.relative_to(output_root.resolve())
            if not source.is_file() or source.stat().st_size != int(row["bytes"]) or _sha256(source) != row["sha256"]:
                raise ValueError(f"nine-file review artifact failed hash/byte verification: {row.get('reviewName') or row.get('name')}")
            package.write(source, str(row.get("reviewName") or source.name))
    with zipfile.ZipFile(package_path) as verification:
        entries = [name for name in verification.namelist() if not name.endswith("/")]
        if len(entries) != 9 or set(entries) != set(PUBLIC_REVIEW_NAMES):
            raise ValueError("regenerated manual-review package is not exactly the nine-file contract")
    return package_path


def finalize_complete_result(request_path: Path, result: dict, output_root: Path) -> dict:
    """Fill EN-1 people, print/QA the 16-page EN-1 PDF, and publish nine clean review files."""
    request = json.loads(Path(request_path).read_text(encoding="utf-8"))
    context = _context(request)
    applicant = _sanitize_person(context.get("en1Applicant"), modeler=False)
    modeler = _sanitize_person(context.get("en1Modeler"), modeler=True)

    filing = output_root / "05_FILING"
    xlsx = filing / "EN-1_READY_TO_INSERT.xlsx"
    if not xlsx.is_file():
        raise ValueError("completed Energy run has no EN-1_READY_TO_INSERT.xlsx")
    identity_audit = _fill_people_identity(xlsx, applicant, modeler)
    pdf = filing / "EN-1_READY_TO_INSERT.pdf"
    print_audit = _print_en1_pdf(xlsx, pdf, output_root)

    rows = [dict(row) for row in list(result.get("artifacts") or [])]
    refreshed: list[dict] = []
    for row in rows:
        path = output_root / str(row.get("path") or "")
        if path.resolve() == xlsx.resolve():
            row.update(bytes=xlsx.stat().st_size, sha256=_sha256(xlsx), userVisible=True)
        else:
            row.setdefault("userVisible", True)
        if row.get("name") in {"BASELINE_UPDATED_GEOMETRY.osm", "PROPOSED_UPDATED_GEOMETRY.osm"}:
            row["kind"] = "compiled-model"
        refreshed.append(row)
    rows = refreshed

    if any(str(row.get("reviewName") or "") == "EN-1.pdf" for row in rows):
        raise ValueError("EN-1 PDF already exists in the review contract unexpectedly")
    rows.append(_artifact(pdf, output_root, "en1-pdf", review_name="EN-1.pdf", user_visible=True))

    # These two files are integrity/filing evidence required by the worker/broker completion
    # boundary. They remain hidden from the clean nine-file review package.
    internal = (
        (filing / "COMcheck_PROJECT_INPUT_READY.cxl", "filing-input"),
        (filing / "COMcheck_BACKSTOP_RESULT.json", "engine-evidence"),
    )
    for path, kind in internal:
        if not path.is_file():
            raise ValueError(f"completed Energy run is missing hidden integrity artifact {path.name}")
        if not any(str(row.get("name") or "") == path.name for row in rows):
            rows.append(_artifact(path, output_root, kind, user_visible=False))

    public_rows = [row for row in rows if row.get("userVisible") is not False]
    package = _regenerate_review_package(output_root, public_rows, result)
    result["artifacts"] = rows
    result["en1Identity"] = {
        "schema": "liber.revex.en1-user-identity.v1",
        "authority": "explicit-user-input-during-revision-scoped-comcheck-authorization",
        **identity_audit,
        "transmittedToComcheck": False,
    }
    result["en1Pdf"] = {
        "name": pdf.name,
        "path": pdf.relative_to(output_root).as_posix(),
        "bytes": pdf.stat().st_size,
        "sha256": _sha256(pdf),
        "pageCount": print_audit["pageCount"],
        "printAudit": "EN-1_PRINT_AUDIT.json",
    }
    result["userOutput"] = {
        "schema": "liber.revex.energy-user-output.v1",
        "count": 9,
        "files": [str(row.get("reviewName") or row.get("name")) for row in public_rows],
        "hiddenIntegrityArtifacts": [path.name for path, _ in internal],
    }
    result["manualReviewPackage"] = {
        "status": "CREATED",
        "name": package.name,
        "path": package.name,
        "bytes": package.stat().st_size,
        "sha256": _sha256(package),
        "entryCount": 9,
        "topLevelFiles": 8,
        "topLevelArchives": 1,
        "referenceTemplatesIncluded": False,
        "referenceIdentityExcluded": True,
    }
    return result
