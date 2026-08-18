#!/usr/bin/env python3
"""Current WALLT EN-1 print contract.

The visual/reference contract is the previously working EN-1 export: 17 selected sheets,
Color legend first, 63% print scale, no Version Control / project-description / instructions /
unused system tabs. The source workbook is never altered; only a disposable print copy is used.
"""
from __future__ import annotations

import json
from pathlib import Path
import shutil
import subprocess
import tempfile
import uuid

VERSION = "20260818-wallt-en1-print63"
PRINT_SCALE = 63
PRINT_SHEETS = (
    "Color legend",
    "1,2,3 Information",
    "3d. LL97 GHG Summary",
    "4. Compliance",
    "5a. Baseline Rotations",
    "5b. Usage Summary",
    "6a. Ext. Wall Areas",
    "6b. Fenestration",
    "6c1. Wall Types",
    "6c2. Wall Types - Addl Rows",
    "6d.1 Interior LPD-Space Method",
    "6d.2 Interior LPD-Bldg  Method",
    "6e. Ext LPD",
    "6f. Process Equip.",
    "6g. Service HW",
    "HVAC_Cover",
    "HVAC Air-side",
)
PAGE_MARKERS = (
    "WORKBOOK COLOR LEGEND",
    "1 Location Information",
    "Carbon Emissions",
    "Energy Modeling Protocol",
    "Baseline Rotations",
    "Energy Modeling Usage Summary",
    "Above-Grade Wall & Fenestration Areas",
    "Vertical Fenestration",
    "Opaque Elements",
    "Opaque Elements - Weighted Average",
    "Interior LPD: Space-by-Space Method",
    "Interior LPD: Building Area Method",
    "Exterior Lighting",
    "Miscellaneous Equipment",
    "Service Hot Water Systems",
    "HVAC Cover Sheet",
    "Air-Side HVAC",
)


def _page_text(page) -> str:
    return " ".join((page.extract_text() or "").split())


def _spill_diagnostic(reader, source_layout: list[dict], output_root: Path) -> dict:
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        text = _page_text(page)
        matched = [
            {"sheet": sheet, "marker": marker}
            for sheet, marker in zip(PRINT_SHEETS, PAGE_MARKERS)
            if marker.lower() in text.lower()
        ]
        pages.append({
            "page": index,
            "characters": len(text),
            "matchedExpectedMarkers": matched,
            "textStart": text[:320],
            "widthPt": round(float(page.mediabox.width), 2),
            "heightPt": round(float(page.mediabox.height), 2),
        })
    diagnostic = {
        "schema": "liber.revex.en1-print-spill-diagnostic.v1",
        "version": VERSION,
        "status": "PAGE_COUNT_MISMATCH",
        "pageCount": len(reader.pages),
        "expectedPageCount": len(PRINT_SHEETS),
        "printScalePercent": PRINT_SCALE,
        "fitToPage": False,
        "selectedSheets": list(PRINT_SHEETS),
        "sourceLayout": source_layout,
        "pages": pages,
    }
    path = output_root / "EN-1_PRINT_SPILL_DIAGNOSTIC.json"
    path.write_text(json.dumps(diagnostic, ensure_ascii=True, indent=2), encoding="utf-8")
    print(json.dumps({
        "stage": "EN1_PRINT_SPILL_DIAGNOSTIC",
        "pageCount": diagnostic["pageCount"],
        "expectedPageCount": diagnostic["expectedPageCount"],
        "pages": [
            {
                "page": row["page"],
                "characters": row["characters"],
                "markers": [item["sheet"] for item in row["matchedExpectedMarkers"]],
                "textStart": row["textStart"][:120],
            }
            for row in pages
        ],
        "sourceLayout": source_layout,
    }, ensure_ascii=True, separators=(",", ":")), flush=True)
    return diagnostic


def _print_pdf(xlsx: Path, pdf: Path, output_root: Path) -> dict:
    from openpyxl import load_workbook
    from openpyxl.worksheet.properties import PageSetupProperties
    from pypdf import PdfReader

    xlsx = Path(xlsx)
    pdf = Path(pdf)
    output_root = Path(output_root)
    print_copy = xlsx.with_name("EN-1_PRINT_SOURCE.xlsx")
    shutil.copy2(xlsx, print_copy)
    workbook = load_workbook(print_copy)
    missing = [name for name in PRINT_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError("EN-1 print set is missing proven filing sheets: " + ", ".join(missing))

    source_layout = []
    for sheet in workbook.worksheets:
        if sheet.title in PRINT_SHEETS:
            source_layout.append({
                "sheet": sheet.title,
                "dimension": sheet.calculate_dimension(),
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "printAreaBefore": str(sheet.print_area or ""),
                "scaleBefore": sheet.page_setup.scale,
                "fitToWidthBefore": sheet.page_setup.fitToWidth,
                "fitToHeightBefore": sheet.page_setup.fitToHeight,
                "orientation": sheet.page_setup.orientation,
            })
            sheet.sheet_state = "visible"
            props = sheet.sheet_properties.pageSetUpPr
            if props is None:
                props = PageSetupProperties()
                sheet.sheet_properties.pageSetUpPr = props
            props.fitToPage = False
            props.autoPageBreaks = False
            sheet.page_setup.fitToWidth = None
            sheet.page_setup.fitToHeight = None
            sheet.page_setup.scale = PRINT_SCALE
        else:
            sheet.sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(PRINT_SHEETS[0])
    workbook.save(print_copy)

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise ValueError("LibreOffice Calc is unavailable for EN-1 PDF export")
    profile = Path(tempfile.gettempdir()) / f"revex-lo-{uuid.uuid4().hex}"
    export_dir = Path(tempfile.mkdtemp(prefix="revex-en1-print63-"))
    try:
        command = [
            soffice, "--headless", f"-env:UserInstallation=file://{profile.as_posix()}",
            "--convert-to", "pdf", "--outdir", str(export_dir), str(print_copy)
        ]
        completed = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                   text=True, encoding="utf-8", errors="replace", timeout=180)
        candidate = export_dir / (print_copy.stem + ".pdf")
        if completed.returncode != 0 or not candidate.is_file() or candidate.stat().st_size < 1024:
            raise ValueError("EN-1 63% LibreOffice PDF export failed: " + (completed.stdout or "").strip()[-1200:])
        shutil.move(str(candidate), str(pdf))
    finally:
        print_copy.unlink(missing_ok=True)
        shutil.rmtree(export_dir, ignore_errors=True)
        shutil.rmtree(profile, ignore_errors=True)

    reader = PdfReader(str(pdf))
    if len(reader.pages) != len(PRINT_SHEETS):
        _spill_diagnostic(reader, source_layout, output_root)
        raise ValueError(f"EN-1 PDF has {len(reader.pages)} pages; proven print contract requires {len(PRINT_SHEETS)}")

    page_boxes = []
    marker_audit = []
    for index, (page, marker) in enumerate(zip(reader.pages, PAGE_MARKERS), start=1):
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        crop_width = float(page.cropbox.width)
        crop_height = float(page.cropbox.height)
        if min(width, height, crop_width, crop_height) < 100:
            raise ValueError(f"EN-1 PDF page {index} has an invalid/cut page box")
        if crop_width > width + 1 or crop_height > height + 1:
            raise ValueError(f"EN-1 PDF page {index} crop box exceeds media box")
        text = _page_text(page)
        marker_ok = marker.lower() in text.lower()
        if not marker_ok:
            raise ValueError(f"EN-1 PDF page {index} does not match expected sheet {PRINT_SHEETS[index-1]!r}; missing marker {marker!r}")
        page_boxes.append({"page": index, "widthPt": round(width, 2), "heightPt": round(height, 2),
                           "cropWidthPt": round(crop_width, 2), "cropHeightPt": round(crop_height, 2)})
        marker_audit.append({"page": index, "sheet": PRINT_SHEETS[index-1], "marker": marker, "passed": True})

    audit = {
        "schema": "liber.revex.en1-print-audit.v2",
        "version": VERSION,
        "status": "PASSED",
        "sourceWorkbook": xlsx.name,
        "pdf": pdf.name,
        "pageCount": len(reader.pages),
        "expectedPageCount": len(PRINT_SHEETS),
        "printScalePercent": PRINT_SCALE,
        "fitToPage": False,
        "selectedSheets": list(PRINT_SHEETS),
        "hiddenNonFilingSheetsInPrintCopy": True,
        "sourceWorkbookSheetVisibilityPreserved": True,
        "referenceContract": "USER_CONFIRMED_PREVIOUSLY_WORKING_EN1_EXPORT_63_PERCENT",
        "pageMarkers": marker_audit,
        "pageBoxes": page_boxes,
    }
    audit_path = output_root / "EN-1_PRINT_AUDIT.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=True, indent=2), encoding="utf-8")
    return audit


def install(en1_module) -> None:
    if getattr(en1_module, "__revex_wallt_print63_installed__", False):
        return
    en1_module.EN1_PRINT_SHEETS = PRINT_SHEETS
    en1_module._print_en1_pdf = _print_pdf
    en1_module.__revex_wallt_print63_installed__ = True
